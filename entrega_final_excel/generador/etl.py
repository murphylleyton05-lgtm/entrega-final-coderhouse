"""
Hojas `Datos_Origen`, `Datos_Limpios` y `Control_Calidad`.

`Datos_Origen` es el export crudo, tal cual sale del sistema transaccional:
duplicados, fechas en tres formatos, textos sin estandarizar, unidades
cargadas como texto y ciudades vacias.

`Datos_Limpios` no es una copia pegada: cada celda es una formula que
reconstruye el valor correcto desde el origen. Asi el ETL queda a la vista,
es auditable celda por celda y se rehace solo si el export cambia.

`Control_Calidad` cuenta los defectos del origen y verifica, con formulas,
que despues del ETL no quede ninguno.
"""
from openpyxl.formatting.rule import FormulaRule
from openpyxl.styles import Font, PatternFill
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.table import Table, TableStyleInfo

import datos
from estilos import (AMBER, AMBER_BG, FECHA, FUENTE, GRAY, INK, LIGHT, LINE,
                     MONEY, MONEY2, NAVY, NUM, PCT1, WHITE, banner,
                     encabezado_tabla, f, imprimible)

ORIGEN = "Datos_Origen"
LIMPIOS = "Datos_Limpios"
CONTROL = "Control_Calidad"

# columnas del export crudo (A..J) + una auxiliar de control (K)
CO = {c: get_column_letter(i + 1) for i, c in enumerate(datos.COLUMNAS_CRUDAS)}
CO_VENTAS = get_column_letter(len(datos.COLUMNAS_CRUDAS) + 1)

# columnas de la tabla limpia
CL = {c: get_column_letter(i + 1) for i, c in enumerate(datos.COLUMNAS)}
N_LIMPIAS = len(datos.COLUMNAS)
C_POS = get_column_letter(N_LIMPIAS + 1)       # posicion en el export
C_CTRL = get_column_letter(N_LIMPIAS + 2)      # control de deduplicacion
C_FECHA_ORI = get_column_letter(N_LIMPIAS + 3)  # traza: fecha como se recibio
C_UNID_ORI = get_column_letter(N_LIMPIAS + 4)   # traza: unidades como se recibieron
C_CIUDAD_ORI = get_column_letter(N_LIMPIAS + 5)  # traza: ciudad como se recibio
N_TOTAL = N_LIMPIAS + 5


# =============================================================================
#  Datos_Origen
# =============================================================================
def hoja_origen(wb, columnas, filas):
    ws = wb.create_sheet(ORIGEN)
    ultima = len(filas) + 1

    for i, c in enumerate(columnas, start=1):
        f(ws, f"{get_column_letter(i)}1", c, bold=True, size=9.5, color=WHITE,
          fill=NAVY, align="center")
    f(ws, f"{CO_VENTAS}1", "Ventas_Origen", bold=True, size=9.5, color=WHITE,
      fill=GRAY, align="center")

    for r, fila in enumerate(filas, start=2):
        for c, v in enumerate(fila, start=1):
            ws.cell(r, c, v)
        # auxiliar de control: cuanto factura el export TAL COMO LLEGA, con
        # duplicados incluidos. Sirve para medir cuanta facturacion inflaban.
        ws.cell(r, len(columnas) + 1,
                f"=IF(ISTEXT({CO['Unidades']}{r}),VALUE({CO['Unidades']}{r}),"
                f"{CO['Unidades']}{r})*{CO['Precio_Unitario']}{r}")

    for r in range(2, ultima + 1):
        ws.cell(r, len(columnas) + 1).number_format = MONEY2
        for col in ("Precio_Unitario", "Costo_Unitario"):
            ws[f"{CO[col]}{r}"].number_format = MONEY2
        c = ws[f"{CO['Fecha']}{r}"]
        if not isinstance(c.value, str):
            c.number_format = FECHA

    anchos = {"ID_Venta": 10, "Fecha": 13, "Region": 13, "Ciudad": 15,
              "Canal": 16, "Categoria": 17, "Producto": 21, "Unidades": 11,
              "Precio_Unitario": 15, "Costo_Unitario": 15}
    for nombre, w in anchos.items():
        ws.column_dimensions[CO[nombre]].width = w
    ws.column_dimensions[CO_VENTAS].width = 15

    # --- los problemas se marcan solos ------------------------------------
    rango = f"A2:{CO_VENTAS}{ultima}"
    amber = PatternFill(start_color=AMBER_BG, end_color=AMBER_BG,
                        fill_type="solid")
    fuente_amber = Font(name=FUENTE, sz=9.5, color="A96A12")
    reglas = [
        (f"{CO['Fecha']}2:{CO['Fecha']}{ultima}",
         f"ISTEXT({CO['Fecha']}2)"),
        (f"{CO['Unidades']}2:{CO['Unidades']}{ultima}",
         f"ISTEXT({CO['Unidades']}2)"),
        (f"{CO['Ciudad']}2:{CO['Ciudad']}{ultima}",
         f'{CO["Ciudad"]}2=""'),
        (f"{CO['ID_Venta']}2:{CO['ID_Venta']}{ultima}",
         f"COUNTIF(${CO['ID_Venta']}$2:${CO['ID_Venta']}${ultima},"
         f"{CO['ID_Venta']}2)>1"),
    ]
    for destino, formula in reglas:
        ws.conditional_formatting.add(
            destino, FormulaRule(formula=[formula], fill=amber,
                                 font=fuente_amber))

    ws.freeze_panes = "A2"
    ws.sheet_view.zoomScale = 90
    imprimible(ws, f"A1:{CO_VENTAS}{ultima}")
    return ws, ultima


def leyenda_origen(ws, ultima):
    """Cartel flotante sobre la hoja cruda, en las columnas libres."""
    col = get_column_letter(len(datos.COLUMNAS_CRUDAS) + 3)   # M
    col_fin = get_column_letter(len(datos.COLUMNAS_CRUDAS) + 8)
    ws.column_dimensions[col].width = 3
    for i in range(len(datos.COLUMNAS_CRUDAS) + 4,
                   len(datos.COLUMNAS_CRUDAS) + 9):
        ws.column_dimensions[get_column_letter(i)].width = 17
    inicio = get_column_letter(len(datos.COLUMNAS_CRUDAS) + 4)

    banner(ws, 2, len(datos.COLUMNAS_CRUDAS) + 4,
           len(datos.COLUMNAS_CRUDAS) + 8,
           "EXPORT CRUDO  ·  NO EDITAR  ·  entrada del proceso")
    textos = [
        ("Que es esta hoja",
         "El archivo tal como lo entrega el sistema transaccional. Se deja "
         "intacto a proposito: es la evidencia de que la limpieza se hizo y "
         "permite rehacerla si cambia una regla."),
        ("Las celdas ambar",
         "Estan resaltadas por formato condicional. Marcan los cuatro "
         "defectos: ID repetido, fecha cargada como texto, unidades cargadas "
         "como texto y ciudad vacia."),
        ("Los textos sin estandarizar",
         "Region, Canal y Categoria llegan con mayusculas, minusculas, "
         "acentos y espacios mezclados. El diccionario de la hoja Parametros "
         "los unifica."),
        ("Que pasa despues",
         "La hoja Datos_Limpios reconstruye cada fila con formulas y la hoja "
         "Control_Calidad verifica que no quede ningun defecto."),
    ]
    r = 4
    for titulo, texto in textos:
        f(ws, f"{inicio}{r}", titulo, bold=True, size=9.5, color=NAVY,
          align="left")
        ws.merge_cells(f"{inicio}{r}:{col_fin}{r}")
        r += 1
        f(ws, f"{inicio}{r}", texto, size=9, color=INK, wrap=True,
          align="left", vertical="top")
        ws.merge_cells(f"{inicio}{r}:{col_fin}{r + 2}")
        r += 4


# =============================================================================
#  Datos_Limpios
# =============================================================================
def hoja_limpios(wb, P, posiciones, n_crudas):
    """
    P          : rangos de la hoja Parametros
    posiciones : para cada operacion unica, su posicion (1..n) dentro del
                 export crudo. Es el resultado del paso "quitar duplicados":
                 se conserva la primera aparicion de cada ID_Venta.
    """
    ws = wb.create_sheet(LIMPIOS)
    n = len(posiciones)
    ultima = n + 1
    fin_crudo = n_crudas + 1   # ultima fila con datos en Datos_Origen

    def ori(campo):
        """Rango del export crudo para una columna, ya absoluto."""
        return f"{ORIGEN}!${CO[campo]}$2:${CO[campo]}${fin_crudo}"

    def trae(campo, fila):
        return f"INDEX({ori(campo)},${C_POS}{fila})"

    def par(rango):
        return f"Parametros!{rango}"

    encabezados = list(datos.COLUMNAS) + [
        "Pos_Origen", "Control_ETL", "Fecha_Origen", "Unidades_Origen",
        "Ciudad_Origen"]
    for i, c in enumerate(encabezados, start=1):
        color = NAVY if i <= N_LIMPIAS else GRAY
        f(ws, f"{get_column_letter(i)}1", c, bold=True, size=9.5, color=WHITE,
          fill=color, align="center")

    for k, pos in enumerate(posiciones):
        r = k + 2
        # --- traza del origen (las tres que se usan mas de una vez) --------
        ws[f"{C_POS}{r}"] = pos
        ws[f"{C_FECHA_ORI}{r}"] = f"={trae('Fecha', r)}"
        ws[f"{C_UNID_ORI}{r}"] = f"={trae('Unidades', r)}"
        # el &"" es necesario: sin el, una celda vacia del origen llega como 0
        # y TRIM(0) devuelve "0", que no es vacio y rompe el reemplazo
        ws[f"{C_CIUDAD_ORI}{r}"] = f'={trae("Ciudad", r)}&""'
        # el control prueba que Pos_Origen es de verdad la PRIMERA aparicion
        ws[f"{C_CTRL}{r}"] = (
            f"=IF(MATCH(${CL['ID_Venta']}{r},{ori('ID_Venta')},0)"
            f"=${C_POS}{r},\"OK\",\"REVISAR\")")

        # --- identificacion ------------------------------------------------
        ws[f"{CL['ID_Venta']}{r}"] = f"={trae('ID_Venta', r)}"

        # --- fecha: tres formatos de entrada, uno de salida -----------------
        x = f"${C_FECHA_ORI}{r}"
        ws[f"{CL['Fecha']}{r}"] = (
            f"=IF(ISNUMBER({x}),{x},"
            f'IF(MID({x},5,1)="-",'
            f"DATE(VALUE(LEFT({x},4)),VALUE(MID({x},6,2)),VALUE(RIGHT({x},2))),"
            f"DATE(VALUE(RIGHT({x},4)),VALUE(MID({x},4,2)),VALUE(LEFT({x},2)))))")
        fe = f"${CL['Fecha']}{r}"
        ws[f"{CL['Anio']}{r}"] = f"=YEAR({fe})"
        ws[f"{CL['Trimestre']}{r}"] = f'="Q"&ROUNDUP(MONTH({fe})/3,0)'
        ws[f"{CL['Mes_Num']}{r}"] = f"=MONTH({fe})"
        ws[f"{CL['Mes']}{r}"] = (
            f"=INDEX({par(P['meses'])},${CL['Mes_Num']}{r})")
        ws[f"{CL['Periodo']}{r}"] = (
            f'=TEXT(${CL["Anio"]}{r},"0000")&"-"&'
            f'TEXT(${CL["Mes_Num"]}{r},"00")')

        # --- textos: se estandarizan contra el diccionario ------------------
        ws[f"{CL['Region']}{r}"] = (
            f"=INDEX({par(P['dic_reg_estandar'])},"
            f"MATCH(TRIM({trae('Region', r)}),{par(P['dic_reg_recibido'])},0))")
        ws[f"{CL['Canal']}{r}"] = (
            f"=INDEX({par(P['dic_can_estandar'])},"
            f"MATCH(TRIM({trae('Canal', r)}),{par(P['dic_can_recibido'])},0))")
        ws[f"{CL['Categoria']}{r}"] = (
            f"=INDEX({par(P['dic_cat_estandar'])},"
            f"MATCH(TRIM({trae('Categoria', r)}),"
            f"{par(P['dic_cat_recibido'])},0))")
        ws[f"{CL['Producto']}{r}"] = f"=TRIM({trae('Producto', r)})"
        # ciudad vacia -> la de referencia de la region
        ws[f"{CL['Ciudad']}{r}"] = (
            f'=IF(TRIM(${C_CIUDAD_ORI}{r})="",'
            f"INDEX({par(P['reg_ciudad'])},"
            f"MATCH(${CL['Region']}{r},{par(P['reg_region'])},0)),"
            f"TRIM(${C_CIUDAD_ORI}{r}))")
        ws[f"{CL['Cat_Anio']}{r}"] = (
            f'=${CL["Categoria"]}{r}&" "&${CL["Anio"]}{r}')
        ws[f"{CL['Prod_Anio']}{r}"] = (
            f'=${CL["Producto"]}{r}&" "&${CL["Anio"]}{r}')

        # --- numeros: los que llegaron como texto se convierten -------------
        u = f"${C_UNID_ORI}{r}"
        ws[f"{CL['Unidades']}{r}"] = f"=IF(ISTEXT({u}),VALUE({u}),{u})"
        ws[f"{CL['Precio_Unitario']}{r}"] = f"={trae('Precio_Unitario', r)}"
        ws[f"{CL['Costo_Unitario']}{r}"] = f"={trae('Costo_Unitario', r)}"

        # --- medidas: se calculan en la planilla, no se pegan ---------------
        cu, cp, cc = (f"${CL['Unidades']}{r}", f"${CL['Precio_Unitario']}{r}",
                      f"${CL['Costo_Unitario']}{r}")
        ws[f"{CL['Ventas']}{r}"] = f"=ROUND({cu}*{cp},2)"
        ws[f"{CL['Costo_Producto']}{r}"] = f"=ROUND({cu}*{cc},2)"
        cv = f"${CL['Ventas']}{r}"
        ws[f"{CL['Costo_Logistico']}{r}"] = (
            f"=ROUND({cv}*SUMIFS({par(P['tasa_valor'])},"
            f"{par(P['tasa_canal'])},${CL['Canal']}{r},"
            f"{par(P['tasa_anio'])},${CL['Anio']}{r},"
            f"{par(P['tasa_trim'])},${CL['Trimestre']}{r}),2)")
        ws[f"{CL['Margen_Bruto']}{r}"] = (
            f"={cv}-${CL['Costo_Producto']}{r}-${CL['Costo_Logistico']}{r}")
        cm = f"${CL['Margen_Bruto']}{r}"
        ws[f"{CL['Margen_Pct']}{r}"] = f"=IFERROR({cm}/{cv},0)"

        # --- columnas calculadas con logica condicional (SI / Y / O) --------
        cmp_ = f"${CL['Margen_Pct']}{r}"
        ws[f"{CL['Segmento_Rentabilidad']}{r}"] = (
            f"=IF(AND({cmp_}>={par(P['meta_margen'])},"
            f"{cv}>={par(P['ticket_alto'])}),\"Alta rentabilidad\","
            f"IF(OR({cmp_}<{par(P['piso_margen'])},"
            f"{cv}<{par(P['ticket_bajo'])}),\"Revisar\",\"Estandar\"))")
        ws[f"{CL['Alerta_Logistica']}{r}"] = (
            f'=IF(AND(${CL["Canal"]}{r}="E-commerce",'
            f"IFERROR(${CL['Costo_Logistico']}{r}/{cv},0)>"
            f"{par(P['techo_log_ecom'])}),\"Alerta\",\"OK\")")
        ws[f"{CL['Ticket']}{r}"] = (
            f"=IF({cv}>={par(P['ticket_alto'])},\"Alto\","
            f"IF({cv}<{par(P['ticket_bajo'])},\"Bajo\",\"Medio\"))")

    # --- formatos -----------------------------------------------------------
    formatos = {"Fecha": FECHA, "Unidades": NUM, "Precio_Unitario": MONEY2,
                "Costo_Unitario": MONEY2, "Ventas": MONEY,
                "Costo_Producto": MONEY, "Costo_Logistico": MONEY,
                "Margen_Bruto": MONEY, "Margen_Pct": PCT1}
    for r in range(2, ultima + 1):
        for campo, fmt in formatos.items():
            ws[f"{CL[campo]}{r}"].number_format = fmt
        ws[f"{C_FECHA_ORI}{r}"].number_format = "General"
        for col in range(1, N_TOTAL + 1):
            ws.cell(r, col).font = Font(name=FUENTE, sz=9.5, color=INK)

    anchos = {"ID_Venta": 10, "Fecha": 11, "Anio": 7, "Trimestre": 10,
              "Mes_Num": 9, "Mes": 7, "Periodo": 10, "Region": 11,
              "Ciudad": 14, "Canal": 14, "Categoria": 15, "Cat_Anio": 19,
              "Producto": 20, "Prod_Anio": 24, "Unidades": 10,
              "Precio_Unitario": 15, "Costo_Unitario": 15, "Ventas": 13,
              "Costo_Producto": 15, "Costo_Logistico": 15, "Margen_Bruto": 14,
              "Margen_Pct": 11, "Segmento_Rentabilidad": 20,
              "Alerta_Logistica": 16, "Ticket": 9}
    for nombre, w in anchos.items():
        ws.column_dimensions[CL[nombre]].width = w
    for col, w in ((C_POS, 11), (C_CTRL, 12), (C_FECHA_ORI, 13),
                   (C_UNID_ORI, 15), (C_CIUDAD_ORI, 14)):
        ws.column_dimensions[col].width = w

    # tabla de Excel: da nombre al rango y encabezados fijos al filtrar
    tabla = Table(displayName="tbl_Ventas",
                  ref=f"A1:{get_column_letter(N_TOTAL)}{ultima}")
    tabla.tableStyleInfo = TableStyleInfo(name="TableStyleLight1",
                                          showRowStripes=True)
    ws.add_table(tabla)
    ws.freeze_panes = "B2"
    ws.sheet_view.zoomScale = 90

    # --- las columnas logicas se leen de un vistazo -------------------------
    rango_seg = f"{CL['Segmento_Rentabilidad']}2:{CL['Segmento_Rentabilidad']}{ultima}"
    ws.conditional_formatting.add(
        rango_seg,
        FormulaRule(formula=[f'{CL["Segmento_Rentabilidad"]}2="Revisar"'],
                    fill=PatternFill(start_color=AMBER_BG, end_color=AMBER_BG,
                                     fill_type="solid"),
                    font=Font(name=FUENTE, sz=9.5, color="A96A12", b=True)))
    rango_ale = f"{CL['Alerta_Logistica']}2:{CL['Alerta_Logistica']}{ultima}"
    ws.conditional_formatting.add(
        rango_ale,
        FormulaRule(formula=[f'{CL["Alerta_Logistica"]}2="Alerta"'],
                    fill=PatternFill(start_color=AMBER_BG, end_color=AMBER_BG,
                                     fill_type="solid"),
                    font=Font(name=FUENTE, sz=9.5, color="A96A12", b=True)))
    return ws, ultima


# =============================================================================
#  Control_Calidad
# =============================================================================
def hoja_control(wb, n_crudas, n_limpias, informe):
    ws = wb.create_sheet(CONTROL)
    ws.sheet_view.showGridLines = False
    ws.column_dimensions["A"].width = 2
    ws.column_dimensions["B"].width = 34
    ws.column_dimensions["C"].width = 46
    for c in "DE":
        ws.column_dimensions[c].width = 15
    ws.column_dimensions["F"].width = 13
    ws.column_dimensions["G"].width = 3

    fin_c = n_crudas + 1
    fin_l = n_limpias + 1

    def O(campo):
        return f"{ORIGEN}!${CO[campo]}$2:${CO[campo]}${fin_c}"

    def L(campo):
        return f"{LIMPIOS}!${CL[campo]}$2:${CL[campo]}${fin_l}"

    def distintos(rango):
        """
        Cuenta valores distintos sin funciones modernas.

        El &"" y el factor (rango<>"") son la version a prueba de celdas
        vacias: sin ellos, un blanco hace COUNTIF=0 y la division revienta.
        """
        return (f'SUMPRODUCT(({rango}<>"")/COUNTIF({rango},{rango}&""))')

    f(ws, "B2", "Control de calidad del dato", bold=True, size=15, color=NAVY)
    f(ws, "B3", "Cada fila mide un defecto del export crudo y verifica que "
                "despues del ETL ya no exista. Todos los numeros son formulas "
                "sobre las hojas Datos_Origen y Datos_Limpios: si el origen "
                "cambia, este tablero se actualiza solo. Nota: CONTAR.SI no "
                "distingue mayusculas de minusculas, asi que las cuentas de "
                "valores distintos son un piso; el diccionario de Parametros "
                "publica todas las variantes recibidas, una por fila.",
      size=9, color=GRAY, italic=True, wrap=True, vertical="top")
    ws.merge_cells("B3:F4")
    ws.row_dimensions[3].height = 24

    banner(ws, 6, 2, 6, "DIAGNOSTICO Y REMEDIACION")
    encabezado_tabla(ws, 7, 2, 6, ["Problema de calidad", "Regla aplicada",
                                   "En el origen", "Despues del ETL",
                                   "Estado"])

    filas = [
        ("Registros duplicados",
         "Se conserva la primera aparicion de cada ID_Venta y se descartan los "
         "reenvios del sistema.",
         f"={n_crudas}-{distintos(O('ID_Venta'))}",
         f"={fin_l - 1}-{distintos(L('ID_Venta'))}", NUM),
        ("Fechas almacenadas como texto",
         "Se detecta el formato (dd/mm/aaaa o aaaa-mm-dd) y se reconstruye la "
         "fecha real con FECHA/DATE.",
         f"=SUMPRODUCT(--ISTEXT({O('Fecha')}))",
         f"=SUMPRODUCT(--ISTEXT({L('Fecha')}))", NUM),
        ("Unidades cargadas como texto",
         "Se convierten a numero con VALOR/VALUE solo cuando la celda es "
         "texto.",
         f"=SUMPRODUCT(--ISTEXT({O('Unidades')}))",
         f"=SUMPRODUCT(--ISTEXT({L('Unidades')}))", NUM),
        ("Ciudades vacias",
         "Se completan con la ciudad de referencia de la region "
         "(INDICE/COINCIDIR contra Parametros).",
         f"=COUNTBLANK({O('Ciudad')})",
         f"=COUNTBLANK({L('Ciudad')})", NUM),
        ("Valores distintos en Categoria",
         "Diccionario de normalizacion: cada variante recibida se mapea a un "
         "unico valor estandar. Deben quedar tantos valores como categorias "
         "reales.",
         f"={distintos(O('Categoria'))}",
         f"={distintos(L('Categoria'))}", NUM),
        ("Valores distintos en Region",
         "Mismo diccionario, dimension Region.",
         f"={distintos(O('Region'))}",
         f"={distintos(L('Region'))}", NUM),
        ("Valores distintos en Canal",
         "Mismo diccionario, dimension Canal.",
         f"={distintos(O('Canal'))}",
         f"={distintos(L('Canal'))}", NUM),
        ("Valores distintos en Producto",
         "ESPACIOS/TRIM sobre el nombre recibido.",
         f"={distintos(O('Producto'))}",
         f"={distintos(L('Producto'))}", NUM),
    ]
    # el estado esperado de cada fila: 0 defectos, o el N de valores estandar
    esperado = [0, 0, 0, 0, len(datos.CATEGORIAS), len(datos.REGIONES),
                len(datos.CANALES), len(datos.PRODUCTOS)]
    r = 8
    for (nombre, regla, en_origen, post, fmt), esp in zip(filas, esperado):
        f(ws, f"B{r}", nombre, size=9.5, bold=True, color=NAVY, align="left")
        f(ws, f"C{r}", regla, size=9, color=INK, wrap=True, align="left",
          vertical="center")
        f(ws, f"D{r}", en_origen, size=9.5, fmt=fmt, align="center",
          color=AMBER, bold=True)
        f(ws, f"E{r}", post, size=9.5, fmt=fmt, align="center", color=NAVY,
          bold=True)
        f(ws, f"F{r}", f'=IF(E{r}={esp},"OK","REVISAR")', size=9.5,
          align="center", bold=True)
        ws.row_dimensions[r].height = 30
        r += 1

    # --- reconciliacion de volumenes e importes ----------------------------
    r += 1
    banner(ws, r, 2, 6, "RECONCILIACION")
    r += 1
    encabezado_tabla(ws, r, 2, 6, ["Concepto", "Detalle", "Valor", "", ""])
    r += 1
    recon = [
        ("Filas recibidas en el export",
         "Todo lo que entrego el sistema, reenvios incluidos.",
         f"=COUNTA({O('ID_Venta')})", NUM),
        ("Operaciones unicas",
         "Filas distintas por ID_Venta: lo que realmente ocurrio.",
         f"={distintos(O('ID_Venta'))}", NUM),
        ("Filas descartadas por duplicado",
         "Diferencia entre las dos anteriores.",
         None, NUM),
        ("Facturacion del export crudo",
         "Unidades x precio sobre TODAS las filas recibidas.",
         f"=SUM({ORIGEN}!${CO_VENTAS}$2:${CO_VENTAS}${fin_c})", MONEY),
        ("Facturacion depurada",
         "La misma cuenta sobre la tabla limpia.",
         f"=SUM({L('Ventas')})", MONEY),
        ("Facturacion que inflaban los duplicados",
         "Lo que se habria reportado de mas sin depurar.",
         None, MONEY),
        ("Filas de la tabla limpia",
         "Debe coincidir con las operaciones unicas.",
         f"=COUNTA({L('ID_Venta')})", NUM),
        ("Control de deduplicacion",
         "Filas donde Pos_Origen NO apunta a la primera aparicion del "
         "ID_Venta. Tiene que ser cero.",
         f'=COUNTIF({LIMPIOS}!${C_CTRL}$2:${C_CTRL}${fin_l},"REVISAR")', NUM),
    ]
    inicio_recon = r
    for i, (nombre, detalle, formula, fmt) in enumerate(recon):
        f(ws, f"B{r}", nombre, size=9.5, bold=True, color=NAVY, align="left")
        f(ws, f"C{r}", detalle, size=9, color=INK, wrap=True, align="left")
        if formula is None:
            # se resuelve contra las dos filas de arriba
            formula = f"=D{r - 2}-D{r - 1}"
        f(ws, f"D{r}", formula, size=9.5, fmt=fmt, align="center", bold=True,
          color=NAVY)
        ws.row_dimensions[r].height = 26
        r += 1

    # el control de deduplicacion tiene semaforo propio
    f(ws, f"E{r - 1}", f'=IF(D{r - 1}=0,"OK","REVISAR")', size=9.5,
      align="center", bold=True)

    # --- lo que el ETL agrega ----------------------------------------------
    r += 1
    banner(ws, r, 2, 6, "COLUMNAS QUE AGREGA EL ETL")
    r += 1
    encabezado_tabla(ws, r, 2, 6,
                     ["Columna calculada", "Regla", "Resultado", "", ""])
    r += 1
    seg = L("Segmento_Rentabilidad")
    logicas = [
        ("Segmento_Rentabilidad",
         'SI(Y(margen% >= meta; ventas >= ticket alto); "Alta rentabilidad"; '
         'SI(O(margen% < piso; ventas < ticket bajo); "Revisar"; "Estandar"))',
         [("Alta rentabilidad", f'=COUNTIF({seg},"Alta rentabilidad")'),
          ("Estandar", f'=COUNTIF({seg},"Estandar")'),
          ("Revisar", f'=COUNTIF({seg},"Revisar")')]),
        ("Alerta_Logistica",
         'SI(Y(canal = "E-commerce"; costo logistico / ventas > techo); '
         '"Alerta"; "OK")',
         [("Alerta", f'=COUNTIF({L("Alerta_Logistica")},"Alerta")'),
          ("OK", f'=COUNTIF({L("Alerta_Logistica")},"OK")')]),
        ("Ticket",
         'SI(ventas >= ticket alto; "Alto"; SI(ventas < ticket bajo; "Bajo"; '
         '"Medio"))',
         [("Alto", f'=COUNTIF({L("Ticket")},"Alto")'),
          ("Medio", f'=COUNTIF({L("Ticket")},"Medio")'),
          ("Bajo", f'=COUNTIF({L("Ticket")},"Bajo")')]),
    ]
    for nombre, regla, conteos in logicas:
        f(ws, f"B{r}", nombre, size=9.5, bold=True, color=NAVY, align="left")
        f(ws, f"C{r}", regla, size=9, color=INK, wrap=True, align="left",
          vertical="top")
        ws.merge_cells(f"C{r}:C{r + len(conteos) - 1}")
        for j, (etiqueta, formula) in enumerate(conteos):
            f(ws, f"D{r + j}", etiqueta, size=9, color=GRAY, align="right")
            f(ws, f"E{r + j}", formula, size=9.5, fmt=NUM, align="center",
              bold=True, color=NAVY)
            ws.row_dimensions[r + j].height = 15
        ws.row_dimensions[r].height = 22
        r += len(conteos) + 1

    f(ws, f"B{r + 1}",
      "Las tres columnas anteriores se calculan con logica condicional "
      "(SI / Y / O) sobre los umbrales de la hoja Parametros. Cambiar un "
      "umbral reclasifica las 2.400 operaciones sin tocar una sola formula.",
      size=9, color=GRAY, italic=True, wrap=True, align="left")
    ws.merge_cells(f"B{r + 1}:F{r + 2}")

    # semaforo: verde institucional si OK, ambar si hay algo que revisar
    verde = Font(name=FUENTE, sz=9.5, b=True, color="1E7B4F")
    ambar = Font(name=FUENTE, sz=9.5, b=True, color="A96A12")
    imprimible(ws, f"B2:F{r + 2}")
    for rango in (f"F8:F{7 + len(filas)}", f"E{inicio_recon + 7}"):
        ws.conditional_formatting.add(
            rango, FormulaRule(formula=[f'{rango.split(":")[0]}="OK"'],
                               font=verde,
                               fill=PatternFill(start_color="E8F3ED",
                                                end_color="E8F3ED",
                                                fill_type="solid")))
        ws.conditional_formatting.add(
            rango, FormulaRule(formula=[f'{rango.split(":")[0]}="REVISAR"'],
                               font=ambar,
                               fill=PatternFill(start_color=AMBER_BG,
                                                end_color=AMBER_BG,
                                                fill_type="solid")))
    return ws
