"""
Hoja `Parametros`: la unica fuente de constantes del libro.

Ningun numero fijo se escribe dos veces. Metas, umbrales, tasas y los
diccionarios de normalizacion de texto viven aca, y el resto del libro los
referencia por celda. La funcion devuelve un diccionario de rangos para que
las formulas se armen contra posiciones reales y no contra literales copiados
a mano.
"""
from openpyxl.styles import Alignment
from openpyxl.worksheet.datavalidation import DataValidation

import datos
from estilos import (AMBER, GRAY, LIGHT, MONEY, NAVY, NUM, PCT1, PCT2, WHITE,
                     banner, encabezado_tabla, f, imprimible)

HOJA = "Parametros"


def construir(wb):
    ws = wb.create_sheet(HOJA)
    ws.sheet_view.showGridLines = False
    ws.column_dimensions["A"].width = 2
    ws.column_dimensions["B"].width = 40
    for c in "CDE":
        ws.column_dimensions[c].width = 16
    ws.column_dimensions["F"].width = 58
    ws.column_dimensions["G"].width = 3
    ws.column_dimensions["H"].width = 8
    ws.column_dimensions["I"].width = 22

    f(ws, "B2", "Parametros, supuestos, metas y diccionarios", bold=True,
      size=15, color=NAVY)
    f(ws, "B3", "Toda constante del modelo vive en esta hoja. El ETL, los KPIs "
                "y las columnas logicas la referencian por celda; ninguna la "
                "repite escrita a mano.", size=9, color=GRAY, italic=True)

    R = {}

    # --- metas y umbrales ---------------------------------------------------
    banner(ws, 5, 2, 6, "METAS Y UMBRALES DEL EJERCICIO 2025")
    encabezado_tabla(ws, 6, 2, 6,
                     ["Indicador", "Valor", "Se usa en", "", "Origen del supuesto"])
    metas = [
        ("Ventas totales 2025 (meta)", 5_400_000, MONEY, "KPI 1",
         "Presupuesto comercial aprobado para el ejercicio."),
        ("Margen bruto % (meta)", 0.34, PCT1, "KPI 2 / columna logica",
         "Piso historico de la cadena: 2024 cerro en 35,8%."),
        ("Crecimiento interanual de ventas (meta)", 0.15, PCT1, "KPI 3",
         "Objetivo de expansion fijado por la direccion."),
        ("Costo logistico sobre ventas (techo)", 0.025, PCT2, "KPI 4",
         "Techo tolerado a nivel compania; 2024 cerro en 2,18%."),
        ("Margen bruto % piso por operacion", datos.PISO_MARGEN, PCT1,
         "Columna Segmento_Rentabilidad",
         "Por debajo de este margen la operacion entra en revision comercial."),
        ("Ticket alto (USD)", datos.TICKET_ALTO, MONEY,
         "Columnas Segmento_Rentabilidad y Ticket",
         "Corte de venta grande: por encima requiere aprobacion de gerencia."),
        ("Ticket bajo (USD)", datos.TICKET_BAJO, MONEY,
         "Columnas Segmento_Rentabilidad y Ticket",
         "Por debajo el costo de atencion se come el margen."),
        ("Costo logistico / ventas tolerado en e-commerce",
         datos.TECHO_LOG_ECOM, PCT2, "Columna Alerta_Logistica",
         "Umbral operativo de la ultima milla en el canal online."),
        ("Margen bruto % minimo por producto", 0.33, PCT1,
         "Grafico de dispersion",
         "Corte que separa la zona de riesgo en el mapa de productos."),
    ]
    for i, (nombre, valor, fmt, uso, origen) in enumerate(metas):
        r = 7 + i
        f(ws, f"B{r}", nombre, size=9.5)
        f(ws, f"C{r}", valor, bold=True, color=NAVY, fmt=fmt, align="center")
        f(ws, f"D{r}", uso, size=8.5, color=GRAY, align="left")
        ws.merge_cells(f"D{r}:E{r}")
        f(ws, f"F{r}", origen, size=8.5, color=GRAY, wrap=True)
    R.update(meta_ventas="$C$7", meta_margen="$C$8", meta_yoy="$C$9",
             techo_log="$C$10", piso_margen="$C$11", ticket_alto="$C$12",
             ticket_bajo="$C$13", techo_log_ecom="$C$14", corte_prod="$C$15")

    # --- supuestos del dataset ----------------------------------------------
    banner(ws, 17, 2, 6, "SUPUESTOS DEL DATASET")
    supuestos = [
        ("Periodo cubierto", "Ene-2024 a Dic-2025", None),
        ("Operaciones unicas (tabla de hechos)",
         len(datos.PRODUCTOS) * len(datos.REGIONES) * 24 * len(datos.CANALES), NUM),
        ("Aumento de precio de lista 2025", datos.INFLACION_PRECIO, PCT1),
        ("Inflacion de costo de compra 2025", datos.INFLACION_COSTO, PCT1),
        ("Crecimiento de unidades 2025", datos.CRECIMIENTO_UNIDADES, PCT1),
        ("Participacion de e-commerce 2024", datos.MIX_ECOM[2024], PCT1),
        ("Participacion de e-commerce 2025", datos.MIX_ECOM[2025], PCT1),
    ]
    for i, (k, v, fmt) in enumerate(supuestos):
        r = 18 + i
        f(ws, f"B{r}", k, size=9.5)
        f(ws, f"C{r}", v, color=NAVY, align="center", fmt=fmt)
    f(ws, "F18", "Dataset SIMULADO y determinista (semilla fija). Extiende el "
                 "modelo Ventas_Tech_DB del checkpoint de SQL -mismas "
                 "categorias, productos y ciudades- a 24 meses, 5 regiones y "
                 "2 canales, para poder analizar margen, estacionalidad y "
                 "evolucion interanual.", size=8.5, color=GRAY, wrap=True,
      vertical="top")
    ws.merge_cells("F18:F24")

    # --- tasa logistica -----------------------------------------------------
    banner(ws, 26, 2, 6,
           "COSTO LOGISTICO SOBRE VENTAS  ·  tasa que aplica el ETL")
    encabezado_tabla(ws, 27, 2, 5, ["Canal", "Anio", "Trimestre", "Tasa"])
    r = 28
    inicio_tasa = r
    for (canal, anio), tasas in datos.TASA_LOGISTICA.items():
        for q, tasa in tasas.items():
            f(ws, f"B{r}", canal, size=9.5)
            f(ws, f"C{r}", anio, align="center", size=9.5)
            f(ws, f"D{r}", f"Q{q}", align="center", size=9.5)
            f(ws, f"E{r}", tasa, fmt=PCT2, align="center", color=NAVY, size=9.5)
            r += 1
    fin_tasa = r - 1
    R.update(tasa_canal=f"$B${inicio_tasa}:$B${fin_tasa}",
             tasa_anio=f"$C${inicio_tasa}:$C${fin_tasa}",
             tasa_trim=f"$D${inicio_tasa}:$D${fin_tasa}",
             tasa_valor=f"$E${inicio_tasa}:$E${fin_tasa}")
    f(ws, f"F{inicio_tasa}",
      "El salto de e-commerce en 2025 (ultima milla y envio bonificado) y su "
      "baja en Q3-Q4 al fijar un umbral de envio minimo es el hecho que "
      "explica la caida y posterior recuperacion del margen.",
      size=8.5, color=GRAY, wrap=True, vertical="top")
    ws.merge_cells(f"F{inicio_tasa}:F{fin_tasa}")

    # --- meses (para derivar el nombre del mes desde la fecha) --------------
    banner(ws, 3, 8, 9, "MESES")
    encabezado_tabla(ws, 4, 8, 9, ["N", "Mes"])
    for i, mes in enumerate(datos.MESES):
        r = 5 + i
        f(ws, f"H{r}", i + 1, align="center", size=9.5)
        f(ws, f"I{r}", mes, align="center", size=9.5)
    R["meses"] = "$I$5:$I$16"

    # --- ciudad de referencia por region ------------------------------------
    banner(ws, 18, 8, 9, "CIUDAD POR REGION")
    encabezado_tabla(ws, 19, 8, 9, ["Region", "Ciudad"])
    for i, (region, (_, ciudad)) in enumerate(datos.REGIONES.items()):
        r = 20 + i
        f(ws, f"H{r}", region, size=9.5)
        f(ws, f"I{r}", ciudad, size=9.5)
    R["reg_region"] = "$H$20:$H$24"
    R["reg_ciudad"] = "$I$20:$I$24"
    f(ws, "H26", "Se usa para completar las ciudades que el export deja "
                 "vacias.", size=8.5, color=GRAY, wrap=True)
    ws.merge_cells("H26:I27")

    # --- diccionarios de normalizacion de texto -----------------------------
    fila = max(fin_tasa + 3, 46)
    banner(ws, fila, 2, 6,
           "DICCIONARIO DE NORMALIZACION DE TEXTO  ·  valor recibido -> valor "
           "estandar")
    f(ws, f"F{fila + 1}",
      "Cada fila es una variante que el sistema transaccional devuelve para un "
      "mismo valor real. El ETL busca el texto recibido (sin espacios de mas) "
      "en la columna izquierda y escribe el estandar de la derecha. Agregar "
      "una variante nueva es agregar una fila: no hay que tocar ninguna "
      "formula.", size=8.5, color=GRAY, wrap=True, vertical="top")
    ws.merge_cells(f"F{fila + 1}:F{fila + 10}")
    fila += 1

    for clave, titulo, mapa in (
            ("cat", "CATEGORIA", datos.VARIANTES_CATEGORIA),
            ("reg", "REGION", datos.VARIANTES_REGION),
            ("can", "CANAL", datos.VARIANTES_CANAL)):
        f(ws, f"B{fila}", f"{titulo}  ·  valor recibido en el export",
          bold=True, size=9, color=NAVY, fill=LIGHT)
        f(ws, f"C{fila}", "Valor estandar", bold=True, size=9, color=NAVY,
          fill=LIGHT, align="center")
        fila += 1
        ini = fila
        for estandar, variantes in mapa.items():
            for var in variantes:
                # el guion bajo hace visible el espacio sobrante en la planilla
                f(ws, f"B{fila}", var, size=9.5,
                  color=AMBER if var != estandar else NAVY)
                f(ws, f"C{fila}", estandar, size=9.5, align="center")
                fila += 1
        R[f"dic_{clave}_recibido"] = f"$B${ini}:$B${fila - 1}"
        R[f"dic_{clave}_estandar"] = f"$C${ini}:$C${fila - 1}"
        fila += 2

    # --- listas para validacion de datos ------------------------------------
    banner(ws, 3, 11, 13, "LISTAS DE VALIDACION")
    encabezado_tabla(ws, 4, 11, 13, ["Region", "Categoria", "Anio"])
    for i, region in enumerate(datos.REGIONES):
        f(ws, f"K{5 + i}", region, size=9.5)
    for i, cat in enumerate(datos.CATEGORIAS):
        f(ws, f"L{5 + i}", cat, size=9.5)
    for i, anio in enumerate(("Todos", 2024, 2025)):
        f(ws, f"M{5 + i}", anio, size=9.5, align="center")
    for col in "KLM":
        ws.column_dimensions[col].width = 16
    R.update(lista_region="$K$5:$K$9", lista_categoria="$L$5:$L$8",
             lista_anio="$M$5:$M$7")

    # nombres definidos: las listas se referencian por nombre, no por rango
    for nombre, rango in (("Lista_Region", R["lista_region"]),
                          ("Lista_Categoria", R["lista_categoria"]),
                          ("Lista_Anio", R["lista_anio"])):
        wb.defined_names.add(_defined_name(nombre, f"{HOJA}!{rango}"))

    f(ws, "K11", "Alimentan las listas desplegables del bloque de consulta de "
                 "la hoja Tablas_Dinamicas.", size=8.5, color=GRAY, wrap=True)
    ws.merge_cells("K11:M13")

    # el area de impresion incluye los bloques laterales (meses,
    # ciudades y listas): son parte de la documentacion, no adornos
    imprimible(ws, f"B2:M{fila}")
    ws.freeze_panes = "A4"
    return R


def _defined_name(nombre, destino):
    from openpyxl.workbook.defined_name import DefinedName
    return DefinedName(nombre, attr_text=destino)


def validacion(rango_lista, titulo, mensaje):
    """Validacion de lista reutilizable, con mensaje de ayuda."""
    dv = DataValidation(type="list", formula1=rango_lista, allow_blank=False,
                        showDropDown=False, showErrorMessage=True,
                        errorTitle="Valor no permitido", error=mensaje,
                        promptTitle=titulo, prompt=mensaje, showInputMessage=True)
    return dv
