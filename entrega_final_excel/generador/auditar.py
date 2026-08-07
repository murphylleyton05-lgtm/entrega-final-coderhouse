"""
Auditoria estructural del .xlsx terminado, punto por punto de la consigna.

`verificar.py` controla que los NUMEROS sean correctos. Este modulo controla
que el ARCHIVO tenga lo que la consigna pide y que el paquete OOXML este bien
formado, que es lo unico que se puede comprobar sin abrir Excel.
"""
import re
import sys
import zipfile
from xml.etree import ElementTree

import openpyxl

NS_MAIN = "{http://schemas.openxmlformats.org/spreadsheetml/2006/main}"


def auditar(ruta):
    z = zipfile.ZipFile(ruta)
    partes = z.namelist()
    wb = openpyxl.load_workbook(ruta, data_only=True)
    ok, mal = [], []

    def check(cond, texto, detalle=""):
        (ok if cond else mal).append(f"{texto}{'  -> ' + detalle if detalle else ''}")

    # --- 0. el paquete abre y todo XML parsea ------------------------------
    rotos = []
    for n in partes:
        if n.endswith((".xml", ".rels")):
            try:
                ElementTree.fromstring(z.read(n))
            except ElementTree.ParseError as e:
                rotos.append(f"{n}: {e}")
    check(not rotos, f"las {len(partes)} partes del paquete parsean",
          "; ".join(rotos[:3]))

    # cada Override de [Content_Types] tiene que existir de verdad
    ct = z.read("[Content_Types].xml").decode()
    faltantes = [p for p in re.findall(r'PartName="/([^"]+)"', ct)
                 if p not in partes]
    check(not faltantes, "los Content_Types apuntan a partes existentes",
          ", ".join(faltantes))

    # y toda relacion interna tiene destino
    huerfanas = []
    for n in [p for p in partes if p.endswith(".rels")]:
        base = n.rsplit("_rels/", 1)[0]
        for t in re.findall(r'Target="([^"]+)"[^>]*?(?:TargetMode="External")?/>',
                            z.read(n).decode()):
            if t.startswith(("http", "../..")) or "TargetMode" in t:
                continue
            destino = _resolver(base, t)
            if destino and destino not in partes:
                huerfanas.append(f"{n} -> {t}")
    check(not huerfanas, "las relaciones internas apuntan a partes existentes",
          "; ".join(huerfanas[:3]))

    # --- 1. datos limpios, sin duplicados y con formatos correctos ---------
    wsl = wb["Datos_Limpios"]
    ids = [wsl[f"A{r}"].value for r in range(2, 2402)]
    check(len(set(ids)) == 2400 and None not in ids,
          f"Datos_Limpios: {len(set(ids))} operaciones unicas, sin duplicados")
    wf = openpyxl.load_workbook(ruta)          # formulas, no valores
    fechas_fmt = {wf["Datos_Limpios"][f"B{r}"].number_format
                  for r in range(2, 2402)}
    check(fechas_fmt == {"dd/mm/yyyy"},
          f"Datos_Limpios: la fecha tiene formato de fecha", str(fechas_fmt))
    tipos = {type(wsl[f"B{r}"].value).__name__ for r in range(2, 2402)}
    check(tipos == {"datetime"},
          f"Datos_Limpios: las 2.400 fechas son fechas reales", str(tipos))
    unid = {type(wsl[f"O{r}"].value).__name__ for r in range(2, 2402)}
    check(unid <= {"int", "float"},
          "Datos_Limpios: las unidades son numeros", str(unid))
    check("Datos_Origen" in wb.sheetnames,
          "el export crudo se conserva en Datos_Origen")

    wc = wb["Control_Calidad"]
    estados = [wc[f"F{r}"].value for r in range(8, 16)]
    check(all(e == "OK" for e in estados),
          "Control_Calidad: los 8 controles de calidad dan OK", str(estados))

    # --- 2. columna calculada con funciones logicas ------------------------
    wfl = wf["Datos_Limpios"]
    logicas = {}
    for col, nombre in (("W", "Segmento_Rentabilidad"), ("X", "Alerta_Logistica"),
                        ("Y", "Ticket")):
        formula = str(wfl[f"{col}2"].value or "")
        logicas[nombre] = formula
    tiene_y = any("AND(" in f for f in logicas.values())
    tiene_o = any("OR(" in f for f in logicas.values())
    tiene_si = all(f.startswith("=IF(") for f in logicas.values())
    check(tiene_si and tiene_y and tiene_o,
          f"{len(logicas)} columnas calculadas con SI, Y y O",
          f"SI={tiene_si} Y={tiene_y} O={tiene_o}")
    valores = {wsl[f"W{r}"].value for r in range(2, 2402)}
    check(valores == {"Alta rentabilidad", "Estandar", "Revisar"},
          "Segmento_Rentabilidad devuelve las tres clases", str(valores))

    # --- 3. tablas dinamicas -----------------------------------------------
    pivots = sorted(p for p in partes
                    if re.match(r"xl/pivotTables/pivotTable\d+\.xml$", p))
    visibles, medidas = [], set()
    for p in pivots:
        x = z.read(p).decode()
        nombre = re.search(r'name="([^"]+)"', x).group(1)
        if nombre.startswith("TD_"):
            visibles.append(nombre)
            medidas |= set(re.findall(r'<dataField name="([^"]+)"', x))
    check(len(pivots) >= 2, f"{len(pivots)} tablas dinamicas en el libro")
    check(len(visibles) >= 2,
          f"{len(visibles)} tablas dinamicas a la vista: {', '.join(visibles)}")
    check({"Ventas", "Margen", "Unidades"} <= medidas,
          f"resumen Ventas, Margen y Unidades", str(sorted(medidas)))
    check(len(set(re.findall(r'cacheId="(\d+)"',
                             "".join(z.read(p).decode() for p in pivots)))) == 1,
          "todas comparten un unico cache (los segmentadores las mueven juntas)")
    wt = wb["Tablas_Dinamicas"]
    check(isinstance(wt["C8"].value, (int, float)),
          "las tablas dinamicas traen sus valores ya calculados",
          repr(wt["C8"].value))

    # --- 4. dashboard: graficos y segmentadores ----------------------------
    charts = [p for p in partes if re.match(r"xl/charts/chart\d+\.xml$", p)]
    check(len(charts) >= 2, f"{len(charts)} graficos en el tablero")
    slicers = [p for p in partes if p.startswith("xl/slicers/")]
    caches = [p for p in partes if p.startswith("xl/slicerCaches/")]
    n_slicers = 0
    if slicers:
        n_slicers = z.read(slicers[0]).decode().count("<slicer ")
    check(n_slicers >= 1,
          f"{n_slicers} segmentadores, con {len(caches)} caches de segmentador")
    dash = wb["Dashboard"]
    sheet_dash = _sheet_part(z, "Dashboard")
    check("x14:slicerList" in z.read(sheet_dash).decode(),
          "los segmentadores estan anclados en la hoja Dashboard")
    # los graficos leen la grilla que resuelve GETPIVOTDATA
    calc = wf["Calculos"]
    gpd = sum(1 for fila in calc.iter_rows()
              for c in fila if isinstance(c.value, str)
              and "GETPIVOTDATA" in c.value)
    check(gpd > 50,
          f"{gpd} formulas GETPIVOTDATA conectan los graficos con las "
          f"tablas dinamicas")

    # --- 5. diseno profesional ---------------------------------------------
    xml_dash = z.read(sheet_dash).decode()
    check('showGridLines="0"' in xml_dash,
          "el tablero no muestra lineas de cuadricula")
    check('showRowColHeaders="0"' in xml_dash,
          "el tablero no muestra encabezados de fila y columna")
    etiquetas = sum(z.read(c).decode().count('<c:showVal val="1"/>')
                    for c in charts)
    check(etiquetas >= 6,
          f"{etiquetas} series con etiquetas de datos visibles")
    titulos = sum(1 for c in charts if "<c:strRef>" in z.read(c).decode()
                  and "<c:title>" in z.read(c).decode())
    check(titulos == len(charts),
          f"los {titulos} titulos de grafico son formulas que leen los datos")
    ocultas = [s for s in wb.worksheets if s.sheet_state != "visible"]
    check({s.title for s in ocultas} == {"Calculos", "Pivots_Aux"},
          f"las hojas de soporte estan ocultas: "
          f"{[s.title for s in ocultas]}")
    check(wb.sheetnames[0] == "Dashboard",
          f"el libro abre en el tablero", wb.sheetnames[0])

    # --- extras del nivel recomendado --------------------------------------
    dv = wf["Tablas_Dinamicas"].data_validations.dataValidation
    check(len(dv) >= 3, f"{len(dv)} listas de validacion en el bloque de "
                        f"consulta")
    check(len(wf.defined_names) >= 3,
          f"{len(wf.defined_names)} nombres definidos para las listas")
    ytd = any("YTD" in str(c.value) for fila in wf["Calculos"].iter_rows()
              for c in fila)
    check(ytd, "hay medidas de inteligencia de tiempo (acumulado del ano)")
    cf = len(wf["Dashboard"].conditional_formatting._cf_rules)
    check(cf >= 4, f"{cf} reglas de formato condicional en las tarjetas de KPI")

    # --- informe ------------------------------------------------------------
    print(f"\nAUDITORIA DE {ruta}")
    print("=" * 72)
    for t in ok:
        print(f"  OK    {t}")
    for t in mal:
        print(f"  FALLA {t}")
    print("=" * 72)
    print(f"{len(ok)} controles OK, {len(mal)} fallas")
    return not mal


def _resolver(base, target):
    partes = [p for p in base.split("/") if p]
    for tramo in target.split("/"):
        if tramo == "..":
            if partes:
                partes.pop()
        elif tramo not in (".", ""):
            partes.append(tramo)
    return "/".join(partes)


def _sheet_part(z, nombre):
    wbxml = z.read("xl/workbook.xml").decode()
    for i, m in enumerate(re.finditer(r"<sheet\b[^>]*/>", wbxml), start=1):
        if re.search(r'name="%s"' % re.escape(nombre), m.group(0)):
            return f"xl/worksheets/sheet{i}.xml"
    raise KeyError(nombre)


if __name__ == "__main__":
    sys.exit(0 if auditar(sys.argv[1]) else 1)
