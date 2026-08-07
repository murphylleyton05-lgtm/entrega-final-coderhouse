"""
Hoja `Dashboard`: el entregable ejecutivo.

Estructura de arriba hacia abajo, que es el orden en que se lee un reporte:

    cabecera -> titular y bajada -> 4 tarjetas de KPI -> 4 segmentadores
    -> 1. QUE PASO      (evolucion mensual + acumulado del ano)
    -> 2. POR QUE       (puente de margen + margen por categoria)
    -> 3. DONDE ACTUAR  (mapa de productos: crecimiento contra margen)
    -> reporte ejecutivo (evidencia + insight) -> fuente

Sin lineas de cuadricula, sin encabezados de fila y columna, tres colores y
nada de vocabulario tecnico: la gerencia ve metricas de negocio, no comandos.
"""
import datetime as dt

from openpyxl.chart import BarChart, LineChart, Reference, ScatterChart, Series
from openpyxl.chart.data_source import StrRef
from openpyxl.chart.label import DataLabelList
from openpyxl.chart.marker import Marker
from openpyxl.chart.shapes import GraphicalProperties
from openpyxl.chart.text import Text
from openpyxl.chart.title import Title
from openpyxl.drawing.line import LineProperties
from openpyxl.formatting.rule import FormulaRule
from openpyxl.styles import Font, PatternFill
from openpyxl.utils import get_column_letter

import analisis
from analisis import CALC, R_CASC, R_CAT, R_MES, R_PROD, T
from estilos import (AMBER, AMBER_BG, FUENTE, GRAY, INK, LIGHT, LINE, MONEY,
                     MONEY_K, NAVY, NAVY_SOFT, PCT1, PCT2, WHITE, banda, caja,
                     f)

DASH = "Dashboard"

# Filas de anclaje. Un grafico de 8,8 cm mide 249 pt y las filas del cuerpo
# miden 15 pt, asi que ocupa 16,6 filas: por eso las secciones van cada 19.
FILA_SEG_DESDE, FILA_SEG_HASTA = 14, 23      # base 0 -> filas 15..23
FILA_S1, FILA_S2, FILA_S3 = 25, 44, 63       # banners de seccion
FILA_PIE_DISP = 82       # pie del mapa de productos
FILA_NOTAS = 84          # las tres notas ocupan 85-86, 87-88 y 89-90
FILA_FUENTE = 92         # la fuente ocupa 92-93


def _titulo_dinamico(chart, ref_celda):
    t = Title()
    t.tx = Text(strRef=StrRef(ref_celda))
    chart.title = t


def _serie_color(s, color, linea=False):
    s.graphicalProperties = GraphicalProperties(solidFill=color)
    if linea:
        s.graphicalProperties.line = LineProperties(solidFill=color, w=22000)
        s.graphicalProperties.noFill = True
    else:
        s.graphicalProperties.line = LineProperties(noFill=True)


# =============================================================================
def hoja(wb, F, K, D):
    ws = wb.create_sheet(DASH, 0)
    ws.sheet_view.showGridLines = False
    ws.sheet_view.showRowColHeaders = False
    ws.sheet_view.zoomScale = 85
    # que imprima como se ve: apaisado, una hoja de ancho
    ws.page_setup.orientation = "landscape"
    ws.page_setup.fitToWidth = 1
    ws.page_setup.fitToHeight = 0
    ws.sheet_properties.pageSetUpPr.fitToPage = True
    ws.print_area = f"A1:R{FILA_FUENTE + 2}"

    ws.column_dimensions["A"].width = 2.2
    for i in range(2, 18):                       # B..Q
        ws.column_dimensions[get_column_letter(i)].width = 9.4
    ws.column_dimensions["R"].width = 2.2

    alturas = {1: 6, 2: 26, 3: 15, 4: 8, 5: 6, 6: 22, 7: 16, 8: 8,
               9: 14, 10: 24, 11: 12, 12: 14, 13: 10, 14: 16}
    for r, h in alturas.items():
        ws.row_dimensions[r].height = h
    for r in range(15, 24):
        ws.row_dimensions[r].height = 17
    for r in range(24, FILA_FUENTE + 3):
        ws.row_dimensions[r].height = 15

    # --- cabecera -----------------------------------------------------------
    banda(ws, "A1:R4", NAVY)
    ws.merge_cells("B2:K2")
    f(ws, "B2", "TechStore  |  Reporte ejecutivo de ventas y margen",
      bold=True, size=17, color=WHITE, align="left")
    ws.merge_cells("L2:Q3")
    f(ws, "L2", f"Actualizado: {dt.date.today():%d/%m/%Y}   ·   Moneda: USD",
      size=9, color=NAVY_SOFT, align="right")
    ws.merge_cells("B3:K3")
    f(ws, "B3", "Ejercicios 2024-2025  ·  2.400 operaciones depuradas  ·  "
                "5 regiones  ·  2 canales", size=10, color=NAVY_SOFT,
      align="left")

    # --- titular narrativo --------------------------------------------------
    ws.merge_cells("B6:Q6")
    f(ws, "B6", f"={T['titular']}", bold=True, size=13, color=NAVY,
      align="left")
    ws.merge_cells("B7:Q7")
    f(ws, "B7", f"={T['bajada']}", size=9.5, color=GRAY, italic=True,
      align="left")

    # --- tarjetas de KPI ----------------------------------------------------
    tarjetas = [
        ("B", "E", "VENTAS 2025", f"={CALC}!$B${K['ventas']}", MONEY,
         f'="vs meta "&TEXT({CALC}!$C${K["ventas"]},"$#,##0")&"   "&'
         f'TEXT({CALC}!$D${K["ventas"]},"+#,##0;-#,##0")&" USD"',
         K["ventas"]),
        ("F", "I", "MARGEN BRUTO", f"={CALC}!$B${K['margen']}", PCT1,
         f'="vs meta "&TEXT({CALC}!$C${K["margen"]},"0.0%")&"   "&'
         f'TEXT({CALC}!$B${D["margen_pp"]},"+0.0;-0.0")&" p.p. vs 2024"',
         K["margen"]),
        ("J", "M", "CRECIMIENTO INTERANUAL", f"={CALC}!$B${K['yoy']}", PCT1,
         f'="vs meta "&TEXT({CALC}!$C${K["yoy"]},"0.0%")&"   "&'
         f'TEXT({CALC}!$D${K["yoy"]}*100,"+0.0;-0.0")&" p.p."',
         K["yoy"]),
        ("N", "Q", "COSTO LOGISTICO / VENTAS", f"={CALC}!$B${K['log']}", PCT2,
         f'="techo "&TEXT({CALC}!$C${K["log"]},"0.00%")&"   "&'
         f'TEXT({CALC}!$B${D["log_pp"]},"+0.00;-0.00")&" p.p. vs 2024"',
         K["log"]),
    ]
    for c1, c2, etiqueta, valor, fmt, pie, fila_kpi in tarjetas:
        ws.merge_cells(f"{c1}9:{c2}9")
        f(ws, f"{c1}9", etiqueta, bold=True, size=8, color=GRAY, align="left")
        ws.merge_cells(f"{c1}10:{c2}11")
        f(ws, f"{c1}10", valor, bold=True, size=20, color=NAVY, fmt=fmt,
          align="left")
        ws.merge_cells(f"{c1}12:{c2}12")
        f(ws, f"{c1}12", pie, size=8.5, color=GRAY, align="left")
        banda(ws, f"{c1}9:{c2}12", LIGHT)
        caja(ws, f"{c1}9:{c2}12", LINE)
        # espejo local de la brecha: el formato condicional lo lee sin salir
        # de la hoja (mas robusto que una referencia entre hojas)
        f(ws, f"{c1}13", f"={CALC}!$D${fila_kpi}", size=8, color=WHITE)
        # por debajo de la meta la tarjeta entera pasa a ambar
        ws.conditional_formatting.add(
            f"{c1}9:{c2}12",
            FormulaRule(formula=[f"${c1}$13<0"],
                        font=Font(name=FUENTE, color=AMBER, b=True),
                        fill=PatternFill(start_color=AMBER_BG,
                                         end_color=AMBER_BG,
                                         fill_type="solid")))

    # --- segmentadores ------------------------------------------------------
    ws.merge_cells("B14:Q14")
    f(ws, "B14", "FILTROS  ·  cualquier seleccion recalcula a la vez las "
                 "cuatro tarjetas y los cinco graficos",
      bold=True, size=8.5, color=GRAY, align="left")
    banda(ws, "B15:Q23", WHITE)

    # --- banners de seccion -------------------------------------------------
    secciones = [
        (FILA_S1, "1 · QUE PASO  ·  evolucion de las ventas y del margen"),
        (FILA_S2, "2 · POR QUE  ·  de donde sale la caida del margen"),
        (FILA_S3, "3 · DONDE ACTUAR  ·  que productos miran primero"),
    ]
    for fila, texto in secciones:
        ws.merge_cells(f"B{fila}:Q{fila}")
        f(ws, f"B{fila}", texto, bold=True, size=9, color=WHITE, fill=NAVY,
          align="left")
        for i in range(2, 18):
            ws.cell(fila, i).fill = PatternFill("solid", fgColor=NAVY)
        ws[f"B{fila}"].value = texto
    return ws


# =============================================================================
def graficos(wb, ws, productos_riesgo):
    calc = wb[CALC]

    # --- 1. evolucion mensual: columnas 2024/2025 + linea de margen % -------
    barras = BarChart()
    barras.type = "col"
    barras.grouping = "clustered"
    barras.gapWidth = 60
    barras.overlap = -10
    barras.add_data(Reference(calc, min_col=2, max_col=3, min_row=R_MES - 1,
                              max_row=R_MES + 11), titles_from_data=True)
    barras.set_categories(Reference(calc, min_col=1, min_row=R_MES,
                                    max_row=R_MES + 11))
    _serie_color(barras.series[0], GRAY)
    _serie_color(barras.series[1], NAVY)
    barras.y_axis.numFmt = MONEY_K
    barras.y_axis.title = None

    linea = LineChart()
    linea.add_data(Reference(calc, min_col=4, min_row=R_MES - 1,
                             max_row=R_MES + 11), titles_from_data=True)
    linea.y_axis.axId = 200
    linea.y_axis.numFmt = PCT1
    linea.y_axis.majorGridlines = None
    linea.y_axis.crosses = "max"
    _serie_color(linea.series[0], AMBER, linea=True)
    linea.series[0].smooth = False
    linea.series[0].marker = Marker(symbol="circle", size=5)
    linea.series[0].dLbls = DataLabelList(showVal=True, showSerName=False,
                                          showCatName=False, showLegendKey=False)
    barras += linea
    _titulo_dinamico(barras, T["g1"])
    barras.height, barras.width = 8.8, 13.2
    barras.legend.position = "b"
    ws.add_chart(barras, "B26")

    # --- 2. acumulado del ano (YTD): 2024 contra 2025 -----------------------
    ytd = LineChart()
    ytd.add_data(Reference(calc, min_col=7, max_col=8, min_row=R_MES - 1,
                           max_row=R_MES + 11), titles_from_data=True)
    ytd.set_categories(Reference(calc, min_col=10, min_row=R_MES,
                                 max_row=R_MES + 11))
    _serie_color(ytd.series[0], GRAY, linea=True)
    _serie_color(ytd.series[1], NAVY, linea=True)
    for s in ytd.series:
        s.smooth = False
        s.marker = Marker(symbol="circle", size=5)
    ytd.y_axis.numFmt = MONEY_K
    _titulo_dinamico(ytd, T["g2"])
    ytd.height, ytd.width = 8.8, 13.2
    ytd.legend.position = "b"
    ws.add_chart(ytd, "J26")

    # --- 3. puente de margen: apilado con la base invisible ------------------
    casc = BarChart()
    casc.type = "col"
    casc.grouping = "stacked"
    casc.overlap = 100
    casc.gapWidth = 45
    casc.add_data(Reference(calc, min_col=4, max_col=7, min_row=R_CASC - 1,
                            max_row=R_CASC + 5), titles_from_data=True)
    casc.set_categories(Reference(calc, min_col=1, min_row=R_CASC,
                                  max_row=R_CASC + 5))
    base, total, suma, resta = casc.series
    base.graphicalProperties = GraphicalProperties(noFill=True)
    base.graphicalProperties.line = LineProperties(noFill=True)
    _serie_color(total, NAVY)
    _serie_color(suma, GRAY)
    _serie_color(resta, AMBER)
    for s in (total, suma, resta):
        s.dLbls = DataLabelList(showVal=True, showSerName=False,
                                showCatName=False, showLegendKey=False)
    casc.y_axis.numFmt = MONEY_K
    _titulo_dinamico(casc, T["g3"])
    casc.height, casc.width = 8.8, 13.2
    casc.legend.position = "b"
    ws.add_chart(casc, f"B{FILA_S2 + 1}")

    # --- 4. margen % por categoria, 2024 contra 2025 ------------------------
    cat = BarChart()
    cat.type = "bar"
    cat.grouping = "clustered"
    cat.gapWidth = 60
    cat.add_data(Reference(calc, min_col=2, max_col=3, min_row=R_CAT - 1,
                           max_row=R_CAT + 3), titles_from_data=True)
    cat.set_categories(Reference(calc, min_col=1, min_row=R_CAT,
                                 max_row=R_CAT + 3))
    _serie_color(cat.series[0], GRAY)
    _serie_color(cat.series[1], NAVY)
    for s in cat.series:
        s.dLbls = DataLabelList(showVal=True, showSerName=False,
                                showCatName=False, showLegendKey=False)
    cat.x_axis.numFmt = PCT1
    _titulo_dinamico(cat, T["g4"])
    cat.height, cat.width = 8.8, 13.2
    cat.legend.position = "b"
    ws.add_chart(cat, f"J{FILA_S2 + 1}")

    # --- 5. mapa de productos: crecimiento contra margen --------------------
    disp = ScatterChart()
    disp.x_axis.numFmt = PCT1
    disp.y_axis.numFmt = PCT1
    disp.x_axis.title = "Crecimiento de ventas 2025 vs 2024"
    disp.y_axis.title = "Margen bruto % 2025"
    disp.style = None
    for col_x, col_y, nombre, color in ((5, 6, "En meta de margen", NAVY),
                                        (7, 8, "Zona de riesgo", AMBER)):
        s = Series(Reference(calc, min_col=col_y, min_row=R_PROD,
                             max_row=R_PROD + 9),
                   Reference(calc, min_col=col_x, min_row=R_PROD,
                             max_row=R_PROD + 9),
                   title=nombre)
        s.marker = Marker(symbol="circle", size=9)
        s.marker.graphicalProperties = GraphicalProperties(solidFill=color)
        s.marker.graphicalProperties.line = LineProperties(solidFill=color)
        s.graphicalProperties = GraphicalProperties()
        s.graphicalProperties.line = LineProperties(noFill=True)
        disp.series.append(s)
    _titulo_dinamico(disp, T["g5"])
    disp.height, disp.width = 8.8, 27.2
    disp.legend.position = "b"
    ws.add_chart(disp, f"B{FILA_S3 + 1}")


# =============================================================================
def reporte_ejecutivo(ws):
    """Cuadro de texto del tablero: evidencia primero, insight despues."""
    ws.merge_cells(f"B{FILA_NOTAS}:Q{FILA_NOTAS}")
    f(ws, f"B{FILA_NOTAS}",
      "REPORTE EJECUTIVO  ·  cada linea dice primero el dato y despues que "
      "hacer con el", bold=True, size=9, color=WHITE, fill=NAVY, align="left")
    for i in range(2, 18):
        ws.cell(FILA_NOTAS, i).fill = PatternFill("solid", fgColor=NAVY)
    ws[f"B{FILA_NOTAS}"].value = ("REPORTE EJECUTIVO  ·  cada linea dice "
                                  "primero el dato y despues que hacer con el")

    for i, clave in enumerate(("nota1", "nota2", "nota3")):
        r = FILA_NOTAS + 1 + i * 2
        ws.merge_cells(f"B{r}:Q{r + 1}")
        f(ws, f"B{r}", f"={T[clave]}", size=9, color=INK, align="left",
          wrap=True, vertical="center")
        banda(ws, f"B{r}:Q{r + 1}", LIGHT)
        caja(ws, f"B{r}:Q{r + 1}", LINE)
        ws[f"B{r}"].value = f"={T[clave]}"

    # el mapa de productos no puede rotular sus puntos, asi que los nombra aca
    ws.merge_cells(f"B{FILA_PIE_DISP}:Q{FILA_PIE_DISP}")
    f(ws, f"B{FILA_PIE_DISP}", f"={T['pie5']}", size=8.5, color=GRAY,
      italic=True, align="left")

    ws.merge_cells(f"B{FILA_FUENTE}:Q{FILA_FUENTE + 1}")
    f(ws, f"B{FILA_FUENTE}",
      "Fuente: export del sistema transaccional depurado en este mismo libro "
      "(hojas Datos_Origen -> Datos_Limpios -> Control_Calidad). Metas y "
      "supuestos en la hoja Parametros. Dataset simulado con fines "
      "academicos. Elaborado por Lleyton Murphy · Coderhouse, Data Analyst · "
      "Practica Integradora Final.", size=8, color=GRAY, align="left",
      wrap=True, vertical="top")
