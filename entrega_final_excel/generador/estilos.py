"""Paleta, formatos y helpers de formato compartidos por todo el libro."""
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side

# --- paleta corporativa: tres colores principales + neutros ------------------
NAVY = "1F3864"    # 1 - institucional / ano corriente / valor en meta
AMBER = "E8A33D"   # 2 - alerta / efectos que restan
GRAY = "8C8C8C"    # 3 - contexto / ano anterior
LIGHT = "F2F4F7"   # neutro de fondo
LINE = "DCE0E6"    # neutro de borde
WHITE = "FFFFFF"
INK = "2F3437"     # neutro de texto
AMBER_BG = "FDF3E3"
NAVY_SOFT = "C7CEDB"

FUENTE = "Arial"

MONEY = '"$"#,##0'
MONEY2 = '"$"#,##0.00'
MONEY_K = '"$"#,##0," k"'
PCT1 = '0.0%'
PCT2 = '0.00%'
NUM = '#,##0'
FECHA = 'dd/mm/yyyy'
PP = '+0.0" p.p.";-0.0" p.p."'


def f(ws, cell, value=None, *, bold=False, size=10, color=INK, fill=None,
      fmt=None, align=None, wrap=False, italic=False, vertical="center"):
    """Escribe y formatea una celda en un solo paso."""
    c = ws[cell]
    if value is not None:
        c.value = value
    c.font = Font(name=FUENTE, sz=size, b=bold, i=italic, color=color)
    if fill:
        c.fill = PatternFill("solid", fgColor=fill)
    if fmt:
        c.number_format = fmt
    c.alignment = Alignment(horizontal=align or "general", vertical=vertical,
                            wrap_text=wrap)
    return c


def banner(ws, fila, col_desde, col_hasta, texto, color=NAVY):
    """Franja de titulo de seccion, del ancho que se le pida."""
    from openpyxl.utils import get_column_letter
    for i in range(col_desde, col_hasta + 1):
        ws.cell(fila, i).fill = PatternFill("solid", fgColor=color)
    c = f(ws, f"{get_column_letter(col_desde)}{fila}", texto, bold=True,
          size=10, color=WHITE, align="left")
    return c


def banda(ws, rango, color):
    for fila in ws[rango]:
        for c in fila:
            c.fill = PatternFill("solid", fgColor=color)


def caja(ws, rango, color=LINE):
    lado = Side(style="thin", color=color)
    celdas = ws[rango]
    n_f, n_c = len(celdas), len(celdas[0])
    for i, fila in enumerate(celdas):
        for j, c in enumerate(fila):
            c.border = Border(
                top=lado if i == 0 else None,
                bottom=lado if i == n_f - 1 else None,
                left=lado if j == 0 else None,
                right=lado if j == n_c - 1 else None)


def imprimible(ws, area=None, apaisado=True):
    """Deja la hoja lista para imprimir: una pagina de ancho, sin cortes raros."""
    ws.page_setup.orientation = "landscape" if apaisado else "portrait"
    ws.page_setup.fitToWidth = 1
    ws.page_setup.fitToHeight = 0
    ws.sheet_properties.pageSetUpPr.fitToPage = True
    if area:
        ws.print_area = area


def encabezado_tabla(ws, fila, col_desde, col_hasta, titulos, color=NAVY):
    from openpyxl.utils import get_column_letter
    for i, t in enumerate(titulos):
        col = get_column_letter(col_desde + i)
        f(ws, f"{col}{fila}", t, bold=True, size=9, color=WHITE, fill=color,
          align="center", wrap=True)
