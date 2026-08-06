"""
Construye el libro "TechStore - Dashboard Final" (Modulo 9, capa de visualizacion).

Estructura del libro
    Dashboard Final : el entregable (KPIs, 4 graficos, 4 segmentadores, narrativa)
    Analisis_IA     : registro del uso de IA como auditor cognitivo
    Datos           : tabla de hechos (2.400 filas, 2024-2025)
    Parametros      : supuestos, metas y tasas -> toda constante vive aca
    Guia_de_uso     : como leer y operar el tablero
    Calculos        : capa intermedia (oculta) que alimenta los graficos
    Pivots          : tablas dinamicas (oculta) que los segmentadores filtran

Como circula el filtro de un segmentador hasta un grafico:
    Slicer -> tabla dinamica (Pivots) -> GETPIVOTDATA (Calculos) -> grafico
"""
import datetime as dt
import os

import openpyxl
from openpyxl.chart import BarChart, LineChart, Reference, ScatterChart, Series
from openpyxl.chart.data_source import StrRef
from openpyxl.chart.marker import Marker
from openpyxl.chart.shapes import GraphicalProperties
from openpyxl.chart.text import Text
from openpyxl.chart.title import Title
from openpyxl.drawing.line import LineProperties
from openpyxl.formatting.rule import FormulaRule
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.table import Table, TableStyleInfo

import datos
from ooxml import Pivot, Slicer, col_letter

# --- paleta: tres colores principales + neutros ------------------------------
NAVY = "1F3864"   # 1 - institucional / valores en meta
AMBER = "E8A33D"  # 2 - alerta / efectos negativos
GRAY = "8C8C8C"   # 3 - contexto / anio anterior
LIGHT = "F2F4F7"  # neutro de fondo
WHITE = "FFFFFF"
INK = "2F3437"    # neutro de texto

FUENTE = "Arial"
MONEY = '"$"#,##0'
MONEY_K = '"$"#,##0," k"'
PCT1 = '0.0%'
PCT2 = '0.00%'
NUM = '#,##0'
PP = '+0.0" p.p.";-0.0" p.p."'

DASH = "Dashboard Final"
CALC = "Calculos"
PIV = "Pivots"

# --- filas de la hoja Calculos (una sola fuente de verdad) -------------------
R_TOT = 4         # 4..13  totales por anio
R_KPI = 17        # 17..22 KPIs contra meta
R_CASC = 28       # 28..33 puente de margen
R_MES = 38        # 38..49 evolucion mensual
R_CAT = 54        # 54..57 categorias
R_PROD = 62       # 62..71 productos
R_TXT = 75        # 75..83 textos narrativos

# orden de los textos narrativos; fija en que celda cae cada uno
CLAVES_TXT = ["titular", "bajada", "g1", "g2", "g3", "g4",
              "nota1", "nota2", "nota3"]
T_TITULOS = {k: f"{CALC}!$B${R_TXT + i}" for i, k in enumerate(CLAVES_TXT)}


def _f(ws, cell, value, *, bold=False, size=10, color=INK, fill=None,
       fmt=None, align=None, wrap=False, italic=False):
    c = ws[cell]
    c.value = value
    c.font = Font(name=FUENTE, sz=size, b=bold, i=italic, color=color)
    if fill:
        c.fill = PatternFill("solid", fgColor=fill)
    if fmt:
        c.number_format = fmt
    if align or wrap:
        c.alignment = Alignment(horizontal=align or "general",
                                vertical="center", wrap_text=wrap)
    return c


# =============================================================================
#  Datos / Parametros
# =============================================================================
def hoja_datos(wb, columnas, filas):
    ws = wb.create_sheet("Datos")
    ix = {c: i + 1 for i, c in enumerate(columnas)}
    ws.append(columnas)
    for c in range(1, len(columnas) + 1):
        cell = ws.cell(1, c)
        cell.font = Font(name=FUENTE, sz=10, b=True, color=WHITE)
        cell.fill = PatternFill("solid", fgColor=NAVY)
        cell.alignment = Alignment(horizontal="center", vertical="center")

    cU, cP, cC = (get_column_letter(ix["Unidades"]),
                  get_column_letter(ix["Precio_Unitario"]),
                  get_column_letter(ix["Costo_Unitario"]))
    cV, cCP, cCL = (get_column_letter(ix["Ventas"]),
                    get_column_letter(ix["Costo_Producto"]),
                    get_column_letter(ix["Costo_Logistico"]))
    cCanal, cAnio, cTrim = (get_column_letter(ix["Canal"]),
                            get_column_letter(ix["Anio"]),
                            get_column_letter(ix["Trimestre"]))

    for r, fila in enumerate(filas, start=2):
        for c, v in enumerate(fila, start=1):
            ws.cell(r, c, v)
        # las medidas se calculan en la planilla, no se pegan como numeros
        ws.cell(r, ix["Ventas"], f"=ROUND({cU}{r}*{cP}{r},2)")
        ws.cell(r, ix["Costo_Producto"], f"=ROUND({cU}{r}*{cC}{r},2)")
        # la tasa no se repite fila a fila: se busca en la tabla de Parametros
        ws.cell(r, ix["Costo_Logistico"],
                f"=ROUND({cV}{r}*SUMIFS(Parametros!$E$23:$E$38,"
                f"Parametros!$B$23:$B$38,{cCanal}{r},"
                f"Parametros!$C$23:$C$38,{cAnio}{r},"
                f"Parametros!$D$23:$D$38,{cTrim}{r}),2)")
        ws.cell(r, ix["Margen_Bruto"], f"={cV}{r}-{cCP}{r}-{cCL}{r}")

    last = len(filas) + 1
    for col, fmt in ((ix["Precio_Unitario"], MONEY), (ix["Costo_Unitario"], MONEY),
                     (ix["Ventas"], MONEY), (ix["Costo_Producto"], MONEY),
                     (ix["Costo_Logistico"], MONEY), (ix["Margen_Bruto"], MONEY),
                     (ix["Unidades"], NUM)):
        for r in range(2, last + 1):
            ws.cell(r, col).number_format = fmt
    for r in range(2, last + 1):
        ws.cell(r, ix["Fecha"]).number_format = "dd/mm/yyyy"

    anchos = {"ID_Venta": 9, "Fecha": 11, "Anio": 7, "Trimestre": 10,
              "Mes_Num": 8, "Mes": 7, "Periodo": 10, "Region": 11,
              "Ciudad": 14, "Canal": 14, "Categoria": 15, "Cat_Anio": 19,
              "Producto": 20, "Prod_Anio": 24, "Unidades": 10,
              "Precio_Unitario": 14, "Costo_Unitario": 14, "Ventas": 13,
              "Costo_Producto": 15, "Costo_Logistico": 15, "Margen_Bruto": 14}
    for name, w in anchos.items():
        ws.column_dimensions[get_column_letter(ix[name])].width = w

    ref = f"A1:{get_column_letter(len(columnas))}{last}"
    t = Table(displayName="tbl_Ventas", ref=ref)
    t.tableStyleInfo = TableStyleInfo(name="TableStyleLight1", showRowStripes=True)
    ws.add_table(t)
    ws.freeze_panes = "A2"
    return ws


def hoja_parametros(wb):
    ws = wb.create_sheet("Parametros")
    ws.sheet_view.showGridLines = False
    ws.column_dimensions["A"].width = 2
    ws.column_dimensions["B"].width = 42
    for c in "CDE":
        ws.column_dimensions[c].width = 16
    ws.column_dimensions["F"].width = 62

    _f(ws, "B2", "Parametros, supuestos y metas", bold=True, size=15, color=NAVY)
    _f(ws, "B3", "Todo numero fijo del modelo vive en esta hoja. Los calculos "
                 "del tablero lo referencian; ninguno lo repite escrito a mano.",
       size=9, color=GRAY, italic=True)

    _f(ws, "B5", "METAS DEL EJERCICIO 2025", bold=True, size=10, color=WHITE,
       fill=NAVY)
    for c in "CDEF":
        ws[f"{c}5"].fill = PatternFill("solid", fgColor=NAVY)
    _f(ws, "B6", "Indicador", bold=True, fill=LIGHT)
    _f(ws, "C6", "Meta", bold=True, fill=LIGHT, align="center")
    _f(ws, "F6", "Origen del supuesto", bold=True, fill=LIGHT)

    metas = [
        ("Ventas totales 2025", 5_400_000, MONEY,
         "Presupuesto comercial aprobado para el ejercicio (dato de negocio)."),
        ("Margen bruto %", 0.34, PCT1,
         "Piso historico de la cadena: el margen 2024 cerro en 35,8%."),
        ("Crecimiento interanual de ventas", 0.15, PCT1,
         "Objetivo de expansion fijado por la direccion."),
        ("Costo logistico sobre ventas", 0.025, PCT2,
         "Techo tolerado; 2024 cerro en 2,18%."),
        ("Margen bruto % minimo por producto", 0.33, PCT1,
         "Corte usado en el grafico de dispersion para marcar la zona de riesgo."),
    ]
    for i, (nombre, valor, fmt, origen) in enumerate(metas):
        r = 7 + i
        _f(ws, f"B{r}", nombre)
        _f(ws, f"C{r}", valor, bold=True, color=NAVY, fmt=fmt, align="center")
        _f(ws, f"F{r}", origen, size=9, color=GRAY)

    _f(ws, "B14", "SUPUESTOS DEL DATASET SIMULADO", bold=True, size=10,
       color=WHITE, fill=NAVY)
    for c in "CDEF":
        ws[f"{c}14"].fill = PatternFill("solid", fgColor=NAVY)
    sup = [
        ("Periodo cubierto", "Ene-2024 a Dic-2025"),
        ("Filas de la tabla de hechos", len(datos.PRODUCTOS) * 5 * 24 * 2),
        ("Aumento de precio de lista 2025", datos.INFLACION_PRECIO),
        ("Inflacion de costo de compra 2025", datos.INFLACION_COSTO),
        ("Crecimiento de unidades 2025", datos.CRECIMIENTO_UNIDADES),
        ("Participacion de e-commerce 2024", datos.MIX_ECOM[2024]),
    ]
    for i, (k, v) in enumerate(sup):
        r = 15 + i
        _f(ws, f"B{r}", k)
        _f(ws, f"C{r}", v, color=NAVY, align="center",
           fmt=PCT1 if isinstance(v, float) else None)
    _f(ws, "F15", "Dataset SIMULADO con semilla fija. Extiende el modelo "
                  "Ventas_Tech_DB del checkpoint de SQL (mismas categorias, "
                  "productos y ciudades) a 24 meses, 5 regiones y 2 canales.",
       size=9, color=GRAY, wrap=True)
    ws.merge_cells("F15:F20")

    _f(ws, "B21", "COSTO LOGISTICO SOBRE VENTAS (tasa aplicada en Datos)",
       bold=True, size=10, color=WHITE, fill=NAVY)
    for c in "CDEF":
        ws[f"{c}21"].fill = PatternFill("solid", fgColor=NAVY)
    _f(ws, "B22", "Canal", bold=True, fill=LIGHT)
    _f(ws, "C22", "Anio", bold=True, fill=LIGHT, align="center")
    _f(ws, "D22", "Trimestre", bold=True, fill=LIGHT, align="center")
    _f(ws, "E22", "Tasa", bold=True, fill=LIGHT, align="center")
    r = 23
    for (canal, anio), tasas in datos.TASA_LOGISTICA.items():
        for q, tasa in tasas.items():
            _f(ws, f"B{r}", canal)
            _f(ws, f"C{r}", anio, align="center")
            _f(ws, f"D{r}", f"Q{q}", align="center")
            _f(ws, f"E{r}", tasa, fmt=PCT2, align="center", color=NAVY)
            r += 1
    _f(ws, "F23", "El salto de e-commerce en 2025 (ultima milla y envio "
                  "bonificado) y su baja en Q3-Q4 al fijar un umbral de envio "
                  "minimo es el hecho que explica la caida y posterior "
                  "recuperacion del margen.", size=9, color=GRAY, wrap=True)
    ws.merge_cells("F23:F30")
    # la fila 22 es el encabezado; el rango de datos que lee Datos es 23..38
    return ws


# =============================================================================
#  Pivots + Calculos
# =============================================================================
def definir_pivots():
    """Una tabla dinamica por dimension y medida; todas comparten un cache."""
    specs = [
        ("PT_Periodo_Ventas", "Periodo", "Ventas", "Ventas"),
        ("PT_Periodo_Margen", "Periodo", "Margen_Bruto", "Margen"),
        ("PT_CatAnio_Ventas", "Cat_Anio", "Ventas", "Ventas"),
        ("PT_CatAnio_Margen", "Cat_Anio", "Margen_Bruto", "Margen"),
        ("PT_ProdAnio_Ventas", "Prod_Anio", "Ventas", "Ventas"),
        ("PT_ProdAnio_Margen", "Prod_Anio", "Margen_Bruto", "Margen"),
        ("PT_Anio_Ventas", "Anio", "Ventas", "Ventas"),
        ("PT_Anio_Margen", "Anio", "Margen_Bruto", "Margen"),
        ("PT_Anio_Unidades", "Anio", "Unidades", "Unidades"),
        ("PT_Anio_CostoProd", "Anio", "Costo_Producto", "CostoProd"),
        ("PT_Anio_CostoLog", "Anio", "Costo_Logistico", "CostoLog"),
        ("PT_Region_Ventas", "Region", "Ventas", "Ventas"),
    ]
    pivots, ref = [], {}
    for i, (name, rowf, meas, cap) in enumerate(specs):
        p = Pivot(name, PIV, 3, 1 + i * 3, rowf, meas, cap)
        pivots.append(p)
        ref[name] = f"{PIV}!${col_letter(p.anchor_col)}${p.anchor_row}"
    return pivots, ref


def hoja_pivots(wb):
    ws = wb.create_sheet(PIV)
    ws.sheet_state = "hidden"
    _f(ws, "A1", "Tablas dinamicas de soporte. Los segmentadores del tablero "
                 "filtran estas tablas y los graficos las leen con "
                 "GETPIVOTDATA. No editar a mano.", size=9, color=GRAY)
    return ws


def gpd(ref, campo, item, medida):
    """GETPIVOTDATA con red: si un segmentador deja el item fuera, devuelve 0."""
    it = item if isinstance(item, int) else f'"{item}"'
    return f'=IFERROR(GETPIVOTDATA("{medida}",{ref},"{campo}",{it}),0)'


def hoja_calculos(wb, ref, productos_riesgo):
    ws = wb.create_sheet(CALC)
    ws.sheet_state = "hidden"
    ws.sheet_view.showGridLines = False
    ws.column_dimensions["A"].width = 34
    for c in "BCDEFGH":
        ws.column_dimensions[c].width = 15

    _f(ws, "A1", "Capa intermedia del tablero", bold=True, size=13, color=NAVY)
    _f(ws, "A2", "Cadena del filtro:  segmentador -> tabla dinamica (Pivots) -> "
                 "GETPIVOTDATA (esta hoja) -> grafico del tablero.",
       size=9, color=GRAY, italic=True)

    # --- totales por anio --------------------------------------------------
    _f(ws, "A3", "Totales por ejercicio", bold=True, color=WHITE, fill=NAVY)
    _f(ws, "B3", 2024, bold=True, color=WHITE, fill=NAVY, align="center")
    _f(ws, "C3", 2025, bold=True, color=WHITE, fill=NAVY, align="center")
    tot = [("Ventas", "PT_Anio_Ventas", "Ventas", MONEY),
           ("Costo de producto", "PT_Anio_CostoProd", "CostoProd", MONEY),
           ("Costo logistico", "PT_Anio_CostoLog", "CostoLog", MONEY),
           ("Margen bruto", "PT_Anio_Margen", "Margen", MONEY),
           ("Unidades", "PT_Anio_Unidades", "Unidades", NUM)]
    for i, (nombre, pt, med, fmt) in enumerate(tot):
        r = R_TOT + i
        _f(ws, f"A{r}", nombre)
        for col, anio in (("B", 2024), ("C", 2025)):
            _f(ws, f"{col}{r}", gpd(ref[pt], "Anio", anio, med), fmt=fmt)
    v, cp, cl, mg, un = (R_TOT, R_TOT + 1, R_TOT + 2, R_TOT + 3, R_TOT + 4)
    derivadas = [
        ("Margen bruto %", f"=IFERROR({{c}}{mg}/{{c}}{v},0)", PCT1),
        ("Precio promedio por unidad", f"=IFERROR({{c}}{v}/{{c}}{un},0)", MONEY),
        ("Costo de producto por unidad", f"=IFERROR({{c}}{cp}/{{c}}{un},0)", MONEY),
        ("Margen unitario", f"={{c}}{R_TOT + 6}-{{c}}{R_TOT + 7}", MONEY),
        ("Costo logistico / ventas", f"=IFERROR({{c}}{cl}/{{c}}{v},0)", PCT2),
    ]
    for i, (nombre, formula, fmt) in enumerate(derivadas):
        r = R_TOT + 5 + i
        _f(ws, f"A{r}", nombre)
        for col in ("B", "C"):
            _f(ws, f"{col}{r}", formula.format(c=col), fmt=fmt)
    F = {"ventas": v, "costoprod": cp, "costolog": cl, "margen": mg,
         "unid": un, "margen_pct": R_TOT + 5, "precio": R_TOT + 6,
         "costo_unit": R_TOT + 7, "margen_unit": R_TOT + 8,
         "log_pct": R_TOT + 9}

    # --- KPIs contra meta ---------------------------------------------------
    _f(ws, f"A{R_KPI - 1}", "KPI", bold=True, color=WHITE, fill=NAVY)
    for col, tit in (("B", "Valor"), ("C", "Meta"), ("D", "Brecha"),
                     ("E", "Cumple")):
        _f(ws, f"{col}{R_KPI - 1}", tit, bold=True, color=WHITE, fill=NAVY,
           align="center")
    kpis = [
        ("Ventas 2025", f"=C{F['ventas']}", "=Parametros!$C$7", MONEY, "mayor"),
        ("Margen bruto %", f"=C{F['margen_pct']}", "=Parametros!$C$8", PCT1, "mayor"),
        ("Crecimiento interanual",
         f"=IFERROR(C{F['ventas']}/B{F['ventas']}-1,0)", "=Parametros!$C$9",
         PCT1, "mayor"),
        ("Costo logistico / ventas", f"=C{F['log_pct']}", "=Parametros!$C$10",
         PCT2, "menor"),
    ]
    for i, (nombre, val, meta, fmt, sentido) in enumerate(kpis):
        r = R_KPI + i
        _f(ws, f"A{r}", nombre)
        _f(ws, f"B{r}", val, fmt=fmt, bold=True, color=NAVY)
        _f(ws, f"C{r}", meta, fmt=fmt)
        _f(ws, f"D{r}", f"=B{r}-C{r}" if sentido == "mayor" else f"=C{r}-B{r}",
           fmt=fmt)
        _f(ws, f"E{r}", f'=IF(D{r}>=0,"EN META","BAJO META")', align="center")
    K = {"ventas": R_KPI, "margen": R_KPI + 1, "yoy": R_KPI + 2,
         "log": R_KPI + 3}

    r = R_KPI + 4
    _f(ws, f"A{r}", "Variacion del margen % (p.p.)")
    _f(ws, f"B{r}", f"=(C{F['margen_pct']}-B{F['margen_pct']})*100", fmt='0.0')
    _f(ws, f"A{r + 1}", "Variacion del costo logistico / ventas (p.p.)")
    _f(ws, f"B{r + 1}", f"=(C{F['log_pct']}-B{F['log_pct']})*100", fmt='0.0')
    D = {"margen_pp": r, "log_pp": r + 1}

    # --- puente de margen (cascada) ----------------------------------------
    _f(ws, f"A{R_CASC - 1}", "Puente de margen bruto 2024 -> 2025", bold=True,
       color=WHITE, fill=NAVY)
    for col, tit in (("B", "Efecto"), ("C", "Acumulado"), ("D", "Base"),
                     ("E", "Total"), ("F", "Suma"), ("G", "Resta")):
        _f(ws, f"{col}{R_CASC - 1}", tit, bold=True, color=WHITE, fill=NAVY,
           align="center")
    pasos = [
        ("Margen bruto 2024", f"=B{F['margen']}"),
        ("Mas unidades vendidas",
         f"=(C{F['unid']}-B{F['unid']})*B{F['margen_unit']}"),
        ("Precios de venta mas altos",
         f"=(C{F['precio']}-B{F['precio']})*C{F['unid']}"),
        ("Costo de compra mas caro",
         f"=-(C{F['costo_unit']}-B{F['costo_unit']})*C{F['unid']}"),
        ("Costo logistico mas caro", f"=-(C{F['costolog']}-B{F['costolog']})"),
        ("Margen bruto 2025", f"=C{F['margen']}"),
    ]
    for i, (etiqueta, formula) in enumerate(pasos):
        r = R_CASC + i
        _f(ws, f"A{r}", etiqueta)
        _f(ws, f"B{r}", formula, fmt=MONEY)
        es_total = i in (0, len(pasos) - 1)
        if i == 0:
            _f(ws, f"C{r}", f"=B{r}", fmt=MONEY)
        elif not es_total:
            _f(ws, f"C{r}", f"=C{r - 1}+B{r}", fmt=MONEY)
        else:
            _f(ws, f"C{r}", f"=B{r}", fmt=MONEY)
        # base invisible + los tres tramos visibles de la cascada
        _f(ws, f"D{r}", 0 if es_total else f"=MIN(C{r - 1},C{r})", fmt=MONEY)
        _f(ws, f"E{r}", f"=B{r}" if es_total else 0, fmt=MONEY)
        _f(ws, f"F{r}", 0 if es_total else f"=MAX(B{r},0)", fmt=MONEY)
        _f(ws, f"G{r}", 0 if es_total else f"=-MIN(B{r},0)", fmt=MONEY)
    r = R_CASC + len(pasos)
    _f(ws, f"A{r}", "Control: los efectos deben cerrar en cero", size=9,
       color=GRAY, italic=True)
    _f(ws, f"B{r}", f"=ROUND(SUM(B{R_CASC}:B{R_CASC + 4})-B{R_CASC + 5},2)",
       fmt=MONEY, size=9)

    # --- evolucion mensual --------------------------------------------------
    _f(ws, f"A{R_MES - 1}", "Evolucion mensual", bold=True, color=WHITE, fill=NAVY)
    for col, tit in (("B", "Ventas 2024"), ("C", "Ventas 2025"),
                     ("D", "Margen % 2025"), ("E", "Margen 2025"),
                     ("F", "Margen 2024")):
        _f(ws, f"{col}{R_MES - 1}", tit, bold=True, color=WHITE, fill=NAVY,
           align="center")
    for i, mes in enumerate(datos.MESES):
        r = R_MES + i
        _f(ws, f"A{r}", mes)
        _f(ws, f"B{r}", gpd(ref["PT_Periodo_Ventas"], "Periodo",
                            f"2024-{i + 1:02d}", "Ventas"), fmt=MONEY)
        _f(ws, f"C{r}", gpd(ref["PT_Periodo_Ventas"], "Periodo",
                            f"2025-{i + 1:02d}", "Ventas"), fmt=MONEY)
        _f(ws, f"E{r}", gpd(ref["PT_Periodo_Margen"], "Periodo",
                            f"2025-{i + 1:02d}", "Margen"), fmt=MONEY)
        _f(ws, f"F{r}", gpd(ref["PT_Periodo_Margen"], "Periodo",
                            f"2024-{i + 1:02d}", "Margen"), fmt=MONEY)
        _f(ws, f"D{r}", f"=IFERROR(E{r}/C{r},0)", fmt=PCT1)

    # --- categorias ---------------------------------------------------------
    _f(ws, f"A{R_CAT - 1}", "Margen % por categoria", bold=True, color=WHITE,
       fill=NAVY)
    for col, tit in (("B", "Margen % 2024"), ("C", "Margen % 2025"),
                     ("D", "Ventas 2025")):
        _f(ws, f"{col}{R_CAT - 1}", tit, bold=True, color=WHITE, fill=NAVY,
           align="center")
    cats = ["Computacion", "Accesorios", "Audio", "Almacenamiento"]
    for i, cat in enumerate(cats):
        r = R_CAT + i
        _f(ws, f"A{r}", cat)
        for col, anio in (("B", 2024), ("C", 2025)):
            mg = gpd(ref["PT_CatAnio_Margen"], "Cat_Anio", f"{cat} {anio}",
                     "Margen")[1:]
            vt = gpd(ref["PT_CatAnio_Ventas"], "Cat_Anio", f"{cat} {anio}",
                     "Ventas")[1:]
            _f(ws, f"{col}{r}", f"=IFERROR(({mg})/({vt}),0)", fmt=PCT1)
        _f(ws, f"D{r}", gpd(ref["PT_CatAnio_Ventas"], "Cat_Anio",
                            f"{cat} 2025", "Ventas"), fmt=MONEY)

    # --- productos (dispersion) --------------------------------------------
    _f(ws, f"A{R_PROD - 1}", "Productos 2025", bold=True, color=WHITE, fill=NAVY)
    for col, tit in (("B", "Ventas 2025"), ("C", "Crecimiento"),
                     ("D", "Margen %"), ("E", "X en meta"), ("F", "Y en meta"),
                     ("G", "X riesgo"), ("H", "Y riesgo")):
        _f(ws, f"{col}{R_PROD - 1}", tit, bold=True, color=WHITE, fill=NAVY,
           align="center")
    for i, (prod, _, _, _) in enumerate(datos.PRODUCTOS):
        r = R_PROD + i
        v25 = gpd(ref["PT_ProdAnio_Ventas"], "Prod_Anio", f"{prod} 2025",
                  "Ventas")[1:]
        v24 = gpd(ref["PT_ProdAnio_Ventas"], "Prod_Anio", f"{prod} 2024",
                  "Ventas")[1:]
        m25 = gpd(ref["PT_ProdAnio_Margen"], "Prod_Anio", f"{prod} 2025",
                  "Margen")[1:]
        _f(ws, f"A{r}", prod)
        _f(ws, f"B{r}", f"={v25}", fmt=MONEY)
        _f(ws, f"C{r}", f"=IFERROR(({v25})/({v24})-1,0)", fmt=PCT1)
        _f(ws, f"D{r}", f"=IFERROR(({m25})/({v25}),0)", fmt=PCT1)
        # series separadas: las celdas vacias son las que el grafico saltea
        col_x, col_y = ("G", "H") if prod in productos_riesgo else ("E", "F")
        _f(ws, f"{col_x}{r}", f"=C{r}", fmt=PCT1)
        _f(ws, f"{col_y}{r}", f"=D{r}", fmt=PCT1)

    # --- narrativa dinamica -------------------------------------------------
    _f(ws, f"A{R_TXT - 1}", "Textos que el tablero muestra", bold=True,
       color=WHITE, fill=NAVY)
    yoy, mgp = f"B{K['yoy']}", f"B{D['margen_pp']}"
    mg25, mg24 = f"C{F['margen_pct']}", f"B{F['margen_pct']}"
    # semestres: H1 = primeras 6 filas del bloque mensual, H2 = las otras 6
    h1_25 = (f'IFERROR(SUM(E{R_MES}:E{R_MES + 5})/'
             f'SUM(C{R_MES}:C{R_MES + 5}),0)')
    h2_25 = (f'IFERROR(SUM(E{R_MES + 6}:E{R_MES + 11})/'
             f'SUM(C{R_MES + 6}:C{R_MES + 11}),0)')
    h2_24 = (f'IFERROR(SUM(F{R_MES + 6}:F{R_MES + 11})/'
             f'SUM(B{R_MES + 6}:B{R_MES + 11}),0)')
    # tramos del puente: 29 volumen, 30 precio, 31 costo de compra, 32 logistica
    e_vol, e_pre = f"B{R_CASC + 1}", f"B{R_CASC + 2}"
    e_cpr, e_log = f"B{R_CASC + 3}", f"B{R_CASC + 4}"
    peso_log = f"IFERROR(ABS({e_log})/(ABS({e_cpr})+ABS({e_log})),0)"

    textos = [
        # (clave, formula)  -- el titular ya recoge la correccion de la nota 2
        ("titular", f'="Las ventas crecen "&TEXT({yoy},"0.0%")&'
                    f'" y el margen cae "&TEXT(ABS({mgp}),"0.0")&'
                    f'" p.p.: el costo de compra y la logistica del e-commerce'
                    f' se reparten el deterioro"'),
        ("bajada", f'="El margen bruto pasa de "&TEXT({mg24},"0.0%")&" a "&'
                   f'TEXT({mg25},"0.0%")&". La logistica explica "&'
                   f'TEXT({peso_log},"0%")&" del mayor costo y el resto es '
                   f'precio de compra: son dos causas, no una."'),
        ("g1", f'="El margen se recupera en el segundo semestre ("&'
               f'TEXT({h2_25},"0.0%")&") pero no vuelve al nivel de 2024 ("&'
               f'TEXT({h2_24},"0.0%")&")"'),
        ("g2", f'="Volumen y precio sumaron "&'
               f'TEXT({e_vol}+{e_pre},"$#,##0")&" de margen; los costos se '
               f'llevaron "&TEXT(ABS({e_cpr})+ABS({e_log}),"$#,##0")'),
        ("g3", '="Los productos que mas crecen son los que menos margen dejan"'),
        ("g4", f'="Computacion concentra la caida: pierde "&'
               f'TEXT(ABS((C{R_CAT}-B{R_CAT})*100),"0.0")&" p.p. de margen"'),
        ("nota1", f'="1) El crecimiento es real y de volumen: +"&'
                  f'TEXT({yoy},"0.0%")&" en ventas con +"&'
                  f'TEXT(IFERROR(C{F["unid"]}/B{F["unid"]}-1,0),"0.0%")&'
                  f'" en unidades. No es un espejismo de precios."'),
        ("nota2", f'="2) La caida de margen tiene dos causas, no una: el costo '
                  f'de compra se llevo "&TEXT(ABS({e_cpr}),"$#,##0")&'
                  f'" y la logistica "&TEXT(ABS({e_log}),"$#,##0")&'
                  f'". Atribuirla solo al e-commerce sobreestima su peso."'),
        ("nota3", f'="3) La correccion desde el Q3 funciona pero es parcial: el '
                  f'segundo semestre ("&TEXT({h2_25},"0.0%")&") supera al '
                  f'primero ("&TEXT({h1_25},"0.0%")&") y sigue por debajo del '
                  f'mismo periodo de 2024 ("&TEXT({h2_24},"0.0%")&")."'),
    ]
    assert [c for c, _ in textos] == CLAVES_TXT, "orden de textos desalineado"
    for i, (clave, formula) in enumerate(textos):
        r = R_TXT + i
        _f(ws, f"A{r}", clave, size=9, color=GRAY)
        _f(ws, f"B{r}", formula, size=9)
    return ws, F, K, D, T_TITULOS


# =============================================================================
#  Dashboard
# =============================================================================
def _banda(ws, rango, color):
    for fila in ws[rango]:
        for c in fila:
            c.fill = PatternFill("solid", fgColor=color)


def _caja(ws, rango, color=GRAY):
    lado = Side(style="thin", color=color)
    celdas = ws[rango]
    n_f, n_c = len(celdas), len(celdas[0])
    for i, fila in enumerate(celdas):
        for j, c in enumerate(fila):
            c.border = Border(
                top=lado if i == 0 else None,
                bottom=lado if i == n_f - 1 else None,
                left=lado if j == 0 else None,
                right=lado if j == n_c - 1 else None)


def _titulo_dinamico(chart, ref_celda):
    t = Title()
    t.tx = Text(strRef=StrRef(ref_celda))
    chart.title = t


def hoja_dashboard(wb, F, K, D, T):
    ws = wb.create_sheet(DASH, 0)
    ws.sheet_view.showGridLines = False
    ws.sheet_view.zoomScale = 90
    # que imprima como se ve: una hoja de ancho, apaisada
    ws.page_setup.orientation = "landscape"
    ws.page_setup.fitToWidth = 1
    ws.page_setup.fitToHeight = 0
    ws.sheet_properties.pageSetUpPr.fitToPage = True
    ws.print_area = "A1:R75"

    ws.column_dimensions["A"].width = 2.2
    for i in range(2, 18):                       # B..Q
        ws.column_dimensions[get_column_letter(i)].width = 9.4
    ws.column_dimensions["R"].width = 2.2

    alturas = {1: 6, 2: 26, 3: 15, 4: 8, 5: 6, 6: 22, 7: 16, 8: 8,
               9: 14, 10: 24, 11: 12, 12: 14, 13: 10, 14: 16}
    for r, h in alturas.items():
        ws.row_dimensions[r].height = h
    for r in range(15, 24):
        ws.row_dimensions[r].height = 17

    # --- cabecera -----------------------------------------------------------
    _banda(ws, "A1:R4", NAVY)
    ws.merge_cells("B2:K2")
    _f(ws, "B2", "TechStore  |  Sistema de Inteligencia de Negocios",
       bold=True, size=17, color=WHITE, align="left")
    ws.merge_cells("L2:Q3")
    _f(ws, "L2", f"Actualizado: {dt.date.today():%d/%m/%Y}   ·   Moneda: USD",
       size=9, color="C7CEDB", align="right")
    ws.merge_cells("B3:K3")
    _f(ws, "B3", "Reporte ejecutivo de ventas y margen  ·  Ejercicios 2024-2025",
       size=10, color="C7CEDB", align="left")

    # --- titular narrativo --------------------------------------------------
    ws.merge_cells("B6:Q6")
    _f(ws, "B6", f"={T['titular']}", bold=True, size=13, color=NAVY,
       align="left")
    ws.merge_cells("B7:Q7")
    _f(ws, "B7", f"={T['bajada']}", size=9.5, color=GRAY, italic=True,
       align="left")

    # --- tarjetas de KPI ----------------------------------------------------
    tarjetas = [
        ("B", "E", "VENTAS 2025", f"={CALC}!$B${K['ventas']}", MONEY,
         f'="vs meta "&TEXT({CALC}!$C${K["ventas"]},"$#,##0")&"   "&'
         f'TEXT({CALC}!$D${K["ventas"]},"+#,##0;-#,##0")&" USD"'),
        ("F", "I", "MARGEN BRUTO", f"={CALC}!$B${K['margen']}", PCT1,
         f'="vs meta "&TEXT({CALC}!$C${K["margen"]},"0.0%")&"   "&'
         f'TEXT({CALC}!$B${D["margen_pp"]},"+0.0;-0.0")&" p.p. vs 2024"'),
        ("J", "M", "CRECIMIENTO INTERANUAL", f"={CALC}!$B${K['yoy']}", PCT1,
         f'="vs meta "&TEXT({CALC}!$C${K["yoy"]},"0.0%")&"   "&'
         f'TEXT({CALC}!$D${K["yoy"]}*100,"+0.0;-0.0")&" p.p."'),
        ("N", "Q", "COSTO LOGISTICO / VENTAS", f"={CALC}!$B${K['log']}", PCT2,
         f'="techo "&TEXT({CALC}!$C${K["log"]},"0.00%")&"   "&'
         f'TEXT({CALC}!$B${D["log_pp"]},"+0.00;-0.00")&" p.p. vs 2024"'),
    ]
    brechas = [K["ventas"], K["margen"], K["yoy"], K["log"]]
    for (c1, c2, etiqueta, valor, fmt, pie), fila_kpi in zip(tarjetas, brechas):
        ws.merge_cells(f"{c1}9:{c2}9")
        _f(ws, f"{c1}9", etiqueta, bold=True, size=8, color=GRAY, align="left")
        ws.merge_cells(f"{c1}10:{c2}11")
        _f(ws, f"{c1}10", valor, bold=True, size=20, color=NAVY, fmt=fmt,
           align="left")
        ws.merge_cells(f"{c1}12:{c2}12")
        _f(ws, f"{c1}12", pie, size=8.5, color=GRAY, align="left")
        _banda(ws, f"{c1}9:{c2}12", LIGHT)
        _caja(ws, f"{c1}9:{c2}12", "DCE0E6")
        # espejo local de la brecha: el formato condicional lo lee sin salir
        # de la hoja (mas robusto que una referencia entre hojas)
        _f(ws, f"{c1}13", f"={CALC}!$D${fila_kpi}", size=8, color=WHITE)
        # por debajo de la meta la tarjeta entera pasa a ambar
        ws.conditional_formatting.add(
            f"{c1}9:{c2}12",
            FormulaRule(formula=[f"${c1}$13<0"],
                        font=Font(name=FUENTE, color=AMBER, b=True),
                        fill=PatternFill(start_color="FDF3E3",
                                         end_color="FDF3E3", fill_type="solid")))

    # --- segmentadores ------------------------------------------------------
    ws.merge_cells("B14:Q14")
    _f(ws, "B14", "FILTROS  ·  los cuatro segmentadores afectan a la vez a los "
                  "cuatro graficos y a las cuatro tarjetas",
       bold=True, size=8.5, color=GRAY, align="left")
    _banda(ws, "B15:Q23", WHITE)

    # --- graficos -----------------------------------------------------------
    for r in range(24, 76):
        ws.row_dimensions[r].height = 15
    ws.merge_cells("B25:Q25")
    _f(ws, "B25", "QUE PASO Y POR QUE", bold=True, size=9, color=WHITE,
       fill=NAVY, align="left")
    ws.merge_cells("B47:Q47")
    _f(ws, "B47", "DONDE MIRAR AHORA", bold=True, size=9, color=WHITE,
       fill=NAVY, align="left")
    return ws


def graficos(wb, ws, T, productos_riesgo):
    calc = wb[CALC]

    def serie_color(s, color, linea=False):
        s.graphicalProperties = GraphicalProperties(
            solidFill=color)
        if linea:
            s.graphicalProperties.line = LineProperties(solidFill=color, w=22000)
            s.graphicalProperties.noFill = True
        else:
            s.graphicalProperties.line = LineProperties(noFill=True)

    # 1 - evolucion mensual: columnas 2024/2025 + linea de margen %
    barras = BarChart()
    barras.type = "col"
    barras.grouping = "clustered"
    barras.gapWidth = 60
    barras.overlap = -10
    datos_ref = Reference(calc, min_col=2, max_col=3, min_row=R_MES - 1,
                          max_row=R_MES + 11)
    barras.add_data(datos_ref, titles_from_data=True)
    barras.set_categories(Reference(calc, min_col=1, min_row=R_MES,
                                    max_row=R_MES + 11))
    serie_color(barras.series[0], GRAY)
    serie_color(barras.series[1], NAVY)
    barras.y_axis.numFmt = MONEY_K
    barras.y_axis.title = None

    linea = LineChart()
    linea.add_data(Reference(calc, min_col=4, min_row=R_MES - 1,
                             max_row=R_MES + 11), titles_from_data=True)
    linea.y_axis.axId = 200
    linea.y_axis.numFmt = PCT1
    linea.y_axis.majorGridlines = None
    linea.y_axis.crosses = "max"
    serie_color(linea.series[0], AMBER, linea=True)
    linea.series[0].smooth = False
    linea.series[0].marker = Marker(symbol="circle", size=5)
    barras += linea
    _titulo_dinamico(barras, T["g1"])
    barras.height, barras.width = 9.9, 13.2
    barras.legend.position = "b"
    ws.add_chart(barras, "B26")

    # 2 - cascada: apilado con la base invisible
    casc = BarChart()
    casc.type = "col"
    casc.grouping = "stacked"
    casc.overlap = 100
    casc.gapWidth = 45
    casc.add_data(Reference(calc, min_col=4, max_col=7, min_row=R_CASC - 1,
                            max_row=R_CASC + 5), titles_from_data=True)
    casc.set_categories(Reference(calc, min_col=1, min_row=R_CASC,
                                  max_row=R_CASC + 5))
    base, total, suma, resta = casc.series
    base.graphicalProperties = GraphicalProperties(
        noFill=True)
    base.graphicalProperties.line = LineProperties(noFill=True)
    serie_color(total, NAVY)
    serie_color(suma, GRAY)
    serie_color(resta, AMBER)
    casc.y_axis.numFmt = MONEY_K
    _titulo_dinamico(casc, T["g2"])
    casc.height, casc.width = 9.9, 13.2
    casc.legend.position = "b"
    ws.add_chart(casc, "J26")

    # 3 - dispersion de productos: crecimiento contra margen
    disp = ScatterChart()
    disp.x_axis.numFmt = PCT1
    disp.y_axis.numFmt = PCT1
    disp.x_axis.title = "Crecimiento de ventas 2025 vs 2024"
    disp.y_axis.title = "Margen bruto % 2025"
    disp.style = None
    for col_x, col_y, nombre, color in (
            (5, 6, "En meta de margen", NAVY),
            (7, 8, "Zona de riesgo", AMBER)):
        s = Series(Reference(calc, min_col=col_y, min_row=R_PROD,
                             max_row=R_PROD + 9),
                   Reference(calc, min_col=col_x, min_row=R_PROD,
                             max_row=R_PROD + 9),
                   title=nombre)
        s.marker = Marker(symbol="circle", size=9)
        s.marker.graphicalProperties = GraphicalProperties(
            solidFill=color)
        s.marker.graphicalProperties.line = LineProperties(solidFill=color)
        s.graphicalProperties = GraphicalProperties()
        s.graphicalProperties.line = LineProperties(noFill=True)
        disp.series.append(s)
    _titulo_dinamico(disp, T["g3"])
    disp.height, disp.width = 9.9, 13.2
    disp.legend.position = "b"
    ws.add_chart(disp, "B48")

    # 4 - margen % por categoria, 2024 contra 2025
    cat = BarChart()
    cat.type = "bar"
    cat.grouping = "clustered"
    cat.gapWidth = 60
    cat.add_data(Reference(calc, min_col=2, max_col=3, min_row=R_CAT - 1,
                           max_row=R_CAT + 3), titles_from_data=True)
    cat.set_categories(Reference(calc, min_col=1, min_row=R_CAT,
                                 max_row=R_CAT + 3))
    serie_color(cat.series[0], GRAY)
    serie_color(cat.series[1], NAVY)
    cat.x_axis.numFmt = PCT1
    _titulo_dinamico(cat, T["g4"])
    cat.height, cat.width = 9.9, 13.2
    cat.legend.position = "b"
    ws.add_chart(cat, "J48")

    # --- notas de analisis --------------------------------------------------
    ws.merge_cells("B69:Q69")
    _f(ws, "B69", "NOTAS DE ANALISIS  ·  conclusiones propias, sometidas a "
                  "revision critica con IA (ver hoja Analisis_IA)",
       bold=True, size=9, color=WHITE, fill=NAVY, align="left")
    for i, clave in enumerate(("nota1", "nota2", "nota3")):
        r = 70 + i
        ws.row_dimensions[r].height = 15
        ws.merge_cells(f"B{r}:Q{r}")
        _f(ws, f"B{r}", f"={T[clave]}", size=9, color=INK, align="left")
    ws.merge_cells("B74:Q74")
    _f(ws, "B74", "Fuente: Ventas_Tech_DB (dataset simulado, 2.400 operaciones). "
                  "Supuestos y metas en la hoja Parametros. "
                  "Elaborado por Lleyton Murphy - Coderhouse Data Analyst, "
                  "Modulo 9.", size=8, color=GRAY, align="left")


# =============================================================================
def hoja_guia(wb):
    ws = wb.create_sheet("Guia_de_uso")
    ws.sheet_view.showGridLines = False
    ws.column_dimensions["A"].width = 2
    ws.column_dimensions["B"].width = 26
    ws.column_dimensions["C"].width = 96
    _f(ws, "B2", "Guia de uso del tablero", bold=True, size=15, color=NAVY)

    bloques = [
        ("Como se lee", "De arriba hacia abajo: el titular dice que paso, las "
                        "cuatro tarjetas lo cuantifican contra la meta, los dos "
                        "graficos de arriba explican por que, y los dos de abajo "
                        "indican donde actuar."),
        ("Los segmentadores", "Trimestre, Categoria, Region y Canal. Los cuatro "
                              "estan conectados al mismo cache de tablas "
                              "dinamicas, asi que cualquier seleccion recalcula "
                              "a la vez las tarjetas y los cuatro graficos."),
        ("Por que Trimestre y no Ano",
         "El filtro de tiempo es el trimestre justamente para no romper la "
         "comparacion interanual: al elegir Q1 se comparan Q1-2024 contra "
         "Q1-2025. Un filtro de ano dejaria sin sentido el KPI de crecimiento."),
        ("Como viaja el filtro", "Segmentador -> tabla dinamica (hoja Pivots) -> "
                                 "GETPIVOTDATA (hoja Calculos) -> grafico. "
                                 "Pivots y Calculos estan ocultas a proposito: "
                                 "son soporte, no lectura."),
        ("Si los numeros no se mueven",
         "Datos > Actualizar todo. El libro fuerza el recalculo al abrir, pero "
         "algunos visores de terceros no ejecutan tablas dinamicas."),
        ("Paleta", "Tres colores: azul institucional (valor principal y ano "
                   "corriente), gris (contexto y ano anterior) y ambar "
                   "(alerta y efectos que restan). Nada mas."),
        ("Origen de los datos", "Dataset simulado y determinista que extiende el "
                                "modelo Ventas_Tech_DB del checkpoint de SQL. "
                                "Todos los supuestos estan en la hoja "
                                "Parametros."),
    ]
    r = 4
    for titulo, texto in bloques:
        _f(ws, f"B{r}", titulo, bold=True, color=NAVY, align="left")
        _f(ws, f"C{r}", texto, wrap=True, align="left")
        ws.row_dimensions[r].height = 42
        r += 1
    return ws


ANALISIS_IA = [
    {
        "titulo": "Conclusion 1 - El crecimiento de 2025 es solido",
        "propia": "Las ventas pasan de USD 4.726.315 a USD 5.566.046 (+17,8%) y "
                  "superan la meta de +15%. El negocio crecio.",
        "prompt": "Actua como analista senior esceptico. Mi conclusion es: "
                  "\"las ventas de una cadena de retail tecnologico crecieron "
                  "17,8% interanual, superando la meta de 15%, entonces el "
                  "negocio crecio\". Antes de darme la razon, decime que "
                  "variables omitidas podrian hacer que esa lectura sea "
                  "enganosa y que dato concreto tendria que mirar para "
                  "descartar cada una.",
        "objecion": "Un alza en pesos no prueba crecimiento del negocio. Puede "
                    "ser (a) puro efecto precio, (b) cambio de mix hacia "
                    "productos mas caros, o (c) adelanto de demanda por "
                    "descuentos agresivos. Pidio contrastar contra unidades.",
        "verificacion": "Las unidades pasan de 25.681 a 29.379 (+14,4%) contra "
                        "+17,8% en pesos: la mayor parte del crecimiento es "
                        "volumen y solo unos 3 p.p. vienen de precio. La "
                        "objecion (a) queda descartada; la (c) queda en pie "
                        "para Computacion, que tuvo 7% de descuento en el H1.",
        "refinamiento": "La nota 1 del tablero informa ventas Y unidades "
                        "juntas, para que nadie lea el +17,8% como volumen "
                        "puro. La conclusion se mantiene, con el matiz.",
    },
    {
        "titulo": "Conclusion 2 - El margen cae por el costo logistico",
        "propia": "El margen bruto cae de 35,8% a 33,3% (-2,5 p.p.) mientras el "
                  "costo logistico sobre ventas sube de 2,18% a 3,30%. La "
                  "logistica del e-commerce se comio la mejora.",
        "prompt": "Hace de abogado del diablo con esta conclusion: \"el margen "
                  "cayo 2,5 p.p. y en el mismo periodo el costo logistico paso "
                  "de 2,18% a 3,30% de las ventas, por el crecimiento del "
                  "e-commerce; entonces la logistica explica la caida del "
                  "margen\". Buscame el error de logica y decime como "
                  "descomponer la variacion para probar o refutar la causa.",
        "objecion": "Falacia de causa unica: el costo logistico subio 1,12 p.p. "
                    "sobre ventas, o sea menos de la mitad de los 2,5 p.p. "
                    "perdidos. Que dos cosas se muevan juntas no vuelve a una "
                    "la causa de la otra. Propuso un puente de margen que "
                    "separe volumen, precio, costo de compra y logistica.",
        "verificacion": "El puente cierra exacto y desmiente la atribucion: "
                        "volumen +USD 258.146 y precio +USD 159.153 contra "
                        "costo de compra -USD 172.525 y logistica -USD 80.921. "
                        "La logistica explica el 32% del mayor costo, no el "
                        "total: el costo de compra pesa mas.",
        "refinamiento": "Cambio estructural del tablero. (1) El titular "
                        "dinamico dejo de decir \"el costo logistico se llevo "
                        "la mejora\" y ahora reparte la causa. (2) Se agrego el "
                        "grafico de cascada justamente para que la "
                        "descomposicion se vea, en vez de afirmarla. (3) La "
                        "nota 2 explicita las dos causas con sus montos.",
    },
    {
        "titulo": "Conclusion 3 - La correccion del Q3 funciono",
        "propia": "El margen del segundo semestre (34,5%) supera al del primero "
                  "(31,8%) despues de fijar el umbral de envio minimo en Q3. "
                  "La medida funciono.",
        "prompt": "Mi conclusion es: \"tras aplicar en Q3 un umbral de envio "
                  "minimo, el margen del segundo semestre (34,5%) supera al "
                  "del primero (31,8%), asi que la medida funciono\". Decime "
                  "que sesgos hay en esa lectura y con que comparacion "
                  "deberia reemplazarla.",
        "objecion": "Dos problemas. Uno, causalidad asumida: el Q4 es el "
                    "trimestre mas fuerte del ano (Nov-Dic) y un mix "
                    "estacional distinto puede subir el margen sin que la "
                    "medida haya hecho nada. Dos, comparacion contra el "
                    "periodo equivocado: hay que comparar H2-2025 contra "
                    "H2-2024, no contra H1-2025.",
        "verificacion": "H2-2024 cerro en 35,7% y H2-2025 en 34,5%: contra el "
                        "mismo periodo del ano anterior el margen sigue 1,2 "
                        "p.p. abajo. La recuperacion es real pero parcial, y "
                        "parte de la mejora es estacional.",
        "refinamiento": "El grafico mensual muestra 2024 y 2025 juntos (no 2025 "
                        "solo), que es lo unico que deja ver la brecha "
                        "estacional. El titulo dinamico y la nota 3 dicen "
                        "\"sin volver al nivel de 2024\" en lugar de "
                        "\"se recupero\".",
    },
]


def hoja_analisis_ia(wb):
    ws = wb.create_sheet("Analisis_IA")
    ws.sheet_view.showGridLines = False
    ws.column_dimensions["A"].width = 2
    ws.column_dimensions["B"].width = 22
    ws.column_dimensions["C"].width = 104

    _f(ws, "B2", "Analisis asistido por IA - registro del proceso", bold=True,
       size=15, color=NAVY)
    _f(ws, "B3", "La IA se uso como auditor cognitivo, no como redactor: "
                 "primero se escribieron las conclusiones a mano, despues se "
                 "le pidio que las atacara, y recien al final se corrigio el "
                 "tablero. El registro completo, con los prompts textuales, "
                 "esta en el documento anexo.", size=9, color=GRAY, italic=True)
    ws.merge_cells("B3:C3")
    ws.row_dimensions[3].height = 26
    ws["B3"].alignment = Alignment(wrap_text=True, vertical="top")

    etiquetas = [("propia", "1. Conclusion propia"),
                 ("prompt", "2. Prompt de contraste"),
                 ("objecion", "3. Objecion de la IA"),
                 ("verificacion", "4. Verificacion sobre los datos"),
                 ("refinamiento", "5. Cambio en el tablero")]
    r = 5
    for bloque in ANALISIS_IA:
        _f(ws, f"B{r}", bloque["titulo"], bold=True, size=11, color=WHITE,
           fill=NAVY, align="left")
        ws[f"C{r}"].fill = PatternFill("solid", fgColor=NAVY)
        ws.row_dimensions[r].height = 20
        r += 1
        for clave, etiqueta in etiquetas:
            _f(ws, f"B{r}", etiqueta, bold=True, size=9, color=NAVY,
               align="left")
            ws[f"B{r}"].alignment = Alignment(vertical="top", horizontal="left")
            _f(ws, f"C{r}", bloque[clave], size=9.5, wrap=True, align="left")
            ws[f"C{r}"].alignment = Alignment(wrap_text=True, vertical="top")
            ws.row_dimensions[r].height = 15 * (1 + len(bloque[clave]) // 105)
            r += 1
        r += 1

    _f(ws, f"B{r}", "Que cambio el proceso", bold=True, size=11, color=WHITE,
       fill=NAVY, align="left")
    ws[f"C{r}"].fill = PatternFill("solid", fgColor=NAVY)
    r += 1
    _f(ws, f"B{r}", "Saldo", bold=True, size=9, color=NAVY, align="left")
    _f(ws, f"C{r}", "De las tres conclusiones, una se confirmo con un matiz "
                    "(la 1), una se refuto en su atribucion causal (la 2) y "
                    "una se corrigio por comparar contra el periodo equivocado "
                    "(la 3). El grafico de cascada y la comparativa 2024-2025 "
                    "del grafico mensual existen por esas objeciones: no "
                    "estaban en el diseno original.",
       size=9.5, wrap=True, align="left")
    ws[f"C{r}"].alignment = Alignment(wrap_text=True, vertical="top")
    ws.row_dimensions[r].height = 60
    return ws


def main():
    columnas, filas = datos.generar()
    ix = {c: i for i, c in enumerate(columnas)}

    # productos por debajo del corte de margen -> serie ambar de la dispersion
    agg = {}
    for f in filas:
        if f[ix["Anio"]] == 2025:
            k = f[ix["Producto"]]
            v, m = agg.get(k, (0.0, 0.0))
            agg[k] = (v + f[ix["Ventas"]], m + f[ix["Margen_Bruto"]])
    corte = 0.33
    riesgo = {p for p, (v, m) in agg.items() if m / v < corte}

    wb = openpyxl.Workbook()
    wb.remove(wb.active)

    pivots, ref = definir_pivots()
    hoja_pivots(wb)
    _, F, K, D, T = hoja_calculos(wb, ref, riesgo)
    ws_dash = hoja_dashboard(wb, F, K, D, T)
    graficos(wb, ws_dash, T, riesgo)
    hoja_analisis_ia(wb)
    hoja_datos(wb, columnas, filas)
    hoja_parametros(wb)
    hoja_guia(wb)

    # el tablero primero; las hojas de soporte, detras y ocultas
    orden = [DASH, "Analisis_IA", "Datos", "Parametros", "Guia_de_uso",
             CALC, PIV]
    for pos, nombre in enumerate(orden):
        wb.move_sheet(nombre, offset=pos - wb.sheetnames.index(nombre))
    wb.active = 0

    out = os.path.join(os.path.dirname(os.path.abspath(__file__)), "libro.xlsx")
    wb.save(out)
    print("libro base:", out)
    print("productos en zona de riesgo:", sorted(riesgo))
    return out, columnas, filas, pivots


if __name__ == "__main__":
    main()
