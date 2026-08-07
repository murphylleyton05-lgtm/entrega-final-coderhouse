"""
Construye el libro base de la pre-entrega del Modulo 8, limpio y sin partes danadas.

El libro que sale de aca contiene SOLO la documentacion estatica (texto):
    LEEME, Codigo_M, Medidas_DAX, Modelo_Estrella

Todo lo que necesita el motor de Excel -- consultas de Power Query, modelo de
Power Pivot, relaciones, medidas, estadistica descriptiva, escenarios y resumen
de escenarios -- lo crea la macro MaterializarModelo al ejecutarse dentro de Excel.
"""

from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side

AZUL = "1F3864"
AZUL_CLARO = "D9E2F3"
GRIS = "F2F2F2"
VERDE = "375623"

TITULO = Font(name="Calibri", size=14, bold=True, color="FFFFFF")
SUBTITULO = Font(name="Calibri", size=11, bold=True, color=AZUL)
NORMAL = Font(name="Calibri", size=11)
CODIGO = Font(name="Consolas", size=10)
CODIGO_COMENTARIO = Font(name="Consolas", size=10, color="2E7D32", italic=True)
CABECERA = Font(name="Calibri", size=11, bold=True, color="FFFFFF")

FONDO_TITULO = PatternFill("solid", fgColor=AZUL)
FONDO_CABECERA = PatternFill("solid", fgColor=AZUL)
FONDO_SUAVE = PatternFill("solid", fgColor=AZUL_CLARO)
FONDO_GRIS = PatternFill("solid", fgColor=GRIS)

BORDE = Border(*[Side(style="thin", color="BFBFBF")] * 4)
WRAP = Alignment(wrap_text=True, vertical="top")
TOP = Alignment(vertical="top")


def banda_titulo(ws, texto, ancho=8):
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=ancho)
    c = ws.cell(row=1, column=1, value=texto)
    c.font = TITULO
    c.fill = FONDO_TITULO
    c.alignment = Alignment(vertical="center", indent=1)
    ws.row_dimensions[1].height = 28


def bloque_codigo(ws, fila, titulo, lineas):
    c = ws.cell(row=fila, column=1, value=titulo)
    c.font = SUBTITULO
    c.fill = FONDO_SUAVE
    fila += 1
    for linea in lineas:
        c = ws.cell(row=fila, column=1, value=linea)
        c.font = CODIGO_COMENTARIO if linea.lstrip().startswith("//") else CODIGO
        c.alignment = TOP
        fila += 1
    return fila + 2


# ---------------------------------------------------------------------------
# CODIGO M  (el mismo que carga la macro; aca la ruta va como marcador)
# ---------------------------------------------------------------------------
RUTA_CSV = r"<ruta del libro>\Fuentes\Ventas_Historico.csv"
RUTA_XLSX = r"<ruta del libro>\Fuentes\Maestras.xlsx"

M_HECHOS = f'''let
    // ORIGEN: archivo transaccional .csv ubicado en la subcarpeta \\Fuentes del libro.
    // La ruta la reescribe la macro MaterializarModelo al ejecutarse, de modo que el
    // libro se puede mover de carpeta y solo hay que volver a correrla.
    Origen = Csv.Document(File.Contents("{RUTA_CSV}"), [Delimiter=",", Columns=9, Encoding=65001, QuoteStyle=QuoteStyle.Csv]),

    EncabezadosPromovidos = Table.PromoteHeaders(Origen, [PromoteAllScalars=true]),

    // Tipado explicito con Culture="en-US": el .csv usa punto decimal y fecha ISO,
    // asi el tipado no depende de la configuracion regional del equipo que abra el libro.
    TiposDefinidos = Table.TransformColumnTypes(EncabezadosPromovidos,{{
        {{"ID_Venta", type text}}, {{"Fecha", type date}}, {{"ID_Cliente", type text}},
        {{"ID_Producto", type text}}, {{"Canal", type text}}, {{"Cantidad", Int64.Type}},
        {{"Precio_Unitario", type number}}, {{"Descuento", type number}}, {{"Costo_Unitario", type number}}
    }}, "en-US"),

    // PASO RENOMBRADO MANUALMENTE (era "Filas filtradas"): consolida en un unico
    // Table.SelectRows lo que la interfaz genera como tres pasos encadenados
    // (quitar nulos, quitar vacios, filtrar cantidad). Un solo recorrido de la tabla.
    VentasFiltradas = Table.SelectRows(TiposDefinidos, each
        [Fecha] <> null and
        [ID_Cliente] <> null and [ID_Cliente] <> "" and
        [ID_Producto] <> null and [ID_Producto] <> "" and
        [Cantidad] <> null and [Cantidad] > 0),

    // METRICAS DE FILA RESUELTAS EN EL ETL, NO EN DAX.
    // Materializarlas aca evita iteradores (SUMX/FILTER) sobre la tabla de hechos:
    // en el modelo alcanza con SUM() sobre columnas ya calculadas.
    VentaBruta = Table.AddColumn(VentasFiltradas, "Venta_Bruta", each [Cantidad] * [Precio_Unitario], type number),
    VentaNeta = Table.AddColumn(VentaBruta, "Venta_Neta", each [Venta_Bruta] * (1 - [Descuento]), type number),
    CostoTotal = Table.AddColumn(VentaNeta, "Costo_Total", each [Cantidad] * [Costo_Unitario], type number),
    MargenBruto = Table.AddColumn(CostoTotal, "Margen_Bruto", each [Venta_Neta] - [Costo_Total], type number),

    // REDUCCION DE HUELLA: se descartan las columnas que ya quedaron absorbidas en las
    // metricas (Precio_Unitario, Descuento, Costo_Unitario, Venta_Bruta). Menos columnas
    // de alta cardinalidad = menor tamanio del modelo en memoria.
    ColumnasFinales = Table.SelectColumns(MargenBruto,{{
        "ID_Venta", "Fecha", "ID_Cliente", "ID_Producto", "Canal",
        "Cantidad", "Venta_Neta", "Costo_Total", "Margen_Bruto"}})
in
    ColumnasFinales'''

M_CLIENTES = f'''let
    // ORIGEN: libro local de tablas maestras (.xlsx), segunda fuente heterogenea del flujo.
    Origen = Excel.Workbook(File.Contents("{RUTA_XLSX}"), null, true),
    TablaClientes = Origen{{[Item="Tbl_Clientes", Kind="Table"]}}[Data],

    TiposDefinidos = Table.TransformColumnTypes(TablaClientes,{{
        {{"ID_Cliente", type text}}, {{"Cliente", type text}}, {{"Segmento", type text}},
        {{"Provincia", type text}}, {{"Pais", type text}}}}, "en-US"),

    ClavesValidas = Table.SelectRows(TiposDefinidos, each [ID_Cliente] <> null and [ID_Cliente] <> ""),

    // Table.Distinct sobre la PK: sin unicidad garantizada Power Pivot rechaza la
    // relacion 1:N y la degrada a N:N. Este paso es el que sostiene el lado "1".
    ClientesUnicos = Table.Distinct(ClavesValidas, {{"ID_Cliente"}})
in
    ClientesUnicos'''

M_PRODUCTOS = f'''let
    Origen = Excel.Workbook(File.Contents("{RUTA_XLSX}"), null, true),
    TablaProductos = Origen{{[Item="Tbl_Productos", Kind="Table"]}}[Data],

    TiposDefinidos = Table.TransformColumnTypes(TablaProductos,{{
        {{"ID_Producto", type text}}, {{"Producto", type text}}, {{"Categoria", type text}},
        {{"Marca", type text}}, {{"Precio_Lista", type number}}}}, "en-US"),

    ClavesValidas = Table.SelectRows(TiposDefinidos, each [ID_Producto] <> null and [ID_Producto] <> ""),

    // Misma logica que en Dim_Clientes: unicidad de PK garantizada en el ETL.
    ProductosUnicos = Table.Distinct(ClavesValidas, {{"ID_Producto"}})
in
    ProductosUnicos'''

M_CALENDARIO = '''let
    // DIMENSION CALENDARIO generada integramente en M (no hay archivo de origen).
    // El rango cubre el historico completo del .csv transaccional.
    FechaInicio = #date(2023, 1, 1),
    FechaFin = #date(2025, 12, 31),

    // Una fila por dia y sin huecos: requisito para marcarla como Tabla de fechas
    // y para que funcione la inteligencia de tiempo (DATEADD / SAMEPERIODLASTYEAR).
    ListaDias = List.Dates(FechaInicio, Duration.Days(FechaFin - FechaInicio) + 1, #duration(1,0,0,0)),
    TablaFechas = Table.FromList(ListaDias, Splitter.SplitByNothing(), {"Fecha"}),
    FechaTipada = Table.TransformColumnTypes(TablaFechas,{{"Fecha", type date}}, "en-US"),

    ConAnio = Table.AddColumn(FechaTipada, "Anio", each Date.Year([Fecha]), Int64.Type),
    ConTrimestre = Table.AddColumn(ConAnio, "Trimestre", each "T" & Text.From(Date.QuarterOfYear([Fecha])), type text),
    ConMes = Table.AddColumn(ConTrimestre, "Mes", each Date.Month([Fecha]), Int64.Type),
    ConNombreMes = Table.AddColumn(ConMes, "Nombre_Mes", each Date.MonthName([Fecha], "es-AR"), type text),
    ConAnioMes = Table.AddColumn(ConNombreMes, "Anio_Mes", each Text.From(Date.Year([Fecha])) & "-" & Text.PadStart(Text.From(Date.Month([Fecha])), 2, "0"), type text),
    ConDiaSemana = Table.AddColumn(ConAnioMes, "Dia_Semana", each Date.DayOfWeekName([Fecha], "es-AR"), type text),

    // Consistencia con el resto de las dimensiones: PK unica explicita.
    CalendarioUnico = Table.Distinct(ConDiaSemana, {"Fecha"})
in
    CalendarioUnico'''

MEDIDAS = [
    ("Total Ventas", "SUM(Hechos_Ventas[Venta_Neta])", "No",
     "Suma directa sobre la columna que Power Query dejo materializada. No hay iteradores: "
     "el motor VertiPaq resuelve la agregacion sobre la columna comprimida."),
    ("Total Costo", "SUM(Hechos_Ventas[Costo_Total])", "No",
     "Agregacion simple sobre columna pre-calculada en el ETL."),
    ("Margen Bruto", "SUM(Hechos_Ventas[Margen_Bruto])", "No",
     "Reemplaza a SUMX(Hechos_Ventas, [Venta_Neta] - [Costo_Total]). Al estar la resta "
     "resuelta fila a fila en Power Query, el iterador deja de ser necesario."),
    ("Unidades Vendidas", "SUM(Hechos_Ventas[Cantidad])", "No",
     "Volumen fisico vendido."),
    ("% Margen Bruto",
     "VAR VentasTotales = [Total Ventas]\n"
     "VAR MargenTotal = [Margen Bruto]\n"
     "RETURN\n"
     "    DIVIDE(MargenTotal, VentasTotales, 0)", "Si",
     "Cada agregado se evalua UNA sola vez y se reutiliza. DIVIDE devuelve 0 en lugar de "
     "error cuando el denominador es cero."),
    ("Ticket Promedio",
     "VAR VentasTotales = [Total Ventas]\n"
     "VAR CantidadTickets = DISTINCTCOUNT(Hechos_Ventas[ID_Venta])\n"
     "RETURN\n"
     "    DIVIDE(VentasTotales, CantidadTickets, 0)", "Si",
     "El DISTINCTCOUNT -- que es la operacion cara -- se materializa una vez en una "
     "variable en lugar de recalcularse dentro del cociente."),
    ("Ventas Anio Anterior",
     "VAR VentasPrevias =\n"
     "    CALCULATE([Total Ventas], DATEADD(Dim_Calendario[Fecha], -1, YEAR))\n"
     "RETURN\n"
     "    IF(ISBLANK(VentasPrevias), BLANK(), VentasPrevias)", "Si",
     "Inteligencia de tiempo apoyada en la relacion con Dim_Calendario. Requiere que la "
     "dimension este marcada como Tabla de fechas."),
    ("Var % vs Anio Anterior",
     "VAR VentasActuales = [Total Ventas]\n"
     "VAR VentasPrevias = [Ventas Anio Anterior]\n"
     "RETURN\n"
     "    IF(\n"
     "        ISBLANK(VentasPrevias) || VentasPrevias = 0,\n"
     "        BLANK(),\n"
     "        DIVIDE(VentasActuales - VentasPrevias, VentasPrevias)\n"
     "    )", "Si",
     "Guarda ambos periodos en variables: sin VAR, [Ventas Anio Anterior] se evaluaria "
     "tres veces (en el ISBLANK, en la comparacion y en el DIVIDE)."),
]

RELACIONES = [
    ("Dim_Calendario", "Fecha", "Hechos_Ventas", "Fecha", "1:N", "Unica (Dim -> Hechos)"),
    ("Dim_Clientes", "ID_Cliente", "Hechos_Ventas", "ID_Cliente", "1:N", "Unica (Dim -> Hechos)"),
    ("Dim_Productos", "ID_Producto", "Hechos_Ventas", "ID_Producto", "1:N", "Unica (Dim -> Hechos)"),
]


def hoja_leeme(wb):
    ws = wb.create_sheet("LEEME")
    banda_titulo(ws, "  COMO SE COMPLETA ESTE LIBRO  (leer antes de entregar)", 3)
    ws.column_dimensions["A"].width = 6
    ws.column_dimensions["B"].width = 110

    pasos = [
        ("", "Este libro trae la documentacion escrita. El modelo de datos lo construye la macro,"),
        ("", "porque Power Pivot solo puede crearse desde adentro de Excel."),
        ("", ""),
        ("1", "Archivo > Guardar como > tipo 'Libro de Excel habilitado para macros (*.xlsm)'."),
        ("2", "Alt + F11 > menu Insertar > Modulo."),
        ("3", "Pegar el codigo de la macro en la hoja en blanco que se abre. Cerrar con Alt + F11."),
        ("4", "Vista > Macros > Ver macros > MaterializarModelo > Ejecutar. Confirmar el cartel."),
        ("", "   Si Excel pregunta por niveles de privacidad: elegir Publico en los dos origenes."),
        ("5", "Al terminar: Power Pivot > Administrar > tabla Dim_Calendario > pestania Disenar >"),
        ("", "   Marcar como tabla de fechas > columna Fecha.   (unico paso manual)"),
        ("6", "Archivo > Guardar como > .xlsx. Se pierden las macros y se conserva el modelo."),
        ("", ""),
        ("", "LO QUE CREA LA MACRO:"),
        ("", "   - carpeta \\Fuentes al lado del libro: Ventas_Historico.csv + Maestras.xlsx"),
        ("", "   - 4 consultas de Power Query cargadas al Modelo de datos"),
        ("", "   - 3 relaciones 1:N en esquema estrella"),
        ("", "   - 8 medidas DAX en la tabla Hechos_Ventas"),
        ("", "   - hoja TD_Ventas con la tabla dinamica de verificacion"),
        ("", "   - hoja Datos_Estadistica + Estadistica_Descriptiva"),
        ("", "   - hoja Simulacion + Resumen de escenarios (Pesimista / Base / Optimista)"),
        ("", "   - hoja Verificacion_Modelo con el log de todo lo anterior"),
        ("", ""),
        ("", "DONDE SE COMPRUEBA CADA PUNTO DE LA CONSIGNA:"),
        ("", "   Power Query .... Datos > Consultas y conexiones (las 4 consultas)"),
        ("", "   Lenguaje M ..... hoja Codigo_M, y Editor avanzado de cada consulta"),
        ("", "   Esquema estrella .... Power Pivot > Administrar > Vista de diagrama"),
        ("", "   Medidas DAX .... area de calculo de Hechos_Ventas, y hoja Medidas_DAX"),
        ("", "   Estadistica .... hoja Estadistica_Descriptiva"),
        ("", "   Escenarios ..... hoja Resumen de escenarios"),
    ]

    fila = 3
    for num, texto in pasos:
        if num:
            c = ws.cell(row=fila, column=1, value=num)
            c.font = Font(name="Calibri", size=11, bold=True, color=AZUL)
            c.alignment = Alignment(horizontal="center")
        c = ws.cell(row=fila, column=2, value=texto)
        c.font = SUBTITULO if texto.isupper() or texto.endswith(":") else NORMAL
        fila += 1

    return ws


def hoja_codigo_m(wb):
    ws = wb.create_sheet("Codigo_M")
    banda_titulo(ws, "  CODIGO M DE LAS 4 CONSULTAS  (Editor avanzado de Power Query)", 3)
    ws.column_dimensions["A"].width = 118
    ws.sheet_view.showGridLines = False

    fila = 3
    c = ws.cell(row=fila, column=1,
                value="Cada bloque es el contenido completo del Editor avanzado de la consulta. "
                      "La macro los carga tal cual, reemplazando <ruta del libro> por la ruta real.")
    c.font = NORMAL
    fila += 2

    for titulo, codigo in (
        ("shared Hechos_Ventas =", M_HECHOS),
        ("shared Dim_Clientes =", M_CLIENTES),
        ("shared Dim_Productos =", M_PRODUCTOS),
        ("shared Dim_Calendario =", M_CALENDARIO),
    ):
        fila = bloque_codigo(ws, fila, titulo, codigo.split("\n"))

    return ws


def hoja_medidas(wb):
    ws = wb.create_sheet("Medidas_DAX")
    banda_titulo(ws, "  MEDIDAS DAX DEL MODELO  (area de calculo de Hechos_Ventas)", 4)
    for col, ancho in zip("ABCD", (24, 62, 10, 60)):
        ws.column_dimensions[col].width = ancho
    ws.sheet_view.showGridLines = False

    fila = 3
    for col, texto in enumerate(("Medida", "Formula DAX", "VAR/RETURN", "Por que esta escrita asi"), start=1):
        c = ws.cell(row=fila, column=col, value=texto)
        c.font = CABECERA
        c.fill = FONDO_CABECERA
        c.alignment = Alignment(vertical="center")
        c.border = BORDE
    ws.row_dimensions[fila].height = 20

    fila += 1
    for nombre, formula, usa_var, nota in MEDIDAS:
        ws.cell(row=fila, column=1, value=nombre).font = Font(name="Calibri", size=11, bold=True)
        c = ws.cell(row=fila, column=2, value=formula)
        c.font = CODIGO
        c.alignment = WRAP
        c = ws.cell(row=fila, column=3, value=usa_var)
        c.alignment = Alignment(horizontal="center", vertical="top")
        c.font = Font(name="Calibri", size=11, bold=(usa_var == "Si"),
                      color=VERDE if usa_var == "Si" else "808080")
        c = ws.cell(row=fila, column=4, value=nota)
        c.font = NORMAL
        c.alignment = WRAP
        for col in range(1, 5):
            ws.cell(row=fila, column=col).border = BORDE
            if usa_var == "Si":
                ws.cell(row=fila, column=col).fill = FONDO_GRIS
        ws.row_dimensions[fila].height = max(34, 15 * (formula.count("\n") + 1))
        fila += 1

    fila += 1
    c = ws.cell(row=fila, column=1, value="Optimizacion aplicada")
    c.font = SUBTITULO
    fila += 1
    for texto in (
        "1. Las metricas de fila (Venta_Neta, Costo_Total, Margen_Bruto) se calculan en Power Query, no en DAX.",
        "   Eso permite resolver el margen con SUM() en vez de SUMX() sobre la tabla de hechos completa.",
        "2. Las 4 medidas derivadas usan VAR/RETURN: cada subexpresion se evalua una sola vez por contexto de filtro.",
        "3. La tabla de hechos conserva 9 columnas de las 13 originales. Las columnas de alta cardinalidad que ya",
        "   quedaron absorbidas en las metricas se descartan en el ETL para bajar el tamanio del modelo en memoria.",
    ):
        ws.cell(row=fila, column=1, value=texto).font = NORMAL
        fila += 1

    return ws


def hoja_estrella(wb):
    ws = wb.create_sheet("Modelo_Estrella")
    banda_titulo(ws, "  DISENO DIMENSIONAL  -  ESQUEMA EN ESTRELLA", 6)
    for col, ancho in zip("ABCDEF", (22, 16, 22, 16, 12, 26)):
        ws.column_dimensions[col].width = ancho
    ws.sheet_view.showGridLines = False

    diagrama = [
        "",
        "                          +------------------------+",
        "                          |     Dim_Calendario     |",
        "                          |  PK  Fecha             |",
        "                          |      Anio, Trimestre,  |",
        "                          |      Mes, Nombre_Mes   |",
        "                          +-----------+------------+",
        "                                      |  1",
        "                                      |",
        "                                      v  N",
        "  +---------------------+   1     N +------------------------+   N     1   +----------------------+",
        "  |    Dim_Clientes     |---------->|     Hechos_Ventas      |<------------|    Dim_Productos     |",
        "  |  PK  ID_Cliente     |           |  FK  Fecha             |             |  PK  ID_Producto     |",
        "  |      Cliente        |           |  FK  ID_Cliente        |             |      Producto        |",
        "  |      Segmento       |           |  FK  ID_Producto       |             |      Categoria       |",
        "  |      Provincia      |           |      Canal, Cantidad   |             |      Marca           |",
        "  |      Pais           |           |      Venta_Neta        |             |      Precio_Lista    |",
        "  |                     |           |      Costo_Total       |             |                      |",
        "  |                     |           |      Margen_Bruto      |             |                      |",
        "  +---------------------+           +------------------------+             +----------------------+",
        "",
        "  Las tres dimensiones estan del lado 1. La tabla de hechos esta del lado N.",
        "  El filtro fluye siempre en una sola direccion: de las dimensiones hacia los hechos.",
    ]

    fila = 3
    for linea in diagrama:
        ws.cell(row=fila, column=1, value=linea).font = CODIGO
        fila += 1

    fila += 1
    c = ws.cell(row=fila, column=1, value="Tabla de relaciones")
    c.font = SUBTITULO
    fila += 1

    cabeceras = ("Tabla origen (1)", "Clave primaria", "Tabla destino (N)", "Clave foranea",
                 "Cardinalidad", "Direccion del filtro")
    for col, texto in enumerate(cabeceras, start=1):
        c = ws.cell(row=fila, column=col, value=texto)
        c.font = CABECERA
        c.fill = FONDO_CABECERA
        c.border = BORDE
    fila += 1

    for rel in RELACIONES:
        for col, texto in enumerate(rel, start=1):
            c = ws.cell(row=fila, column=col, value=texto)
            c.font = NORMAL
            c.border = BORDE
        fila += 1

    fila += 1
    for texto in (
        "Integridad referencial: cada dimension termina su consulta M con Table.Distinct sobre la clave primaria.",
        "Sin esa unicidad Power Pivot rechaza la relacion 1:N y la degrada a muchos-a-muchos.",
        "Dim_Calendario se marca como Tabla de fechas (Power Pivot > Disenar) para habilitar DATEADD y SAMEPERIODLASTYEAR.",
    ):
        ws.cell(row=fila, column=1, value=texto).font = NORMAL
        fila += 1

    return ws


def main():
    wb = Workbook()
    wb.remove(wb.active)

    hoja_leeme(wb)
    hoja_codigo_m(wb)
    hoja_medidas(wb)
    hoja_estrella(wb)

    wb.properties.title = "Pre-entrega Modulo 8 - Modelo BI"
    wb.properties.creator = "Lleyton Murphy"

    destino = "PreEntrega_ModeloBI_MurphyLleyton.xlsx"
    wb.save(destino)
    print("Escrito:", destino)


if __name__ == "__main__":
    main()
