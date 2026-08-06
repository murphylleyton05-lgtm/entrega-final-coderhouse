"""
Arma el libro `PreEntrega_ModeloBI_MurphyLleyton.xlsx` con openpyxl.

Estructura del libro (el orden de las pestanas es el orden de lectura):

    Inicio                   que es el archivo, indice y el paso final en Excel
    Parametros               ruta de las fuentes y supuestos del modelo
    Modelo_Estrella          diagrama del esquema y tabla de relaciones 1:N
    Medidas_DAX              las medidas, con el antes y despues de optimizar
    Codigo_M                 el Section1.m completo, legible sin abrir el editor
    Estadistica_Descriptiva  salida del Analysis ToolPak + verificacion
    Simulacion               cuenta de resultados 2026, Buscar Objetivo, escenarios
    Resumen de escenarios    comparacion Pesimista / Base / Optimista
    Hechos_Ventas            tabla de hechos (5.200 filas)
    Dim_Clientes             dimension
    Dim_Productos            dimension
    Dim_Calendario           dimension

Las hojas de datos son el resultado de las consultas: se dejan visibles para
que el modelo se pueda auditar y para que el Analysis ToolPak tenga un rango
sobre el que trabajar (el ToolPak no lee el modelo de datos, solo celdas).

Los valores de las celdas con formula se cachean despues, en `ooxml.cachear`,
con los numeros que calcula `simulacion.py`. Por eso este modulo devuelve, al
final, el mapa de valores esperados.
"""
import os

import openpyxl
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter
from openpyxl.workbook.defined_name import DefinedName
from openpyxl.worksheet.table import Table, TableStyleInfo

import consultas_m
import datos
import estadistica
import medidas_dax
import simulacion

AQUI = os.path.dirname(os.path.abspath(__file__))
RAIZ = os.path.dirname(AQUI)
NOMBRE_ARCHIVO = "PreEntrega_ModeloBI_MurphyLleyton.xlsx"

# --- paleta: tres colores, como en el dashboard del Modulo 9 -----------------
AZUL = "1F3864"       # institucional
AZUL_MEDIO = "2F5597"
AZUL_SUAVE = "D6DCE4"
GRIS = "8C8C8C"
GRIS_SUAVE = "F2F2F2"
AMBAR = "E8A33D"
AMBAR_SUAVE = "FCE9CF"
BLANCO = "FFFFFF"
TEXTO = "262626"

F_TITULO = Font(name="Calibri", size=20, bold=True, color=AZUL)
F_BAJADA = Font(name="Calibri", size=11, italic=True, color=GRIS)
F_SECCION = Font(name="Calibri", size=12, bold=True, color=BLANCO)
F_CABECERA = Font(name="Calibri", size=10, bold=True, color=BLANCO)
F_ETIQUETA = Font(name="Calibri", size=11, color=TEXTO)
F_ETIQUETA_B = Font(name="Calibri", size=11, bold=True, color=TEXTO)
F_VALOR = Font(name="Calibri", size=11, color=TEXTO)
F_NOTA = Font(name="Calibri", size=10, color=GRIS)
F_CODIGO = Font(name="Consolas", size=9, color=TEXTO)
F_CODIGO_COM = Font(name="Consolas", size=9, color="3F7F5F")   # comentarios
F_DESTACADO = Font(name="Calibri", size=11, bold=True, color=AZUL)

R_SECCION = PatternFill("solid", fgColor=AZUL)
R_CABECERA = PatternFill("solid", fgColor=AZUL_MEDIO)
R_SUAVE = PatternFill("solid", fgColor=GRIS_SUAVE)
R_AZUL_SUAVE = PatternFill("solid", fgColor=AZUL_SUAVE)
R_CAMBIANTE = PatternFill("solid", fgColor=AMBAR_SUAVE)
R_AMBAR = PatternFill("solid", fgColor=AMBAR)
R_BLANCO = PatternFill("solid", fgColor=BLANCO)

_linea = Side(style="thin", color="BFBFBF")
BORDE = Border(left=_linea, right=_linea, top=_linea, bottom=_linea)

IZQ = Alignment(horizontal="left", vertical="center")
IZQ_ARRIBA = Alignment(horizontal="left", vertical="top", wrap_text=True)
DER = Alignment(horizontal="right", vertical="center")
CENTRO = Alignment(horizontal="center", vertical="center")
CENTRO_WRAP = Alignment(horizontal="center", vertical="center", wrap_text=True)

FMT_MONEDA = '#,##0.00'
FMT_MONEDA0 = '#,##0'
FMT_ENTERO = '#,##0'
FMT_PCT1 = '0.0%'
FMT_PCT2 = '0.00%'
FMT_FECHA = 'yyyy-mm-dd'
FMT_NUM4 = '#,##0.0000'


# --- formato de numeros en los textos ----------------------------------------
# Los textos del libro estan en espanol: miles con punto y decimales con coma.
# Se formatea numero por numero y nunca con un replace sobre la frase entera,
# que se comeria las comas de la prosa.

def num(valor, decimales=0):
    return (f"{valor:,.{decimales}f}"
            .replace(",", "\x00").replace(".", ",").replace("\x00", "."))


def pct(valor, decimales=1, signo=False):
    texto = f"{valor * 100:{'+' if signo else ''}.{decimales}f}".replace(".", ",")
    return texto + "%"


# --- utilidades de escritura -------------------------------------------------

class Libro:
    """Envoltorio fino sobre openpyxl que ademas junta los valores a cachear."""

    def __init__(self):
        self.wb = openpyxl.Workbook()
        self.wb.remove(self.wb.active)
        self.cache = {}          # {hoja: {celda: valor calculado}}

    def hoja(self, nombre, anchos=None, sin_grilla=True):
        ws = self.wb.create_sheet(nombre)
        ws.sheet_view.showGridLines = not sin_grilla
        for letra, ancho in (anchos or {}).items():
            ws.column_dimensions[letra].width = ancho
        return ws

    def formula(self, ws, celda, formula, valor, formato=None, fuente=None,
                relleno=None, alineacion=None, borde=False):
        """Escribe una formula y anota el valor que debe quedar cacheado."""
        self.escribir(ws, celda, formula, formato=formato, fuente=fuente,
                      relleno=relleno, alineacion=alineacion, borde=borde)
        self.cache.setdefault(ws.title, {})[celda] = valor
        return valor

    @staticmethod
    def escribir(ws, celda, valor, formato=None, fuente=None, relleno=None,
                 alineacion=None, borde=False):
        c = ws[celda]
        c.value = valor
        if formato:
            c.number_format = formato
        c.font = fuente or F_VALOR
        if relleno:
            c.fill = relleno
        if alineacion:
            c.alignment = alineacion
        if borde:
            c.border = BORDE
        return c


def titulo(lb, ws, texto, bajada, fila=2, col="B", ancho=8):
    lb.escribir(ws, f"{col}{fila}", texto, fuente=F_TITULO, alineacion=IZQ)
    ws.row_dimensions[fila].height = 28
    if bajada:
        lb.escribir(ws, f"{col}{fila + 1}", bajada, fuente=F_BAJADA, alineacion=IZQ)
    return fila + 3


def seccion(lb, ws, fila, texto, desde="B", hasta="E"):
    ws.merge_cells(f"{desde}{fila}:{hasta}{fila}")
    lb.escribir(ws, f"{desde}{fila}", "  " + texto, fuente=F_SECCION,
                relleno=R_SECCION, alineacion=IZQ)
    ws.row_dimensions[fila].height = 22
    return fila + 1


def cabecera(lb, ws, fila, valores, columnas):
    for col, valor in zip(columnas, valores):
        lb.escribir(ws, f"{col}{fila}", valor, fuente=F_CABECERA,
                    relleno=R_CABECERA, alineacion=CENTRO_WRAP, borde=True)
    ws.row_dimensions[fila].height = 20
    return fila + 1


def nota(lb, ws, fila, texto, desde="B", hasta="E", alto=None):
    ws.merge_cells(f"{desde}{fila}:{hasta}{fila}")
    lb.escribir(ws, f"{desde}{fila}", texto, fuente=F_NOTA, alineacion=IZQ_ARRIBA)
    if alto:
        ws.row_dimensions[fila].height = alto
    return fila + 1


# --- hojas de datos ----------------------------------------------------------

FACT_COLS = ["ID_Venta", "Fecha", "ID_Cliente", "ID_Producto", "Canal",
             "Unidades", "Precio_Unitario", "Descuento", "Costo_Unitario",
             "Costo_Envio", "Venta_Bruta", "Venta_Neta", "Costo_Total",
             "Margen_Bruto"]

FORMATOS_FACT = {
    "Fecha": FMT_FECHA, "Unidades": FMT_ENTERO, "Precio_Unitario": FMT_MONEDA,
    "Descuento": FMT_PCT1, "Costo_Unitario": FMT_MONEDA, "Costo_Envio": FMT_MONEDA,
    "Venta_Bruta": FMT_MONEDA, "Venta_Neta": FMT_MONEDA,
    "Costo_Total": FMT_MONEDA, "Margen_Bruto": FMT_MONEDA,
}


def hoja_datos(lb, nombre, columnas, filas, tabla, anchos, formatos, comentario):
    """
    Una hoja por consulta cargada. La fila 1 lleva una nota de origen y la
    tabla arranca en la fila 3, para que quede claro que el contenido es el
    resultado de una consulta y no datos escritos a mano.
    """
    ws = lb.hoja(nombre, sin_grilla=False)
    lb.escribir(ws, "A1", comentario, fuente=F_NOTA, alineacion=IZQ)

    for j, col in enumerate(columnas, start=1):
        lb.escribir(ws, f"{get_column_letter(j)}3", col,
                    fuente=F_CABECERA, relleno=R_CABECERA,
                    alineacion=CENTRO, borde=True)

    for i, fila in enumerate(filas, start=4):
        for j, col in enumerate(columnas, start=1):
            c = ws.cell(row=i, column=j, value=fila[col])
            c.font = F_VALOR
            if col in formatos:
                c.number_format = formatos[col]

    ref = f"A3:{get_column_letter(len(columnas))}{len(filas) + 3}"
    t = Table(displayName=tabla, ref=ref)
    t.tableStyleInfo = TableStyleInfo(name="TableStyleLight9", showRowStripes=True)
    ws.add_table(t)

    for letra, ancho in anchos.items():
        ws.column_dimensions[letra].width = ancho
    ws.freeze_panes = "A4"
    return ws, len(filas)


# --- hoja Inicio -------------------------------------------------------------

INDICE = [
    ("Parametros", "Ruta de los archivos de origen y supuestos del modelo",
     "Paso 1 - conexion reproducible"),
    ("Modelo_Estrella", "Diagrama del esquema en estrella y las 3 relaciones 1:N",
     "Paso 3 - diseno dimensional"),
    ("Medidas_DAX", "Las 8 medidas, con VAR/RETURN y el antes/despues de optimizar",
     "Paso 4 - optimizacion DAX"),
    ("Codigo_M", "El script Section1.m completo, con los pasos renombrados",
     "Paso 2 - lenguaje M"),
    ("Estadistica_Descriptiva", "Salida del Analysis ToolPak sobre Venta_Neta y Margen_Bruto",
     "Paso 5 - diagnostico"),
    ("Simulacion", "Cuenta de resultados 2026, Buscar Objetivo y celdas cambiantes",
     "Paso 6 - analisis de hipotesis"),
    ("Resumen de escenarios", "Pesimista / Base / Optimista, uno al lado del otro",
     "Paso 6 - administrador de escenarios"),
    ("Hechos_Ventas", "Tabla de hechos: 5.200 operaciones limpias, 2024-2025",
     "Salida de la consulta Hechos_Ventas"),
    ("Dim_Clientes", "60 clientes. PK: ID_Cliente", "Salida de la consulta Dim_Clientes"),
    ("Dim_Productos", "10 productos. PK: ID_Producto", "Salida de la consulta Dim_Productos"),
    ("Dim_Calendario", "731 dias, 2024-2025. PK: Fecha", "Salida de la consulta Dim_Calendario"),
]

PASOS_EXCEL = [
    ("1", "Cargar las consultas al modelo de datos",
     "Datos > Consultas y conexiones. Para cada consulta (Hechos_Ventas, "
     "Dim_Clientes, Dim_Productos, Dim_Calendario): clic derecho > Cargar en... "
     "> Crear solo conexion + tildar Agregar estos datos al Modelo de datos. "
     "pRutaFuentes queda como solo conexion, sin modelo: es un parametro, no una tabla."),
    ("2", "Crear las tres relaciones 1:N",
     "Power Pivot > Administrar > Vista de diagrama. Arrastrar la PK de cada "
     "dimension sobre la FK de Hechos_Ventas, segun la tabla de la hoja "
     "Modelo_Estrella. Marcar Dim_Calendario como Tabla de fechas (columna Fecha)."),
    ("3", "Pegar las medidas DAX",
     "En la ventana de Power Pivot, pestana Hechos_Ventas, area de calculo: "
     "pegar una medida por celda. El codigo esta completo en la hoja Medidas_DAX "
     "y en el archivo dax/medidas.dax."),
]


def hoja_inicio(lb, resumen_datos):
    ws = lb.hoja("Inicio", anchos={"A": 2.5, "B": 26, "C": 62, "D": 34, "E": 3})
    f = titulo(lb, ws, "Pre-entrega: Optimizacion y Modelado de Datos",
               "Modulo 8 - Arquitectura del dato, optimizacion de modelos y "
               "simulacion de escenarios | Lleyton Murphy", col="B")

    f = seccion(lb, ws, f, "De que se trata", hasta="D")
    f = nota(lb, ws,  f,
             "TechStore es una cadena de retail de tecnologia. Este libro reconstruye su "
             "ecosistema de datos de punta a punta: toma el historico transaccional en "
             "CSV y las tablas maestras en XLSX, los pasa por un flujo ETL en Power Query, "
             "los estructura como modelo en estrella para Power Pivot, define las metricas "
             "en DAX optimizado, diagnostica la distribucion de la venta con el Analysis "
             "ToolPak y cierra con la capa de simulacion financiera para 2026.",
             hasta="D", alto=62)
    f += 1

    f = seccion(lb, ws, f, "Indice de hojas", hasta="D")
    f = cabecera(lb, ws, f, ["Hoja", "Que contiene", "Punto de la consigna"],
                 ["B", "C", "D"])
    for i, (hoja, que, punto) in enumerate(INDICE):
        relleno = R_SUAVE if i % 2 else None
        lb.escribir(ws, f"B{f}", hoja, fuente=F_ETIQUETA_B, relleno=relleno,
                    alineacion=IZQ, borde=True)
        lb.escribir(ws, f"C{f}", que, relleno=relleno, alineacion=IZQ, borde=True)
        lb.escribir(ws, f"D{f}", punto, fuente=F_NOTA, relleno=relleno,
                    alineacion=IZQ, borde=True)
        f += 1
    f += 1

    f = seccion(lb, ws, f, "Paso final dentro de Excel (5 minutos)", hasta="D")
    f = nota(lb, ws, f,
             "El modelo tabular de Power Pivot es una base en memoria que solo Excel puede "
             "escribir: no existe forma de generarlo desde afuera del programa. Todo lo "
             "demas ya esta en el archivo (las consultas M, los datos, las medidas escritas, "
             "las relaciones especificadas, los escenarios y el resumen). Estos tres pasos "
             "materializan el modelo dentro de Excel y no cambian ningun numero del libro.",
             hasta="D", alto=52)
    f = cabecera(lb, ws, f, ["Paso", "Que hacer", "Como"], ["B", "C", "D"])
    for numero, que, como in PASOS_EXCEL:
        lb.escribir(ws, f"B{f}", f"Paso {numero}", fuente=F_DESTACADO,
                    alineacion=CENTRO, borde=True, relleno=R_AZUL_SUAVE)
        lb.escribir(ws, f"C{f}", que, fuente=F_ETIQUETA_B, alineacion=IZQ_ARRIBA, borde=True)
        lb.escribir(ws, f"D{f}", como, fuente=F_NOTA, alineacion=IZQ_ARRIBA, borde=True)
        ws.row_dimensions[f].height = 74
        f += 1
    f += 1

    f = seccion(lb, ws, f, "Antes de empezar", hasta="D")
    avisos = [
        ("Mantener la carpeta fuentes junto al libro",
         "Las consultas no tienen la ruta escrita adentro: la leen de la celda RutaFuentes "
         "de la hoja Parametros, que se arma sola con la carpeta donde este el archivo. "
         "Si se mueve el libro sin la carpeta fuentes, alcanza con corregir esa unica celda."),
        ("Si aparece el error Formula.Firewall",
         "Datos > Obtener datos > Opciones de consulta > Privacidad > Omitir siempre la "
         "configuracion de niveles de privacidad. Pasa porque pRutaFuentes lee una celda "
         "del libro y las demas consultas leen archivos: son dos origenes distintos."),
        ("Complementos necesarios",
         "Analysis ToolPak (Herramientas para analisis) y Power Pivot, ambos desde "
         "Archivo > Opciones > Complementos. Requiere Excel de escritorio 2016 o posterior."),
    ]
    for i, (que, detalle) in enumerate(avisos):
        relleno = R_SUAVE if i % 2 else None
        lb.escribir(ws, f"B{f}", que, fuente=F_ETIQUETA_B, relleno=relleno,
                    alineacion=IZQ_ARRIBA, borde=True)
        ws.merge_cells(f"C{f}:D{f}")
        lb.escribir(ws, f"C{f}", detalle, fuente=F_NOTA, relleno=relleno,
                    alineacion=IZQ_ARRIBA, borde=True)
        ws.row_dimensions[f].height = 46
        f += 1
    f += 1

    f = nota(lb, ws, f,
             f"Los datos son simulados y deterministas (semilla fija). El CSV de origen trae "
             f"{num(resumen_datos['crudas'])} filas; despues de la limpieza quedan "
             f"{num(resumen_datos['limpias'])}. Venta neta acumulada 2024-2025: USD "
             f"{num(resumen_datos['venta_neta'], 2)}.",
             hasta="D", alto=30)
    return ws


# --- hoja Parametros ---------------------------------------------------------

def hoja_parametros(lb, base):
    ws = lb.hoja("Parametros", anchos={"A": 2.5, "B": 46, "C": 30, "D": 62, "E": 3})
    f = titulo(lb, ws, "Parametros del modelo",
               "Toda constante del libro vive en esta hoja. Ninguna consulta ni "
               "formula lleva un valor escrito adentro.")

    f = seccion(lb, ws, f, "1 | Ruta de los archivos de origen", hasta="D")
    f = cabecera(lb, ws, f, ["Parametro", "Valor", "Para que sirve"], ["B", "C", "D"])
    fila_carpeta = f
    lb.escribir(ws, f"B{f}", "Carpeta de este libro", fuente=F_ETIQUETA, alineacion=IZQ, borde=True)
    lb.formula(ws, f"C{f}", '=LEFT(CELL("filename"),FIND("[",CELL("filename"))-1)',
               "C:\\TechStore\\PreEntrega_ModeloBI\\", alineacion=IZQ, borde=True)
    lb.escribir(ws, f"D{f}", "Se resuelve sola al abrir el archivo: CELL(\"filename\") "
                             "devuelve la ruta completa del libro guardado.",
                fuente=F_NOTA, alineacion=IZQ_ARRIBA, borde=True)
    f += 1

    lb.escribir(ws, f"B{f}", "Subcarpeta de las fuentes", fuente=F_ETIQUETA, alineacion=IZQ, borde=True)
    lb.escribir(ws, f"C{f}", "fuentes\\", alineacion=IZQ, borde=True)
    lb.escribir(ws, f"D{f}", "Donde estan ventas_historico.csv y maestros.xlsx.",
                fuente=F_NOTA, alineacion=IZQ_ARRIBA, borde=True)
    fila_sub = f
    f += 1

    fila_ruta = f
    lb.escribir(ws, f"B{f}", "RutaFuentes  (la lee Power Query)",
                fuente=F_ETIQUETA_B, alineacion=IZQ, borde=True, relleno=R_CAMBIANTE)
    lb.formula(ws, f"C{f}", f"=C{fila_carpeta}&C{fila_sub}",
               "C:\\TechStore\\PreEntrega_ModeloBI\\fuentes\\",
               alineacion=IZQ, borde=True, relleno=R_CAMBIANTE, fuente=F_ETIQUETA_B)
    lb.escribir(ws, f"D{f}", "Celda con nombre. La consulta pRutaFuentes la lee con "
                             "Excel.CurrentWorkbook y el resto de las consultas la usa. "
                             "Si el libro se movio sin la carpeta, se escribe la ruta aca a mano.",
                fuente=F_NOTA, alineacion=IZQ_ARRIBA, borde=True)
    ws.row_dimensions[f].height = 44
    f += 1

    for etiqueta, valor, detalle in [
        ("Archivo de hechos", "ventas_historico.csv", "Historico transaccional, formato CSV."),
        ("Libro de tablas maestras", "maestros.xlsx",
         "Tablas tbl_Clientes y tbl_Productos, formato XLSX."),
    ]:
        lb.escribir(ws, f"B{f}", etiqueta, fuente=F_ETIQUETA, alineacion=IZQ, borde=True)
        lb.escribir(ws, f"C{f}", valor, alineacion=IZQ, borde=True)
        lb.escribir(ws, f"D{f}", detalle, fuente=F_NOTA, alineacion=IZQ_ARRIBA, borde=True)
        f += 1
    f += 1

    f = seccion(lb, ws, f, "2 | Supuestos de la simulacion financiera", hasta="D")
    f = cabecera(lb, ws, f, ["Supuesto", "Valor", "De donde sale"], ["B", "C", "D"])
    fila_meta = f
    supuestos = [
        ("Meta de beneficio neto 2026", simulacion.META_BENEFICIO, FMT_MONEDA0,
         "La define el directorio. Es el objetivo que resuelve Buscar Objetivo."),
        ("Gastos fijos de estructura (% s/venta neta 2025)", simulacion.RATIO_GASTOS_FIJOS,
         FMT_PCT1,
         "Sueldos, alquileres, marketing y administracion. Unico supuesto que no "
         "sale de los datos."),
        ("Tasa de interes financiero de referencia", simulacion.INTERES_REFERENCIA,
         FMT_PCT1,
         "Costo del capital de trabajo del escenario base."),
    ]
    for etiqueta, valor, formato, detalle in supuestos:
        lb.escribir(ws, f"B{f}", etiqueta, fuente=F_ETIQUETA, alineacion=IZQ, borde=True)
        lb.escribir(ws, f"C{f}", valor, formato=formato, alineacion=DER, borde=True,
                    fuente=F_ETIQUETA_B)
        lb.escribir(ws, f"D{f}", detalle, fuente=F_NOTA, alineacion=IZQ_ARRIBA, borde=True)
        ws.row_dimensions[f].height = 30
        f += 1
    f += 1

    f = seccion(lb, ws, f, "3 | Variacion de la demanda por escenario", hasta="D")
    f = cabecera(lb, ws, f, ["Escenario", "Variacion 2026", "Criterio"], ["B", "C", "D"])
    criterios = {
        "Pesimista": "Se corta el ciclo de reposicion de equipos y cae el consumo.",
        "Base": "Plan conservador: no extrapola el rebote de 2025, que fue "
                f"del {pct(base['crecimiento_unidades'])} en unidades.",
        "Optimista": "Se sostiene la demanda de 2025 con mejor conversion en e-commerce.",
    }
    for nombre in ("Pesimista", "Base", "Optimista"):
        lb.escribir(ws, f"B{f}", nombre, fuente=F_ETIQUETA_B, alineacion=IZQ, borde=True)
        lb.escribir(ws, f"C{f}", simulacion.PLAN_DEMANDA[nombre], formato=FMT_PCT1,
                    alineacion=DER, borde=True)
        lb.escribir(ws, f"D{f}", criterios[nombre], fuente=F_NOTA,
                    alineacion=IZQ_ARRIBA, borde=True)
        f += 1
    f += 1

    f = nota(lb, ws, f,
             "Los datos son simulados y deterministas: el generador usa semilla fija, "
             "asi que dos corridas producen exactamente los mismos archivos y los mismos "
             "numeros que se citan en el README.", hasta="D", alto=30)

    lb.wb.defined_names.add(DefinedName("RutaFuentes",
                                        attr_text=f"Parametros!$C${fila_ruta}"))
    return {"fila_meta": fila_meta,
            "fila_ratio_fijos": fila_meta + 1,
            "fila_interes": fila_meta + 2}


# --- hoja Modelo_Estrella ----------------------------------------------------

# (tabla lado 1, PK, tabla lado N, FK, cardinalidad, direccion del filtro)
RELACIONES = [
    ("Dim_Calendario  (731 claves)", "Fecha", "Hechos_Ventas", "Fecha",
     "Uno a varios (1:N)", "Dim_Calendario --> Hechos_Ventas"),
    ("Dim_Clientes  (60 claves)", "ID_Cliente", "Hechos_Ventas", "ID_Cliente",
     "Uno a varios (1:N)", "Dim_Clientes --> Hechos_Ventas"),
    ("Dim_Productos  (10 claves)", "ID_Producto", "Hechos_Ventas", "ID_Producto",
     "Uno a varios (1:N)", "Dim_Productos --> Hechos_Ventas"),
]

# columnas del libro que ocupa cada campo de la tabla de relaciones
COLUMNAS_REL = [("B", "C"), ("D", "E"), ("F", "G"), ("H", "I"), ("J", "K"), ("L", "M")]

CAJAS = {
    "Dim_Calendario": ["Fecha  (PK)", "Anio", "Trimestre", "Mes", "Nombre_Mes",
                       "Anio_Mes", "Dia_Semana", "Es_Fin_Semana"],
    "Dim_Clientes": ["ID_Cliente  (PK)", "Cliente", "Segmento", "Region",
                     "Ciudad", "Alta"],
    "Dim_Productos": ["ID_Producto  (PK)", "Producto", "Categoria", "Marca",
                      "Precio_Lista", "Costo_Estandar"],
    "Hechos_Ventas": ["ID_Venta", "Fecha  (FK)", "ID_Cliente  (FK)",
                      "ID_Producto  (FK)", "Canal", "Unidades",
                      "Precio_Unitario", "Descuento", "Costo_Unitario",
                      "Costo_Envio", "Venta_Bruta", "Venta_Neta",
                      "Costo_Total", "Margen_Bruto"],
}


def _caja(lb, ws, nombre, columnas, col_desde, col_hasta, fila, es_hechos=False):
    """Dibuja una tabla del diagrama con celdas: titulo + lista de columnas."""
    d, h = col_desde, col_hasta
    ws.merge_cells(f"{d}{fila}:{h}{fila}")
    lb.escribir(ws, f"{d}{fila}", nombre,
                fuente=Font(name="Calibri", size=11, bold=True, color=BLANCO),
                relleno=PatternFill("solid", fgColor=AZUL if es_hechos else AZUL_MEDIO),
                alineacion=CENTRO, borde=True)
    for i, col in enumerate(columnas, start=1):
        ws.merge_cells(f"{d}{fila + i}:{h}{fila + i}")
        clave = "(PK)" in col or "(FK)" in col
        lb.escribir(ws, f"{d}{fila + i}", "  " + col,
                    fuente=Font(name="Consolas", size=9, bold=clave,
                                color=AZUL if clave else TEXTO),
                    relleno=R_AZUL_SUAVE if clave else R_BLANCO,
                    alineacion=IZQ, borde=True)
    return fila + len(columnas) + 1


def hoja_modelo(lb):
    anchos = {"A": 2.5}
    for letra in "BCDEFGHIJKLM":
        anchos[letra] = 13
    ws = lb.hoja("Modelo_Estrella", anchos=anchos)

    f = titulo(lb, ws, "Modelo dimensional en estrella",
               "Una tabla de hechos al centro, tres dimensiones conformadas "
               "alrededor y tres relaciones 1:N que filtran hacia el centro.")

    # dimension de arriba (Calendario), centrada sobre la tabla de hechos
    fila_cal = f + 1
    fin_cal = _caja(lb, ws, "Dim_Calendario", CAJAS["Dim_Calendario"], "F", "H", fila_cal)

    # flecha hacia abajo
    lb.escribir(ws, f"G{fin_cal}", "1", fuente=F_DESTACADO, alineacion=CENTRO)
    lb.escribir(ws, f"G{fin_cal + 1}", "|", fuente=F_DESTACADO, alineacion=CENTRO)
    lb.escribir(ws, f"G{fin_cal + 2}", "v", fuente=F_DESTACADO, alineacion=CENTRO)
    lb.escribir(ws, f"G{fin_cal + 3}", "N", fuente=F_DESTACADO, alineacion=CENTRO)
    ws.column_dimensions["G"].width = 13

    fila_centro = fin_cal + 4
    _caja(lb, ws, "Hechos_Ventas", CAJAS["Hechos_Ventas"], "F", "H",
          fila_centro, es_hechos=True)

    # dimensiones laterales, alineadas con el centro de la tabla de hechos
    fila_lat = fila_centro + 3
    _caja(lb, ws, "Dim_Clientes", CAJAS["Dim_Clientes"], "B", "D", fila_lat)
    _caja(lb, ws, "Dim_Productos", CAJAS["Dim_Productos"], "J", "L", fila_lat)

    flecha = fila_lat + 3
    lb.escribir(ws, f"E{flecha}", "1 --> N", fuente=F_DESTACADO, alineacion=CENTRO)
    lb.escribir(ws, f"I{flecha}", "N <-- 1", fuente=F_DESTACADO, alineacion=CENTRO)

    f = fila_centro + len(CAJAS["Hechos_Ventas"]) + 3
    f = max(f, fila_lat + len(CAJAS["Dim_Clientes"]) + 3)

    f = seccion(lb, ws, f, "Relaciones a crear en la Vista de diagrama", hasta="M")
    encabezados = ["Tabla del lado 1", "Clave primaria (PK)", "Tabla del lado N",
                   "Clave foranea (FK)", "Cardinalidad", "Direccion del filtro"]
    for (desde, hasta), texto in zip(COLUMNAS_REL, encabezados):
        ws.merge_cells(f"{desde}{f}:{hasta}{f}")
        lb.escribir(ws, f"{desde}{f}", texto, fuente=F_CABECERA, relleno=R_CABECERA,
                    alineacion=CENTRO_WRAP, borde=True)
    ws.row_dimensions[f].height = 20
    f += 1

    fuentes_rel = [F_ETIQUETA_B, F_CODIGO, F_ETIQUETA_B, F_CODIGO, F_ETIQUETA, F_ETIQUETA]
    for i, rel in enumerate(RELACIONES):
        relleno = R_SUAVE if i % 2 else None
        for (desde, hasta), valor, fuente in zip(COLUMNAS_REL, rel, fuentes_rel):
            ws.merge_cells(f"{desde}{f}:{hasta}{f}")
            lb.escribir(ws, f"{desde}{f}", valor, fuente=fuente, relleno=relleno,
                        alineacion=IZQ, borde=True)
        f += 1
    f += 1
    f = nota(lb, ws, f,
             "Las tres relaciones son de direccion unica: el filtro va de la dimension a "
             "los hechos y nunca al reves. Ninguna dimension se relaciona con otra, que es "
             "lo que distingue una estrella de un copo de nieve.", desde="B", hasta="M")
    f += 1

    f = seccion(lb, ws, f, "Por que este diseno y no una tabla plana", hasta="M")
    argumentos = [
        ("Una sola tabla de hechos, sin redundancia",
         "El nombre del cliente, su region y la categoria del producto se guardan una "
         "vez en la dimension, no 5.200 veces en cada fila de venta. Eso es lo que hace "
         "que el motor VertiPaq comprima bien y que el archivo no crezca."),
        ("El filtro fluye siempre desde la dimension hacia los hechos",
         "Las tres relaciones son de direccion unica. Filtrar por Categoria o por "
         "Trimestre alcanza a la tabla de hechos, y no al reves: no hay caminos "
         "ambiguos ni relaciones circulares que el motor deba desempatar."),
        ("Las claves estan garantizadas del lado 1",
         "Cada consulta de dimension termina con Table.Distinct sobre su PK. Sin ese "
         "paso Power Pivot rechaza la relacion 1:N y la degrada a muchos a muchos."),
        ("Dim_Calendario cubre anios completos",
         "Se genera en M desde el 1 de enero del primer anio hasta el 31 de diciembre "
         "del ultimo. Un calendario recortado rompe SAMEPERIODLASTYEAR y cualquier "
         "acumulado del ejercicio."),
    ]
    for i, (que, detalle) in enumerate(argumentos):
        relleno = R_SUAVE if i % 2 else None
        ws.merge_cells(f"B{f}:D{f}")
        lb.escribir(ws, f"B{f}", que, fuente=F_ETIQUETA_B, relleno=relleno,
                    alineacion=IZQ_ARRIBA, borde=True)
        ws.merge_cells(f"E{f}:M{f}")
        lb.escribir(ws, f"E{f}", detalle, fuente=F_NOTA, relleno=relleno,
                    alineacion=IZQ_ARRIBA, borde=True)
        ws.row_dimensions[f].height = 44
        f += 1
    return ws


# --- hoja Medidas_DAX --------------------------------------------------------

def hoja_dax(lb):
    ws = lb.hoja("Medidas_DAX", anchos={"A": 2.5, "B": 26, "C": 72, "D": 58, "E": 3})
    f = titulo(lb, ws, "Medidas DAX del modelo",
               "Ocho medidas. Cuatro usan VAR/RETURN. Ninguna itera la tabla "
               "de hechos.", col="B")

    f = seccion(lb, ws, f, "Medidas", hasta="D")
    f = cabecera(lb, ws, f, ["Medida", "Codigo DAX", "Por que esta escrita asi"],
                 ["B", "C", "D"])
    for i, (nombre, codigo, explicacion) in enumerate(medidas_dax.MEDIDAS):
        relleno = R_SUAVE if i % 2 else None
        usa_var = codigo.strip().startswith(nombre) and "VAR " in codigo
        lb.escribir(ws, f"B{f}", nombre,
                    fuente=F_DESTACADO if usa_var else F_ETIQUETA_B,
                    relleno=R_CAMBIANTE if usa_var else relleno,
                    alineacion=IZQ_ARRIBA, borde=True)
        lb.escribir(ws, f"C{f}", codigo, fuente=F_CODIGO, relleno=relleno,
                    alineacion=IZQ_ARRIBA, borde=True)
        lb.escribir(ws, f"D{f}", explicacion, fuente=F_NOTA, relleno=relleno,
                    alineacion=IZQ_ARRIBA, borde=True)
        ws.row_dimensions[f].height = max(30, 13 * (codigo.count("\n") + 1) + 8)
        f += 1
    f += 1
    f = nota(lb, ws, f, "Las medidas resaltadas en ambar son las que usan la "
                        "estructura VAR / RETURN.", hasta="D")
    f += 1

    f = seccion(lb, ws, f, "Antes y despues: la misma metrica, dos costos distintos",
                hasta="D")
    f = cabecera(lb, ws, f, ["Version", "Codigo DAX", "Lectura"], ["B", "C", "D"])

    lb.escribir(ws, f"B{f}", "Sin optimizar", fuente=F_ETIQUETA_B,
                relleno=PatternFill("solid", fgColor="F8CBAD"),
                alineacion=IZQ_ARRIBA, borde=True)
    lb.escribir(ws, f"C{f}", medidas_dax.COMPARACION["lenta"], fuente=F_CODIGO,
                alineacion=IZQ_ARRIBA, borde=True)
    lb.escribir(ws, f"D{f}",
                "\n".join("- " + p for p in medidas_dax.COMPARACION["problemas"]),
                fuente=F_NOTA, alineacion=IZQ_ARRIBA, borde=True)
    ws.row_dimensions[f].height = 210
    f += 1

    lb.escribir(ws, f"B{f}", "Optimizada", fuente=F_ETIQUETA_B,
                relleno=PatternFill("solid", fgColor="C6E0B4"),
                alineacion=IZQ_ARRIBA, borde=True)
    lb.escribir(ws, f"C{f}", medidas_dax.COMPARACION["rapida"], fuente=F_CODIGO,
                alineacion=IZQ_ARRIBA, borde=True)
    lb.escribir(ws, f"D{f}", medidas_dax.COMPARACION["resultado"], fuente=F_NOTA,
                alineacion=IZQ_ARRIBA, borde=True)
    ws.row_dimensions[f].height = 84
    f += 2

    f = seccion(lb, ws, f, "Que se hizo para que el modelo no pese", hasta="D")
    tecnicas = [
        ("La aritmetica de fila se resuelve en el ETL",
         "Venta_Bruta, Venta_Neta, Costo_Total y Margen_Bruto se calculan en Power Query. "
         "Las medidas quedan como SUM() sobre columnas ya comprimidas."),
        ("Se descartan columnas que no aportan",
         "La columna Estado sale del modelo: despues del filtro vale Confirmada en las "
         "5.200 filas. Una columna constante ocupa memoria y no discrimina nada."),
        ("VAR para no repetir subtotales",
         "Cuando una medida necesita el mismo subtotal dos o tres veces, se guarda en una "
         "variable. El motor lo evalua una vez por contexto de filtro en lugar de una vez "
         "por aparicion."),
        ("Nada de FILTER sobre la tabla completa",
         "FILTER(Hechos_Ventas, ...) materializa la tabla entera antes de filtrar. Los "
         "filtros del modelo se resuelven sobre columnas, que es lo que el motor sabe "
         "hacer barato."),
        ("Se reutilizan medidas en vez de repetir expresiones",
         "Crecimiento Interanual %, % Margen Bruto y Ticket Promedio se apoyan en "
         "[Total Ventas]. Si cambia la definicion de venta, cambia en un solo lugar."),
    ]
    for i, (que, detalle) in enumerate(tecnicas):
        relleno = R_SUAVE if i % 2 else None
        lb.escribir(ws, f"B{f}", que, fuente=F_ETIQUETA_B, relleno=relleno,
                    alineacion=IZQ_ARRIBA, borde=True)
        ws.merge_cells(f"C{f}:D{f}")
        lb.escribir(ws, f"C{f}", detalle, fuente=F_NOTA, relleno=relleno,
                    alineacion=IZQ_ARRIBA, borde=True)
        ws.row_dimensions[f].height = 44
        f += 1
    return ws


# --- hoja Codigo_M -----------------------------------------------------------

def hoja_codigo_m(lb):
    ws = lb.hoja("Codigo_M", anchos={"A": 2.5, "B": 118})
    f = titulo(lb, ws, "Codigo M del flujo ETL",
               "Contenido literal de Section1.m. Es el mismo script que esta "
               "embebido en el libro y en powerquery/Section1.m.")

    f = seccion(lb, ws, f, "Pasos editados a mano en el Editor avanzado", hasta="B")
    ediciones = [
        "Los pasos llevan nombre de negocio: VentasFiltradas, SinDuplicados, "
        "TiposAjustados, ClavesUnicas, EncabezadosPromovidos. La interfaz los habria "
        "llamado #\"Filas filtradas\", #\"Filas filtradas1\", #\"Tipo cambiado2\".",
        "Se comento cada bloque con // explicando la decision, no la sintaxis.",
        "Se consolido en un solo Table.SelectRows lo que la interfaz genera como tres "
        "pasos de filtrado encadenados.",
        "Se saco la ruta del codigo: ninguna consulta tiene una ruta escrita adentro, "
        "todas piden pRutaFuentes.",
        "Se agrego Culture = \"en-US\" a los tipados para que el libro abra igual con "
        "cualquier configuracion regional.",
        "Se reordeno el flujo para filtrar antes de tipar, que es lo eficiente y ademas "
        "evita que una celda vacia rompa la conversion.",
    ]
    for i, texto in enumerate(ediciones, start=1):
        lb.escribir(ws, f"B{f}", f"{i}.  {texto}", fuente=F_NOTA, alineacion=IZQ_ARRIBA)
        ws.row_dimensions[f].height = 30
        f += 1
    f += 1

    f = seccion(lb, ws, f, "Section1.m", hasta="B")
    for linea in consultas_m.SECCION.split("\n"):
        es_comentario = linea.lstrip().startswith("//")
        lb.escribir(ws, f"B{f}", linea or " ",
                    fuente=F_CODIGO_COM if es_comentario else F_CODIGO,
                    alineacion=Alignment(horizontal="left", vertical="center"))
        ws.row_dimensions[f].height = 12.5
        f += 1
    return ws


# --- hoja Estadistica_Descriptiva -------------------------------------------

def hoja_estadistica(lb, ventas, fila_datos_desde, fila_datos_hasta):
    """
    Reproduce la salida del Analysis ToolPak y la deja verificable.

    Bloque izquierdo (A1:D15): las mismas 14 filas que escribe
    Datos > Analisis de datos > Estadistica descriptiva, con valores estaticos,
    que es como el complemento entrega su reporte.

    Bloque derecho: las cuatro medidas que pide la consigna recalculadas con
    formulas nativas sobre la tabla de hechos, con una columna de control. Si
    las formulas y el reporte coinciden, el reporte es correcto.
    """
    ws = lb.hoja("Estadistica_Descriptiva",
                 anchos={"A": 27, "B": 16, "C": 27, "D": 16, "E": 4,
                         "F": 30, "G": 18, "H": 18, "I": 12, "J": 50})

    variables = [
        ("Venta_Neta", "L", [v["Venta_Neta"] for v in ventas]),
        ("Margen_Bruto", "N", [v["Margen_Bruto"] for v in ventas]),
    ]

    # --- bloque ToolPak ------------------------------------------------------
    for j, (nombre, _, valores) in enumerate(variables):
        col_etiqueta = get_column_letter(1 + j * 2)
        col_valor = get_column_letter(2 + j * 2)
        ws.merge_cells(f"{col_etiqueta}1:{col_valor}1")
        lb.escribir(ws, f"{col_etiqueta}1", nombre, fuente=F_CABECERA,
                    relleno=R_CABECERA, alineacion=CENTRO, borde=True)

        resultado = estadistica.describir(valores)
        for i, etiqueta in enumerate(estadistica.ETIQUETAS):
            fila = 2 + i
            lb.escribir(ws, f"{col_etiqueta}{fila}", etiqueta,
                        fuente=F_ETIQUETA, alineacion=IZQ, borde=True)
            valor = resultado[etiqueta]
            formato = FMT_ENTERO if etiqueta == "Cuenta" else FMT_MONEDA
            if etiqueta in ("Curtosis", "Coeficiente de asimetria"):
                formato = FMT_NUM4
            lb.escribir(ws, f"{col_valor}{fila}", valor,
                        formato=None if isinstance(valor, str) else formato,
                        alineacion=DER, borde=True)

    lb.escribir(ws, "A17",
                "Origen: Datos > Analisis de datos > Estadistica descriptiva, sobre "
                "Hechos_Ventas[Venta_Neta] y Hechos_Ventas[Margen_Bruto], con "
                "Rotulos en la primera fila y Resumen de estadisticas activados.",
                fuente=F_NOTA, alineacion=IZQ_ARRIBA)
    ws.merge_cells("A17:D18")
    ws.row_dimensions[17].height = 30

    # --- bloque de verificacion con formulas ---------------------------------
    lb.escribir(ws, "F1", "  Verificacion con formulas nativas",
                fuente=F_SECCION, relleno=R_SECCION, alineacion=IZQ, borde=True)
    ws.merge_cells("F1:I1")
    cabecera(lb, ws, 2, ["Estadistico", "Venta_Neta", "Margen_Bruto", "Coincide"],
             ["F", "G", "H", "I"])

    rango = {"Venta_Neta": f"Hechos_Ventas!$L${fila_datos_desde}:$L${fila_datos_hasta}",
             "Margen_Bruto": f"Hechos_Ventas!$N${fila_datos_desde}:$N${fila_datos_hasta}"}

    controles = [
        ("Media", "AVERAGE({r})", "Media"),
        ("Mediana", "MEDIAN({r})", "Mediana"),
        ("Desviacion estandar (muestral)", "STDEV.S({r})", "Desviacion estandar"),
        ("Rango (max - min)", "MAX({r})-MIN({r})", "Rango"),
    ]
    resultados = {n: estadistica.describir(v) for n, _, v in variables}
    fila = 3
    for etiqueta, plantilla, clave in controles:
        lb.escribir(ws, f"F{fila}", etiqueta, fuente=F_ETIQUETA, alineacion=IZQ, borde=True)
        for col, (nombre, _, _) in zip(("G", "H"), variables):
            lb.formula(ws, f"{col}{fila}",
                       "=" + plantilla.format(r=rango[nombre]),
                       resultados[nombre][clave], formato=FMT_MONEDA,
                       alineacion=DER, borde=True)
        lb.formula(ws, f"I{fila}",
                   f'=IF(AND(ABS(G{fila}-B{2 + estadistica.ETIQUETAS.index(clave)})<0.01,'
                   f'ABS(H{fila}-D{2 + estadistica.ETIQUETAS.index(clave)})<0.01),"OK","revisar")',
                   "OK", alineacion=CENTRO, borde=True, fuente=F_ETIQUETA_B)
        fila += 1

    # --- lectura del diagnostico --------------------------------------------
    r = resultados["Venta_Neta"]
    ratio = r["Media"] / r["Mediana"]
    fila += 1
    lb.escribir(ws, f"F{fila}", "  Lectura del diagnostico", fuente=F_SECCION,
                relleno=R_SECCION, alineacion=IZQ, borde=True)
    ws.merge_cells(f"F{fila}:J{fila}")
    fila += 1

    confianza = r["Nivel de confianza (95,0%)"]
    lecturas = [
        ("La media no resume esta variable",
         f"La media de la venta neta es USD {num(r['Media'], 2)} y la mediana USD "
         f"{num(r['Mediana'], 2)}: la media es {num(ratio, 1)} veces la mediana. La "
         f"mitad de las operaciones esta muy por debajo del promedio."),
        ("La distribucion esta fuertemente sesgada a la derecha",
         f"Coeficiente de asimetria {num(r['Coeficiente de asimetria'], 2)} y curtosis "
         f"{num(r['Curtosis'], 2)}. Son pocas operaciones corporativas muy grandes las "
         f"que estiran la cola: el maximo (USD {num(r['Maximo'], 2)}) es "
         f"{num(r['Maximo'] / r['Mediana'])} veces la mediana."),
        ("La dispersion es mayor que el nivel",
         f"La desviacion estandar (USD {num(r['Desviacion estandar'], 2)}) supera a la "
         f"media: el coeficiente de variacion es del "
         f"{pct(r['Desviacion estandar'] / r['Media'], 0)}. Cualquier proyeccion tiene "
         f"que trabajar con rangos, no con un unico numero puntual."),
        ("Por eso la simulacion se hace por escenarios",
         f"Con esta dispersion, planificar 2026 con un solo valor esperado seria "
         f"enganoso. De ahi el paso siguiente: tres escenarios y un Buscar Objetivo "
         f"que dice cuanto hay que vender para la meta de USD "
         f"{num(simulacion.META_BENEFICIO)}."),
        ("El intervalo de confianza de la media",
         f"Con 95% de confianza la venta neta promedio esta entre USD "
         f"{num(r['Media'] - confianza, 2)} y USD {num(r['Media'] + confianza, 2)} "
         f"(la media mas menos {num(confianza, 2)})."),
    ]
    for i, (que, detalle) in enumerate(lecturas):
        relleno = R_SUAVE if i % 2 else None
        lb.escribir(ws, f"F{fila}", que, fuente=F_ETIQUETA_B, relleno=relleno,
                    alineacion=IZQ_ARRIBA, borde=True)
        ws.merge_cells(f"G{fila}:J{fila}")
        lb.escribir(ws, f"G{fila}", detalle, fuente=F_NOTA, relleno=relleno,
                    alineacion=IZQ_ARRIBA, borde=True)
        ws.row_dimensions[fila].height = 46
        fila += 1
    return ws


# --- hoja Simulacion ---------------------------------------------------------

def hoja_simulacion(lb, base, vars_, proyecciones, objetivo, filas_param,
                    fila_datos_desde, fila_datos_hasta):
    ws = lb.hoja("Simulacion", anchos={"A": 2.5, "B": 46, "C": 18, "D": 62, "E": 3})
    f = titulo(lb, ws, "Simulacion financiera 2026",
               "Cierre 2025 como base, tres escenarios y la pregunta del "
               "directorio resuelta con Buscar Objetivo.")

    d, h = fila_datos_desde, fila_datos_hasta
    anio = "2025"

    def sumaproducto(col):
        return (f'=SUMPRODUCT((YEAR(Hechos_Ventas!$B${d}:$B${h})={anio})'
                f'*Hechos_Ventas!${col}${d}:${col}${h})')

    a25 = base["anio_2025"]

    # --- 1. cierre 2025 ------------------------------------------------------
    f = seccion(lb, ws, f, "1 | Cierre 2025: la base del plan", hasta="D")
    f = cabecera(lb, ws, f, ["Concepto", "Valor", "Como se calcula"], ["B", "C", "D"])

    fila_unidades = f
    filas = {}
    base_items = [
        ("unidades", "Unidades vendidas", sumaproducto("F"), a25["unidades"], FMT_ENTERO,
         "Suma de Unidades de la tabla de hechos, filtrando el ejercicio 2025."),
        ("venta", "Venta neta", sumaproducto("L"), a25["venta_neta"], FMT_MONEDA,
         "Suma de Venta_Neta, la columna que calcula Power Query."),
        ("costo", "Costo de ventas", sumaproducto("M"), a25["costo_ventas"], FMT_MONEDA,
         "Suma de Costo_Total."),
        ("envio", "Costo logistico (envios)", sumaproducto("J"), a25["costo_envio"],
         FMT_MONEDA, "Suma de Costo_Envio. Es la variable que mas mueve el margen."),
    ]
    for clave, etiqueta, formula, valor, formato, detalle in base_items:
        lb.escribir(ws, f"B{f}", etiqueta, fuente=F_ETIQUETA, alineacion=IZQ, borde=True)
        lb.formula(ws, f"C{f}", formula, valor, formato=formato, alineacion=DER, borde=True)
        lb.escribir(ws, f"D{f}", detalle, fuente=F_NOTA, alineacion=IZQ_ARRIBA, borde=True)
        filas[clave] = f
        f += 1

    derivados = [
        ("precio", "Precio neto promedio por unidad",
         f"=C{filas['venta']}/C{filas['unidades']}", base["precio_neto_promedio"],
         FMT_MONEDA, "Venta neta dividida unidades."),
        ("costo_unit", "Costo unitario promedio",
         f"=C{filas['costo']}/C{filas['unidades']}", base["costo_unitario_promedio"],
         FMT_MONEDA, "Costo de ventas dividido unidades."),
        ("tasa_log", "Tasa de costo logistico observada",
         f"=C{filas['envio']}/C{filas['venta']}", base["tasa_logistica"],
         FMT_PCT2, "Costo de envios sobre venta neta. Es el punto de partida de los escenarios."),
        ("fijos", "Gastos fijos anuales de estructura",
         f"=C{filas['venta']}*Parametros!$C${filas_param['fila_ratio_fijos']}",
         base["gastos_fijos"], FMT_MONEDA,
         "Venta neta 2025 por el ratio de estructura de la hoja Parametros."),
    ]
    for clave, etiqueta, formula, valor, formato, detalle in derivados:
        lb.escribir(ws, f"B{f}", etiqueta, fuente=F_ETIQUETA, alineacion=IZQ, borde=True)
        lb.formula(ws, f"C{f}", formula, valor, formato=formato, alineacion=DER, borde=True)
        lb.escribir(ws, f"D{f}", detalle, fuente=F_NOTA, alineacion=IZQ_ARRIBA, borde=True)
        filas[clave] = f
        f += 1
    f += 1

    # --- 2. celdas cambiantes ------------------------------------------------
    f = seccion(lb, ws, f, "2 | Variables del escenario (celdas cambiantes)", hasta="D")
    f = nota(lb, ws, f,
             "Estas tres celdas son las que mueve el Administrador de escenarios. "
             "Ahora muestran el escenario Base.", hasta="D")
    f = cabecera(lb, ws, f, ["Variable", "Valor", "Que representa"], ["B", "C", "D"])

    base_vars = vars_["Base"]
    cambiantes = [
        ("Var_Demanda", "Variacion de la demanda 2026", base_vars[0], FMT_PCT1,
         "Cuanto crecen o caen las unidades vendidas respecto de 2025."),
        ("Tasa_Logistica", "Tasa de costo logistico", base_vars[1], FMT_PCT2,
         "Costo de envios como porcentaje de la venta neta."),
        ("Tasa_Interes", "Tasa de interes financiero", base_vars[2], FMT_PCT1,
         "Costo del capital de trabajo como porcentaje de la venta neta."),
    ]
    celdas_cambiantes = []
    for nombre, etiqueta, valor, formato, detalle in cambiantes:
        lb.escribir(ws, f"B{f}", etiqueta, fuente=F_ETIQUETA_B, relleno=R_CAMBIANTE,
                    alineacion=IZQ, borde=True)
        lb.escribir(ws, f"C{f}", valor, formato=formato, fuente=F_ETIQUETA_B,
                    relleno=R_CAMBIANTE, alineacion=DER, borde=True)
        lb.escribir(ws, f"D{f}", detalle, fuente=F_NOTA, alineacion=IZQ_ARRIBA, borde=True)
        lb.wb.defined_names.add(DefinedName(nombre, attr_text=f"Simulacion!$C${f}"))
        celdas_cambiantes.append((nombre, f"C{f}", formato))
        filas[nombre] = f
        f += 1
    f += 1

    # --- 3. proyeccion -------------------------------------------------------
    p = proyecciones["Base"]
    f = seccion(lb, ws, f, "3 | Cuenta de resultados proyectada 2026", hasta="D")
    f = cabecera(lb, ws, f, ["Concepto", "Proyeccion 2026", "Formula"], ["B", "C", "D"])

    proy = [
        ("unidades_p", "Unidades a vender",
         f"=C{filas['unidades']}*(1+C{filas['Var_Demanda']})", p["unidades"], FMT_ENTERO,
         "Unidades 2025 por (1 + variacion de la demanda)", False),
        ("ventas_p", "Venta neta",
         None, p["ventas"], FMT_MONEDA, "Unidades por precio neto promedio", True),
        ("costo_p", "Costo de ventas",
         None, p["costo_ventas"], FMT_MONEDA, "Unidades por costo unitario promedio", False),
        ("margen_p", "Margen bruto",
         None, p["margen_bruto"], FMT_MONEDA, "Venta neta menos costo de ventas", True),
        ("log_p", "Costo logistico",
         None, p["costo_logistico"], FMT_MONEDA, "Venta neta por tasa de costo logistico", False),
        ("fijos_p", "Gastos fijos de estructura",
         None, p["gastos_fijos"], FMT_MONEDA, "Constante: no varia con el volumen", False),
        ("fin_p", "Costo financiero",
         None, p["costo_financiero"], FMT_MONEDA, "Venta neta por tasa de interes", False),
        ("beneficio_p", "Beneficio neto",
         None, p["beneficio"], FMT_MONEDA,
         "Margen bruto menos logistica, estructura y costo financiero", True),
        ("margen_neto_p", "Margen neto",
         None, p["margen_neto"], FMT_PCT2, "Beneficio neto sobre venta neta", True),
    ]

    fila_inicio_proy = f
    for clave, etiqueta, formula, valor, formato, detalle, destacar in proy:
        filas[clave] = f
        f += 1
    f = fila_inicio_proy

    formulas_proy = {
        "unidades_p": f"=C{filas['unidades']}*(1+C{filas['Var_Demanda']})",
        "ventas_p": f"=C{filas['unidades_p']}*C{filas['precio']}",
        "costo_p": f"=C{filas['unidades_p']}*C{filas['costo_unit']}",
        "margen_p": f"=C{filas['ventas_p']}-C{filas['costo_p']}",
        "log_p": f"=C{filas['ventas_p']}*C{filas['Tasa_Logistica']}",
        "fijos_p": f"=C{filas['fijos']}",
        "fin_p": f"=C{filas['ventas_p']}*C{filas['Tasa_Interes']}",
        "beneficio_p": (f"=C{filas['margen_p']}-C{filas['log_p']}"
                        f"-C{filas['fijos_p']}-C{filas['fin_p']}"),
        "margen_neto_p": f"=C{filas['beneficio_p']}/C{filas['ventas_p']}",
    }
    for clave, etiqueta, _, valor, formato, detalle, destacar in proy:
        relleno = R_AZUL_SUAVE if destacar else None
        lb.escribir(ws, f"B{f}", etiqueta,
                    fuente=F_ETIQUETA_B if destacar else F_ETIQUETA,
                    relleno=relleno, alineacion=IZQ, borde=True)
        lb.formula(ws, f"C{f}", formulas_proy[clave], valor, formato=formato,
                   fuente=F_ETIQUETA_B if destacar else F_VALOR,
                   relleno=relleno, alineacion=DER, borde=True)
        lb.escribir(ws, f"D{f}", detalle, fuente=F_NOTA, alineacion=IZQ_ARRIBA, borde=True)
        f += 1

    for nombre, clave in [("Ventas_Netas_2026", "ventas_p"),
                          ("Margen_Bruto_2026", "margen_p"),
                          ("Beneficio_Neto_2026", "beneficio_p"),
                          ("Margen_Neto_2026", "margen_neto_p")]:
        lb.wb.defined_names.add(DefinedName(nombre, attr_text=f"Simulacion!$C${filas[clave]}"))
    f += 1

    # --- 4. buscar objetivo --------------------------------------------------
    f = seccion(lb, ws, f, "4 | Buscar Objetivo: llegar a un beneficio neto de "
                           f"USD {num(simulacion.META_BENEFICIO)}", hasta="D")
    f = nota(lb, ws, f,
             "Datos > Analisis de hipotesis > Buscar objetivo. Definir la celda del "
             "beneficio neto de este bloque, con el valor de la meta, cambiando la celda "
             "de unidades a vender. El bloque usa constantes propias tomadas del cierre "
             "2025, asi que el resultado no se mueve cuando se cambia de escenario.",
             hasta="D", alto=44)
    f = cabecera(lb, ws, f, ["Concepto", "Valor", "Detalle"], ["B", "C", "D"])

    fo = {}
    entradas = [
        ("meta", "Meta de beneficio neto",
         f"=Parametros!$C${filas_param['fila_meta']}", objetivo["meta"], FMT_MONEDA,
         "La define el directorio. Vive en la hoja Parametros.", False),
        ("unidades_o", "Unidades a vender  (celda que cambia)",
         None, objetivo["unidades"], FMT_ENTERO,
         "Constante. Es el valor que encontro Buscar Objetivo.", "cambiante"),
        ("precio_o", "Precio neto promedio", f"=C{filas['precio']}", objetivo["precio"],
         FMT_MONEDA, "Del cierre 2025.", False),
        ("costo_o", "Costo unitario promedio", f"=C{filas['costo_unit']}",
         objetivo["costo_unitario"], FMT_MONEDA, "Del cierre 2025.", False),
        ("tasa_o", "Tasa de costo logistico", f"=C{filas['tasa_log']}",
         objetivo["tasa_logistica"], FMT_PCT2, "La observada en 2025, sin escenario.", False),
        ("interes_o", "Tasa de interes financiero",
         f"=Parametros!$C${filas_param['fila_interes']}", objetivo["tasa_interes"],
         FMT_PCT1, "Tasa de referencia de la hoja Parametros.", False),
        ("fijos_o", "Gastos fijos anuales", f"=C{filas['fijos']}",
         objetivo["gastos_fijos"], FMT_MONEDA, "Del cierre 2025.", False),
    ]
    for clave, etiqueta, formula, valor, formato, detalle, marcar in entradas:
        fo[clave] = f
        relleno = R_CAMBIANTE if marcar == "cambiante" else None
        fuente = F_ETIQUETA_B if marcar else F_ETIQUETA
        lb.escribir(ws, f"B{f}", etiqueta, fuente=fuente, relleno=relleno,
                    alineacion=IZQ, borde=True)
        if formula is None:
            lb.escribir(ws, f"C{f}", valor, formato=formato, fuente=F_ETIQUETA_B,
                        relleno=relleno, alineacion=DER, borde=True)
        else:
            lb.formula(ws, f"C{f}", formula, valor, formato=formato,
                       fuente=fuente, relleno=relleno, alineacion=DER, borde=True)
        lb.escribir(ws, f"D{f}", detalle, fuente=F_NOTA, alineacion=IZQ_ARRIBA, borde=True)
        f += 1

    salidas = [
        ("ventas_o", "Venta neta necesaria",
         f"=C{fo['unidades_o']}*C{fo['precio_o']}",
         objetivo["unidades"] * objetivo["precio"], FMT_MONEDA, False),
        ("cventas_o", "Costo de ventas",
         f"=C{fo['unidades_o']}*C{fo['costo_o']}",
         objetivo["unidades"] * objetivo["costo_unitario"], FMT_MONEDA, False),
        ("clog_o", "Costo logistico", None, None, FMT_MONEDA, False),
        ("cfin_o", "Costo financiero", None, None, FMT_MONEDA, False),
        ("cfijos_o", "Gastos fijos", f"=C{fo['fijos_o']}", objetivo["gastos_fijos"],
         FMT_MONEDA, False),
        ("beneficio_o", "Beneficio neto alcanzado", None, None, FMT_MONEDA, True),
        ("dif_o", "Diferencia contra la meta", None, None, FMT_MONEDA, False),
    ]
    for clave, *_ in salidas:
        fo[clave] = f
        f += 1

    ventas_o = objetivo["unidades"] * objetivo["precio"]
    valores_salida = {
        "ventas_o": (f"=C{fo['unidades_o']}*C{fo['precio_o']}", ventas_o,
                     "Unidades por precio neto promedio"),
        "cventas_o": (f"=C{fo['unidades_o']}*C{fo['costo_o']}",
                      objetivo["unidades"] * objetivo["costo_unitario"],
                      "Unidades por costo unitario"),
        "clog_o": (f"=C{fo['ventas_o']}*C{fo['tasa_o']}",
                   ventas_o * objetivo["tasa_logistica"], "Venta por tasa logistica"),
        "cfin_o": (f"=C{fo['ventas_o']}*C{fo['interes_o']}",
                   ventas_o * objetivo["tasa_interes"], "Venta por tasa de interes"),
        "cfijos_o": (f"=C{fo['fijos_o']}", objetivo["gastos_fijos"], "Constante"),
        "beneficio_o": (f"=C{fo['ventas_o']}-C{fo['cventas_o']}-C{fo['clog_o']}"
                        f"-C{fo['cfin_o']}-C{fo['cfijos_o']}",
                        simulacion.META_BENEFICIO,
                        "Venta menos costo de ventas, logistica, financiero y estructura"),
        "dif_o": (f"=C{fo['beneficio_o']}-C{fo['meta']}", 0.0,
                  "Tiene que dar cero: es la comprobacion de Buscar Objetivo"),
    }
    for clave, etiqueta, _, _, formato, destacar in salidas:
        formula, valor, detalle = valores_salida[clave]
        relleno = R_AZUL_SUAVE if destacar else None
        lb.escribir(ws, f"B{fo[clave]}", etiqueta,
                    fuente=F_ETIQUETA_B if destacar else F_ETIQUETA,
                    relleno=relleno, alineacion=IZQ, borde=True)
        lb.formula(ws, f"C{fo[clave]}", formula, valor, formato=formato,
                   fuente=F_ETIQUETA_B if destacar else F_VALOR,
                   relleno=relleno, alineacion=DER, borde=True)
        lb.escribir(ws, f"D{fo[clave]}", detalle, fuente=F_NOTA,
                    alineacion=IZQ_ARRIBA, borde=True)

    for etiqueta, formula, valor, formato in [
        ("Unidades adicionales sobre 2025",
         f"=C{fo['unidades_o']}-C{filas['unidades']}",
         objetivo["unidades"] - objetivo["unidades_2025"], FMT_ENTERO),
        ("Crecimiento necesario sobre 2025",
         f"=C{fo['unidades_o']}/C{filas['unidades']}-1",
         objetivo["variacion_unidades"], FMT_PCT1),
        ("Crecimiento necesario sobre el plan base",
         f"=C{fo['unidades_o']}/C{filas['unidades_p']}-1",
         objetivo["unidades"] / (objetivo["unidades_2025"] * (1 + simulacion.PLAN_DEMANDA["Base"])) - 1,
         FMT_PCT1),
    ]:
        lb.escribir(ws, f"B{f}", etiqueta, fuente=F_ETIQUETA, alineacion=IZQ, borde=True)
        lb.formula(ws, f"C{f}", formula, valor, formato=formato, alineacion=DER, borde=True)
        lb.escribir(ws, f"D{f}", "", borde=True)
        f += 1
    f += 1

    sobre_base = objetivo["unidades"] / objetivo["unidades_base"] - 1
    conclusion = (
        f"Respuesta a la gerencia: para cerrar 2026 con un beneficio neto de USD "
        f"{num(simulacion.META_BENEFICIO)} hay que vender {num(objetivo['unidades'])} "
        f"unidades, un {pct(objetivo['variacion_unidades'])} mas que en 2025 y un "
        f"{pct(sobre_base)} por encima del plan base. La meta queda entre el escenario "
        f"Base (USD {num(proyecciones['Base']['beneficio'])}) y el Optimista (USD "
        f"{num(proyecciones['Optimista']['beneficio'])}): es alcanzable, pero no con el "
        f"plan conservador."
    )
    f = nota(lb, ws, f, conclusion, hasta="D", alto=48)

    return {"celdas_cambiantes": celdas_cambiantes,
            "filas": filas,
            "resultado": [("Ventas_Netas_2026", f"C{filas['ventas_p']}", FMT_MONEDA),
                          ("Margen_Bruto_2026", f"C{filas['margen_p']}", FMT_MONEDA),
                          ("Beneficio_Neto_2026", f"C{filas['beneficio_p']}", FMT_MONEDA),
                          ("Margen_Neto_2026", f"C{filas['margen_neto_p']}", FMT_PCT2)]}


# --- hoja Resumen de escenarios ---------------------------------------------

NOTA_EXCEL = (
    "Notas: La columna Valores actuales representa los valores de las celdas cambiantes "
    "en el momento en que se creo el Informe resumen de escenarios. Las celdas cambiantes "
    "de cada escenario se muestran en gris."
)


def hoja_resumen(lb, vars_, proyecciones, info_sim):
    ws = lb.hoja("Resumen de escenarios",
                 anchos={"A": 2.5, "B": 4, "C": 30, "D": 20, "E": 18, "F": 18,
                         "G": 18, "H": 4, "I": 56})
    orden = ["Pesimista", "Base", "Optimista"]

    lb.escribir(ws, "C2", "Resumen de escenarios", fuente=F_TITULO, alineacion=IZQ)
    ws.row_dimensions[2].height = 28

    f = 4
    lb.escribir(ws, f"C{f}", "", relleno=R_CABECERA, borde=True)
    lb.escribir(ws, f"D{f}", "Valores actuales:", fuente=F_CABECERA,
                relleno=R_CABECERA, alineacion=CENTRO, borde=True)
    for col, nombre in zip("EFG", orden):
        lb.escribir(ws, f"{col}{f}", nombre, fuente=F_CABECERA, relleno=R_CABECERA,
                    alineacion=CENTRO, borde=True)
    f += 1

    lb.escribir(ws, f"C{f}", "Celdas cambiantes:", fuente=F_ETIQUETA_B,
                relleno=R_SUAVE, alineacion=IZQ, borde=True)
    for col in "DEFG":
        lb.escribir(ws, f"{col}{f}", "", relleno=R_SUAVE, borde=True)
    f += 1

    etiquetas_cambiantes = {"Var_Demanda": 0, "Tasa_Logistica": 1, "Tasa_Interes": 2}
    for nombre, _, formato in info_sim["celdas_cambiantes"]:
        idx = etiquetas_cambiantes[nombre]
        lb.escribir(ws, f"C{f}", "    " + nombre, fuente=F_CODIGO, alineacion=IZQ, borde=True)
        lb.escribir(ws, f"D{f}", vars_["Base"][idx], formato=formato,
                    alineacion=DER, borde=True)
        for col, escenario in zip("EFG", orden):
            lb.escribir(ws, f"{col}{f}", vars_[escenario][idx], formato=formato,
                        relleno=R_SUAVE, alineacion=DER, borde=True)
        f += 1

    lb.escribir(ws, f"C{f}", "Celdas de resultado:", fuente=F_ETIQUETA_B,
                relleno=R_SUAVE, alineacion=IZQ, borde=True)
    for col in "DEFG":
        lb.escribir(ws, f"{col}{f}", "", relleno=R_SUAVE, borde=True)
    f += 1

    claves = {"Ventas_Netas_2026": "ventas", "Margen_Bruto_2026": "margen_bruto",
              "Beneficio_Neto_2026": "beneficio", "Margen_Neto_2026": "margen_neto"}
    for nombre, _, formato in info_sim["resultado"]:
        clave = claves[nombre]
        destacar = nombre == "Beneficio_Neto_2026"
        fuente = F_DESTACADO if destacar else F_CODIGO
        lb.escribir(ws, f"C{f}", "    " + nombre, fuente=fuente, alineacion=IZQ, borde=True)
        lb.escribir(ws, f"D{f}", proyecciones["Base"][clave], formato=formato,
                    alineacion=DER, borde=True,
                    fuente=F_ETIQUETA_B if destacar else F_VALOR)
        for col, escenario in zip("EFG", orden):
            lb.escribir(ws, f"{col}{f}", proyecciones[escenario][clave], formato=formato,
                        alineacion=DER, borde=True,
                        relleno=R_AZUL_SUAVE if destacar else None,
                        fuente=F_ETIQUETA_B if destacar else F_VALOR)
        f += 1

    f += 1
    ws.merge_cells(f"C{f}:G{f + 1}")
    lb.escribir(ws, f"C{f}", NOTA_EXCEL, fuente=F_NOTA, alineacion=IZQ_ARRIBA)
    ws.row_dimensions[f].height = 30
    f += 2
    ws.merge_cells(f"C{f}:G{f + 1}")
    lb.escribir(ws, f"C{f}",
                "Los tres escenarios estan guardados en el libro. Para regenerar este "
                "resumen: hoja Simulacion > Datos > Analisis de hipotesis > Administrador "
                "de escenarios > Resumen, con las celdas de resultado "
                "Ventas_Netas_2026, Margen_Bruto_2026, Beneficio_Neto_2026 y "
                "Margen_Neto_2026.", fuente=F_NOTA, alineacion=IZQ_ARRIBA)
    ws.row_dimensions[f].height = 30
    f += 3

    # --- lectura gerencial (agregado propio, no lo escribe Excel) -----------
    f = seccion(lb, ws, f, "Lectura del resumen", desde="C", hasta="G")
    meta = simulacion.META_BENEFICIO
    base_b = proyecciones["Base"]["beneficio"]
    opt_b = proyecciones["Optimista"]["beneficio"]
    pes_b = proyecciones["Pesimista"]["beneficio"]
    lecturas = [
        ("La meta no se alcanza con el plan base",
         f"El escenario Base cierra en USD {num(base_b)} de beneficio neto: faltan USD "
         f"{num(meta - base_b)} para la meta de USD {num(meta)}, un "
         f"{pct(meta / base_b - 1)} mas."),
        ("Solo el escenario optimista supera la meta",
         f"Con USD {num(opt_b)}, el Optimista deja un margen de USD {num(opt_b - meta)} "
         f"sobre el objetivo. Es el unico de los tres que lo logra."),
        ("El piso del rango es la mitad del techo",
         f"Entre Pesimista (USD {num(pes_b)}) y Optimista (USD {num(opt_b)}) hay USD "
         f"{num(opt_b - pes_b)} de diferencia: el resultado de 2026 depende mas de estas "
         f"tres variables que de cualquier otra decision del ejercicio."),
        ("La palanca que la empresa si controla",
         f"Entre Base y Optimista la tasa logistica baja "
         f"{num((vars_['Base'][1] - vars_['Optimista'][1]) * 100, 2)} puntos y la de "
         f"interes {num((vars_['Base'][2] - vars_['Optimista'][2]) * 100, 1)}. Esas dos "
         f"se negocian con proveedores y bancos; la demanda no se controla."),
    ]
    for i, (que, detalle) in enumerate(lecturas):
        relleno = R_SUAVE if i % 2 else None
        lb.escribir(ws, f"C{f}", que, fuente=F_ETIQUETA_B, relleno=relleno,
                    alineacion=IZQ_ARRIBA, borde=True)
        ws.merge_cells(f"D{f}:I{f}")
        lb.escribir(ws, f"D{f}", detalle, fuente=F_NOTA,
                    relleno=relleno, alineacion=IZQ_ARRIBA, borde=True)
        ws.row_dimensions[f].height = 44
        f += 1
    return ws


# --- orquestacion ------------------------------------------------------------

def main(destino=None):
    d = datos.construir()
    ventas = datos.enriquecer(d["ventas"])

    base = simulacion.base_desde_datos(ventas)
    vars_ = simulacion.variables(base)
    proyecciones = {n: simulacion.proyectar(base, *vars_[n]) for n in vars_}
    objetivo = simulacion.buscar_objetivo(base)

    lb = Libro()

    resumen_datos = {
        "crudas": len(d["crudas"]),
        "limpias": len(ventas),
        "venta_neta": sum(v["Venta_Neta"] for v in ventas),
    }
    hoja_inicio(lb, resumen_datos)
    filas_param = hoja_parametros(lb, base)
    hoja_modelo(lb)
    hoja_dax(lb)
    hoja_codigo_m(lb)

    fila_desde, fila_hasta = 4, len(ventas) + 3

    hoja_estadistica(lb, ventas, fila_desde, fila_hasta)
    info_sim = hoja_simulacion(lb, base, vars_, proyecciones, objetivo,
                               filas_param, fila_desde, fila_hasta)
    hoja_resumen(lb, vars_, proyecciones, info_sim)

    hoja_datos(lb, "Hechos_Ventas", FACT_COLS, ventas, "tbl_Hechos",
               {"A": 11, "B": 12, "C": 11, "D": 12, "E": 14, "F": 10, "G": 15,
                "H": 11, "I": 14, "J": 12, "K": 13, "L": 13, "M": 13, "N": 13},
               FORMATOS_FACT,
               "Resultado de la consulta Hechos_Ventas (Power Query). "
               "5.200 operaciones confirmadas, 2024-2025. Copia en hoja para poder "
               "auditar el modelo y para que el Analysis ToolPak tenga un rango que leer.")
    hoja_datos(lb, "Dim_Clientes",
               ["ID_Cliente", "Cliente", "Segmento", "Region", "Ciudad", "Alta"],
               d["clientes"], "tbl_Clientes_dim",
               {"A": 12, "B": 26, "C": 20, "D": 12, "E": 16, "F": 12}, {},
               "Resultado de la consulta Dim_Clientes. PK: ID_Cliente (60 claves unicas).")
    hoja_datos(lb, "Dim_Productos",
               ["ID_Producto", "Producto", "Categoria", "Marca", "Precio_Lista",
                "Costo_Estandar"],
               d["productos"], "tbl_Productos_dim",
               {"A": 13, "B": 22, "C": 16, "D": 12, "E": 13, "F": 15},
               {"Precio_Lista": FMT_MONEDA, "Costo_Estandar": FMT_MONEDA},
               "Resultado de la consulta Dim_Productos. PK: ID_Producto (10 claves unicas).")
    hoja_datos(lb, "Dim_Calendario",
               ["Fecha", "Anio", "Trimestre", "Mes", "Nombre_Mes", "Anio_Mes",
                "Dia_Semana", "Es_Fin_Semana"],
               d["calendario"], "tbl_Calendario_dim",
               {"A": 12, "B": 9, "C": 11, "D": 8, "E": 14, "F": 11, "G": 13, "H": 15},
               {"Fecha": FMT_FECHA},
               "Resultado de la consulta Dim_Calendario, generada en M. "
               "PK: Fecha (731 dias, anios completos 2024-2025).")

    lb.wb.active = 0
    lb.wb.properties.title = ("Pre-entrega: Optimizacion y Modelado de Datos "
                              "para Toma de Decisiones")
    lb.wb.properties.subject = ("Modulo 8 - Arquitectura del dato, optimizacion "
                                "de modelos y simulacion de escenarios")
    lb.wb.properties.creator = "Lleyton Murphy"
    lb.wb.properties.description = (
        "Flujo ETL en Power Query, modelo en estrella para Power Pivot, medidas "
        "DAX con VAR/RETURN, estadistica descriptiva y simulacion de escenarios.")

    destino = destino or os.path.join(RAIZ, NOMBRE_ARCHIVO)
    lb.wb.save(destino)

    return {
        "ruta": destino,
        "cache": lb.cache,
        "escenarios": {"celdas": info_sim["celdas_cambiantes"], "valores": vars_},
        "base": base,
        "proyecciones": proyecciones,
        "objetivo": objetivo,
        "resumen_datos": resumen_datos,
    }


if __name__ == "__main__":
    info = main()
    print("libro ->", info["ruta"])
