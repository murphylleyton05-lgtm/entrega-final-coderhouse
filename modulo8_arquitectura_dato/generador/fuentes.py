"""
Escribe los archivos de origen que consume Power Query.

    fuentes/ventas_historico.csv   -> archivo transaccional historico (hechos)
    fuentes/maestros.xlsx          -> libro local con las tablas maestras

Son dos formatos distintos a proposito: la consigna pide combinar un `.csv`
transaccional con tablas maestras guardadas en un `.xlsx` local.

Detalles que importan para que el ETL sea reproducible en cualquier maquina:

  * el CSV se escribe en UTF-8 con separador coma y punto decimal, y las
    consultas M lo leen con `Culture="en-US"`. Asi el libro abre igual en un
    Excel con configuracion regional en espanol que en uno en ingles.
  * las fechas van en ISO (yyyy-mm-dd), que no depende del orden dia/mes.
  * las tablas del `.xlsx` son tablas con nombre (`tbl_Clientes`,
    `tbl_Productos`), no rangos sueltos: asi el M las referencia por nombre y
    no por posicion.
"""
import csv
import os

import openpyxl
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.table import Table, TableStyleInfo

import datos

AQUI = os.path.dirname(os.path.abspath(__file__))
FUENTES = os.path.join(os.path.dirname(AQUI), "fuentes")

COLUMNAS_CSV = ["ID_Venta", "Fecha", "ID_Cliente", "ID_Producto", "Canal",
                "Estado", "Unidades", "Precio_Unitario", "Descuento",
                "Costo_Unitario", "Costo_Envio"]


def _sucio(valor, indice, columna):
    """
    Ensucia el texto de algunas celdas numericas con espacios sobrantes.

    Es el defecto mas comun de un export real y obliga a que la consulta tipe
    las columnas explicitamente en vez de confiar en la deteccion automatica.
    """
    if columna in ("Unidades", "Precio_Unitario", "Costo_Envio") and indice % 17 == 0:
        return f"  {valor} "
    return valor


def escribir_csv(crudas, ruta):
    with open(ruta, "w", newline="", encoding="utf-8") as fh:
        w = csv.writer(fh, delimiter=",", lineterminator="\n")
        w.writerow(COLUMNAS_CSV)
        for i, fila in enumerate(crudas):
            valores = []
            for col in COLUMNAS_CSV:
                v = fila[col]
                if col == "Fecha":
                    v = v.isoformat()
                valores.append(_sucio(v, i, col))
            w.writerow(valores)
    return ruta


def _hoja_tabla(wb, nombre, columnas, filas, titulo_tabla, anchos):
    ws = wb.create_sheet(nombre)
    ws.append(columnas)
    for fila in filas:
        ws.append([fila[c] for c in columnas])
    ref = f"A1:{get_column_letter(len(columnas))}{len(filas) + 1}"
    tabla = Table(displayName=titulo_tabla, ref=ref)
    tabla.tableStyleInfo = TableStyleInfo(
        name="TableStyleMedium2", showRowStripes=True)
    ws.add_table(tabla)
    for i, ancho in enumerate(anchos, start=1):
        ws.column_dimensions[get_column_letter(i)].width = ancho
    ws.freeze_panes = "A2"
    return ws


def escribir_maestros(clientes, productos, ruta):
    wb = openpyxl.Workbook()
    wb.remove(wb.active)
    _hoja_tabla(wb, "Clientes",
                ["ID_Cliente", "Cliente", "Segmento", "Region", "Ciudad", "Alta"],
                clientes, "tbl_Clientes", [12, 26, 20, 12, 16, 12])
    _hoja_tabla(wb, "Productos",
                ["ID_Producto", "Producto", "Categoria", "Marca",
                 "Precio_Lista", "Costo_Estandar"],
                productos, "tbl_Productos", [13, 22, 16, 12, 13, 15])
    wb.save(ruta)
    return ruta


def main():
    os.makedirs(FUENTES, exist_ok=True)
    d = datos.construir()
    csv_ruta = escribir_csv(d["crudas"], os.path.join(FUENTES, "ventas_historico.csv"))
    xlsx_ruta = escribir_maestros(d["clientes"], d["productos"],
                                  os.path.join(FUENTES, "maestros.xlsx"))
    print("csv     ->", csv_ruta, f"({len(d['crudas'])} filas)")
    print("maestros->", xlsx_ruta)
    return d


if __name__ == "__main__":
    main()
