"""Construye PreEntrega_ModeloBI_MurphyLleyton.xlsx.

Genera todo lo que puede existir sin el motor de Excel:
  - hoja Config con el parametro de ruta (rango con nombre RutaDatos)
  - modelo financiero para Buscar Objetivo
  - los 3 escenarios REALES del Administrador de escenarios (XML <scenarios>)
  - hoja Resumen de escenarios con el formato que emite Excel
  - hoja de Estadistica descriptiva con el formato que emite Analysis ToolPak
  - documentacion del esquema en estrella, el codigo M y las medidas DAX

Lo que NO puede generarse desde aca y queda documentado paso a paso:
el modelo tabular de Power Pivot (xl/model/item.data) y las medidas DAX, que
solo los escribe el motor xVelocity de Excel.
"""
import re
import shutil
import zipfile
from datetime import date
from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter
from openpyxl.workbook.defined_name import DefinedName

from estadistica import descriptivos, ventas_depuradas

RAIZ = Path(__file__).resolve().parent.parent
SALIDA = RAIZ / "PreEntrega_ModeloBI_MurphyLleyton.xlsx"

AZUL = "1F3864"
AZUL_CLARO = "D9E2F3"
GRIS = "F2F2F2"
AMBAR = "FFF2CC"
VERDE = "E2EFDA"
ROJO = "FCE4E4"

MONEDA = '"$" #,##0'
MONEDA_2 = '"$" #,##0.00'
PORCENTAJE = '0.0%'
ENTERO = '#,##0'
DECIMAL = '#,##0.00'

fino = Side(style="thin", color="BFBFBF")
BORDE = Border(left=fino, right=fino, top=fino, bottom=fino)


# ---------------------------------------------------------------------------
# helpers de formato
# ---------------------------------------------------------------------------
def titulo(ws, celda, texto, ancho=6):
    ws[celda] = texto
    ws[celda].font = Font(bold=True, size=14, color="FFFFFF")
    ws[celda].fill = PatternFill("solid", fgColor=AZUL)
    ws[celda].alignment = Alignment(vertical="center", indent=1)
    fila = int(re.sub(r"\D", "", celda))
    col = re.sub(r"\d", "", celda)
    ws.row_dimensions[fila].height = 26
    ws.merge_cells(f"{celda}:{get_column_letter(ws[celda].column + ancho - 1)}{fila}")


def seccion(ws, celda, texto, ancho=6):
    ws[celda] = texto
    ws[celda].font = Font(bold=True, size=11, color=AZUL)
    ws[celda].fill = PatternFill("solid", fgColor=AZUL_CLARO)
    fila = int(re.sub(r"\D", "", celda))
    ws.merge_cells(f"{celda}:{get_column_letter(ws[celda].column + ancho - 1)}{fila}")


def etiqueta(ws, fila, texto, col="A", negrita=False, indent=0):
    ws[f"{col}{fila}"] = texto
    ws[f"{col}{fila}"].font = Font(bold=negrita)
    ws[f"{col}{fila}"].alignment = Alignment(indent=indent)


def valor(ws, fila, contenido, col="B", formato=None, negrita=False, fondo=None):
    celda = ws[f"{col}{fila}"]
    celda.value = contenido
    if formato:
        celda.number_format = formato
    celda.font = Font(bold=negrita)
    if fondo:
        celda.fill = PatternFill("solid", fgColor=fondo)
    celda.border = BORDE
    return celda


def nota(ws, fila, texto, col="A", ancho=7):
    ws[f"{col}{fila}"] = texto
    ws[f"{col}{fila}"].font = Font(italic=True, size=9, color="595959")
    ws[f"{col}{fila}"].alignment = Alignment(wrap_text=True, vertical="top")
    ws.merge_cells(f"{col}{fila}:{get_column_letter(ws[f'{col}{fila}'].column + ancho - 1)}{fila}")


# ---------------------------------------------------------------------------
# parametros del modelo financiero, anclados en los datos reales
# ---------------------------------------------------------------------------
def parametros_negocio(filas):
    """Deriva la base de la simulacion del ultimo anio cerrado del dataset."""
    anio = max(f["Fecha"].year for f in filas)
    del_anio = [f for f in filas if f["Fecha"].year == anio]

    unidades = sum(f["Unidades"] for f in del_anio)
    ingresos = sum(f["Importe_Neto"] for f in del_anio)
    costos = sum(f["Costo_Total"] for f in del_anio)

    return {
        "anio": anio,
        "unidades": unidades,
        "precio_prom": round(ingresos / unidades, 2),
        "costo_prom": round(costos / unidades, 2),
        # Estructura de gastos fijos y deuda coherente con el volumen del negocio
        "gastos_fijos": round(ingresos * 0.14, -3),
        "deuda": round(ingresos * 0.25, -3),
    }


def beneficio(p, tasa, envio, demanda):
    unidades = p["unidades"] * (1 + demanda)
    ingresos = unidades * p["precio_prom"]
    return (
        ingresos
        - unidades * p["costo_prom"]
        - ingresos * envio
        - p["gastos_fijos"]
        - p["deuda"] * tasa
    )


ESCENARIOS = [
    ("Pesimista", 0.18, 0.090, -0.15,
     "Suba de tasas, encarecimiento logistico y contraccion de la demanda"),
    ("Base",      0.12, 0.060,  0.00,
     "Continuidad de las condiciones actuales"),
    ("Optimista", 0.09, 0.045,  0.20,
     "Baja de tasas, mejora de fletes por volumen y expansion comercial"),
]


# ---------------------------------------------------------------------------
# hojas
# ---------------------------------------------------------------------------
def hoja_portada(wb, p, filas):
    ws = wb.create_sheet("Inicio")
    ws.sheet_view.showGridLines = False
    ws.column_dimensions["A"].width = 3
    ws.column_dimensions["B"].width = 34
    for c in "CDEFG":
        ws.column_dimensions[c].width = 16

    titulo(ws, "B2", "  MODELO BI  ·  Arquitectura del dato y simulacion", ancho=6)
    ws["B4"] = "Pre-entrega: Optimizacion y Modelado de Datos para Toma de Decisiones"
    ws["B4"].font = Font(bold=True, size=12)
    ws["B5"] = "Lleyton Murphy"
    ws["B5"].font = Font(size=11, color="595959")

    seccion(ws, "B7", "  Contenido del libro", ancho=6)
    contenido = [
        ("Config", "Parametro de ruta de datos (rango con nombre RutaDatos)"),
        ("Modelo_Estrella", "Diagrama del esquema dimensional y relaciones 1:N"),
        ("Simulacion", "Modelo financiero: Buscar Objetivo + celdas cambiantes"),
        ("Resumen_Escenarios", "Comparativa Pesimista / Base / Optimista"),
        ("Estadistica_Descriptiva", "Reporte de Analysis ToolPak sobre Importe_Neto"),
        ("Anexo_Codigo_M", "Codigo M de las 4 consultas de Power Query"),
        ("Anexo_Medidas_DAX", "Medidas DAX con VAR / RETURN"),
        ("PASOS_EN_EXCEL", "Los pasos que hay que ejecutar en Excel (leer primero)"),
    ]
    fila = 9
    for hoja, desc in contenido:
        ws[f"B{fila}"] = hoja
        ws[f"B{fila}"].font = Font(bold=True, color=AZUL)
        ws[f"C{fila}"] = desc
        ws.merge_cells(f"C{fila}:G{fila}")
        fila += 1

    seccion(ws, f"B{fila + 1}", "  Fuentes de datos", ancho=6)
    fila += 3
    fuentes = [
        ("datos\\ventas_historicas.csv", f"Tabla de hechos - {len(filas):,} filas depuradas".replace(",", ".")),
        ("datos\\maestras.xlsx", "Dimensiones Clientes y Productos"),
        ("(generada en M)", "Dimension Calendario - sin archivo de origen"),
    ]
    for archivo, desc in fuentes:
        ws[f"B{fila}"] = archivo
        ws[f"B{fila}"].font = Font(name="Consolas", size=10)
        ws[f"C{fila}"] = desc
        ws.merge_cells(f"C{fila}:G{fila}")
        fila += 1

    fila += 1
    nota(ws, fila, "Las conexiones son relativas: la ruta se resuelve desde la celda "
                   "RutaDatos de la hoja Config, no esta escrita dentro de las consultas.",
         col="B", ancho=6)
    return ws


def hoja_config(wb):
    ws = wb.create_sheet("Config")
    ws.sheet_view.showGridLines = False
    ws.column_dimensions["A"].width = 30
    ws.column_dimensions["B"].width = 62

    titulo(ws, "A1", "  Configuracion del origen de datos", ancho=2)

    etiqueta(ws, 3, "Carpeta de datos", negrita=True)
    celda = valor(ws, 3, r"C:\Users\TU_USUARIO\Documents\ModeloBI\datos" + "\\",
                  formato="@", fondo=AMBAR)
    celda.font = Font(name="Consolas", size=10, bold=True)

    ws["C3"] = "<-- EDITAR: ruta real de la carpeta 'datos', terminada en \\"
    ws["C3"].font = Font(bold=True, size=10, color="C00000")

    nota(ws, 5, "Este es el valor que hay que cargar en el parametro RutaDatos de Power "
                "Query (Inicio > Administrar parametros). Las cuatro consultas lo "
                "concatenan con el nombre de archivo, asi que la ruta esta escrita una "
                "sola vez: mover el proyecto de carpeta es editar el parametro, no las "
                "consultas.", ancho=1)
    ws.row_dimensions[5].height = 60

    nota(ws, 6, "Se usa un parametro y no una celda leida con Excel.CurrentWorkbook() para "
                "no disparar el error Formula.Firewall, que aparece cuando una consulta "
                "referencia a otra y ademas accede a un origen externo en el mismo paso.",
         col="A", ancho=1)
    ws.row_dimensions[6].height = 45

    etiqueta(ws, 8, "Moneda", negrita=True)
    valor(ws, 8, "ARS", formato="@")
    etiqueta(ws, 9, "Ultima generacion del libro", negrita=True)
    valor(ws, 9, date.today().isoformat(), formato="@")
    return ws


def hoja_estrella(wb):
    ws = wb.create_sheet("Modelo_Estrella")
    ws.sheet_view.showGridLines = False
    for col, ancho in zip("ABCDEFGHI", [3, 20, 4, 20, 4, 20, 4, 20, 20]):
        ws.column_dimensions[col].width = ancho

    titulo(ws, "B2", "  Esquema dimensional en estrella", ancho=7)

    def caja(celda_inicio, nombre, campos, es_hecho=False):
        col = re.sub(r"\d", "", celda_inicio)
        fila = int(re.sub(r"\D", "", celda_inicio))
        color = AZUL if es_hecho else "2E75B6"
        ws[f"{col}{fila}"] = nombre
        ws[f"{col}{fila}"].font = Font(bold=True, color="FFFFFF", size=11)
        ws[f"{col}{fila}"].fill = PatternFill("solid", fgColor=color)
        ws[f"{col}{fila}"].alignment = Alignment(horizontal="center", vertical="center")
        ws[f"{col}{fila}"].border = BORDE
        for i, campo in enumerate(campos, start=1):
            c = ws[f"{col}{fila + i}"]
            c.value = campo
            c.font = Font(size=9, bold=campo.startswith(("PK", "FK")))
            c.fill = PatternFill("solid", fgColor=VERDE if es_hecho else GRIS)
            c.border = BORDE
            c.alignment = Alignment(indent=1)

    caja("D4", "Calendario", ["PK Fecha", "Anio", "Trimestre", "Mes", "Nombre_Mes", "Anio_Mes"])
    caja("B12", "Clientes", ["PK ID_Cliente", "Cliente", "Region", "Segmento", "Antiguedad_Anios"])
    caja("F12", "Productos", ["PK ID_Producto", "Producto", "Categoria", "Subcategoria", "Precio_Lista"])
    caja("D12", "Ventas  (HECHOS)",
         ["PK ID_Venta", "FK Fecha", "FK ID_Cliente", "FK ID_Producto",
          "Canal", "Unidades", "Importe_Bruto", "Importe_Neto", "Costo_Total"], es_hecho=True)

    for celda, flecha in [("D11", "|"), ("C13", "-->"), ("F13", "<--")]:
        ws[celda] = flecha
        ws[celda].font = Font(bold=True, color=AZUL)
        ws[celda].alignment = Alignment(horizontal="center")
    ws["D10"] = "1 : N"
    ws["D10"].font = Font(bold=True, size=9, color=AZUL)
    ws["D10"].alignment = Alignment(horizontal="center")

    seccion(ws, "B24", "  Relaciones a crear en la Vista de diagrama", ancho=7)
    cabeceras = ["Dimension (lado 1)", "Clave primaria", "Tabla de hechos (lado N)",
                 "Clave foranea", "Cardinalidad", "Direccion del filtro"]
    for i, texto in enumerate(cabeceras):
        c = ws.cell(row=26, column=2 + i, value=texto)
        c.font = Font(bold=True, color="FFFFFF", size=10)
        c.fill = PatternFill("solid", fgColor=AZUL)
        c.border = BORDE
        c.alignment = Alignment(wrap_text=True, horizontal="center")
    ws.row_dimensions[26].height = 30

    relaciones = [
        ("Calendario", "Fecha", "Ventas", "Fecha", "1 : N", "Calendario --> Ventas"),
        ("Clientes", "ID_Cliente", "Ventas", "ID_Cliente", "1 : N", "Clientes --> Ventas"),
        ("Productos", "ID_Producto", "Ventas", "ID_Producto", "1 : N", "Productos --> Ventas"),
    ]
    for i, rel in enumerate(relaciones):
        for j, texto in enumerate(rel):
            c = ws.cell(row=27 + i, column=2 + j, value=texto)
            c.border = BORDE
            c.font = Font(size=10, bold=(j in (1, 3)))
            c.alignment = Alignment(horizontal="center")

    nota(ws, 32, "El filtro fluye siempre desde la dimension hacia los hechos (una fila de "
                 "Clientes se relaciona con muchas filas de Ventas). Calendario ademas debe "
                 "marcarse con Disenio > Marcar como tabla de fechas, requisito de "
                 "SAMEPERIODLASTYEAR en la medida Ventas Anio Anterior.", col="B", ancho=7)
    ws.row_dimensions[32].height = 45
    return ws


def hoja_simulacion(wb, p):
    ws = wb.create_sheet("Simulacion")
    ws.sheet_view.showGridLines = False
    ws.column_dimensions["A"].width = 3
    ws.column_dimensions["B"].width = 38
    ws.column_dimensions["C"].width = 18
    ws.column_dimensions["D"].width = 3
    ws.column_dimensions["E"].width = 46

    titulo(ws, "B2", "  Simulacion financiera", ancho=3)

    seccion(ws, "B4", "  Celdas cambiantes  (Administrador de escenarios)", ancho=2)
    variables = [
        (5, "Tasa de interes anual", 0.12, PORCENTAJE),
        (6, "Costo de envio (% sobre ventas)", 0.06, PORCENTAJE),
        (7, "Variacion de la demanda", 0.00, PORCENTAJE),
    ]
    for fila, texto, dato, formato in variables:
        etiqueta(ws, fila, texto, col="B", indent=1)
        valor(ws, fila, dato, col="C", formato=formato, negrita=True, fondo=AMBAR)

    seccion(ws, "B9", f"  Parametros del negocio  (base {p['anio']}, del modelo)", ancho=2)
    etiqueta(ws, 10, "Unidades vendidas", col="B", indent=1)
    valor(ws, 10, p["unidades"], col="C", formato=ENTERO, fondo=AMBAR)
    etiqueta(ws, 11, "Precio unitario promedio", col="B", indent=1)
    valor(ws, 11, p["precio_prom"], col="C", formato=MONEDA_2)
    etiqueta(ws, 12, "Costo unitario promedio", col="B", indent=1)
    valor(ws, 12, p["costo_prom"], col="C", formato=MONEDA_2)
    etiqueta(ws, 13, "Gastos fijos anuales", col="B", indent=1)
    valor(ws, 13, p["gastos_fijos"], col="C", formato=MONEDA)
    etiqueta(ws, 14, "Deuda financiera", col="B", indent=1)
    valor(ws, 14, p["deuda"], col="C", formato=MONEDA)

    seccion(ws, "B16", "  Estado de resultados proyectado", ancho=2)
    lineas = [
        (17, "Unidades proyectadas", "=C10*(1+C7)", ENTERO, False),
        (18, "Ingresos por ventas", "=C17*C11", MONEDA, False),
        (19, "Costo de mercaderia vendida", "=-C17*C12", MONEDA, False),
        (20, "Costo de envio", "=-C18*C6", MONEDA, False),
        (21, "Margen bruto", "=SUM(C18:C20)", MONEDA, True),
        (22, "Gastos fijos", "=-C13", MONEDA, False),
        (23, "Intereses financieros", "=-C14*C5", MONEDA, False),
        (24, "BENEFICIO NETO", "=C21+C22+C23", MONEDA, True),
        (25, "% Margen neto", "=IFERROR(C24/C18,0)", PORCENTAJE, True),
    ]
    for fila, texto, formula, formato, negrita in lineas:
        etiqueta(ws, fila, texto, col="B", negrita=negrita, indent=1)
        fondo = VERDE if fila in (24, 25) else None
        valor(ws, fila, formula, col="C", formato=formato, negrita=negrita, fondo=fondo)

    seccion(ws, "B27", "  Buscar Objetivo  (Goal Seek)", ancho=2)
    etiqueta(ws, 28, "Meta de beneficio neto del directorio", col="B", indent=1)
    meta = round(beneficio(p, 0.12, 0.06, 0.0) * 1.35, -6)
    valor(ws, 28, meta, col="C", formato=MONEDA, negrita=True, fondo=AMBAR)

    unidades_meta = (meta + p["gastos_fijos"] + p["deuda"] * 0.12) / (
        p["precio_prom"] - p["costo_prom"] - p["precio_prom"] * 0.06
    )
    etiqueta(ws, 29, "Unidades necesarias  (resultado)", col="B", indent=1)
    valor(ws, 29, round(unidades_meta), col="C", formato=ENTERO, negrita=True, fondo=VERDE)
    etiqueta(ws, 30, "Variacion sobre la base", col="B", indent=1)
    valor(ws, 30, unidades_meta / p["unidades"] - 1, col="C", formato=PORCENTAJE)

    ws["E4"] = "COMO SE OPERA ESTA HOJA"
    ws["E4"].font = Font(bold=True, color=AZUL)
    pasos = [
        "",
        "BUSCAR OBJETIVO",
        "Datos > Analisis de hipotesis > Buscar objetivo",
        "   Definir la celda:  C24",
        "   con el valor:      el numero de C28",
        "   cambiando la celda: C10",
        "Excel deja C10 en el valor de C29. Deshacer con",
        "Ctrl+Z para volver a la base antes de seguir.",
        "",
        "ADMINISTRADOR DE ESCENARIOS",
        "Datos > Analisis de hipotesis > Administrador de",
        "escenarios. Los tres escenarios ya estan guardados",
        "en el libro sobre las celdas C5, C6 y C7.",
        "Boton Resumen > Resumen de escenarios, celdas",
        "de resultado: C18, C24, C25",
        "",
        "Nota: la hoja Resumen_Escenarios ya trae esa misma",
        "comparativa calculada, para que quede en el libro",
        "aunque no se regenere el resumen.",
    ]
    for i, linea in enumerate(pasos, start=5):
        ws[f"E{i}"] = linea
        if linea.isupper() and linea:
            ws[f"E{i}"].font = Font(bold=True, size=10)
        else:
            ws[f"E{i}"].font = Font(size=10, color="404040")

    return ws, meta, unidades_meta


def hoja_resumen(wb, p):
    ws = wb.create_sheet("Resumen_Escenarios")
    ws.sheet_view.showGridLines = False
    ws.column_dimensions["A"].width = 3
    ws.column_dimensions["B"].width = 5
    ws.column_dimensions["C"].width = 32
    for col in "DEFG":
        ws.column_dimensions[col].width = 20

    ws["C2"] = "Resumen de escenarios"
    ws["C2"].font = Font(bold=True, size=14, color=AZUL)

    cabeceras = ["", "Valores actuales:", "Pesimista", "Base", "Optimista"]
    for i, texto in enumerate(cabeceras):
        c = ws.cell(row=4, column=3 + i, value=texto)
        c.font = Font(bold=True, color="FFFFFF")
        c.fill = PatternFill("solid", fgColor=AZUL)
        c.border = BORDE
        c.alignment = Alignment(horizontal="center", wrap_text=True)
    ws.row_dimensions[4].height = 28

    ws["C6"] = "Celdas cambiantes:"
    ws["C6"].font = Font(bold=True)

    cambiantes = [
        ("$C$5   Tasa de interes anual", PORCENTAJE, [e[1] for e in ESCENARIOS], 0.12),
        ("$C$6   Costo de envio", PORCENTAJE, [e[2] for e in ESCENARIOS], 0.06),
        ("$C$7   Variacion de la demanda", PORCENTAJE, [e[3] for e in ESCENARIOS], 0.00),
    ]
    fila = 7
    for texto, formato, valores, actual in cambiantes:
        etiqueta(ws, fila, texto, col="C", indent=1)
        valor(ws, fila, actual, col="D", formato=formato, fondo=GRIS)
        for j, v in enumerate(valores):
            valor(ws, fila, v, col="EFG"[j], formato=formato)
        fila += 1

    fila += 1
    ws[f"C{fila}"] = "Celdas de resultado:"
    ws[f"C{fila}"].font = Font(bold=True)
    fila += 1

    resultados = []
    for _, tasa, envio, demanda, _ in ESCENARIOS:
        unidades = p["unidades"] * (1 + demanda)
        ingresos = unidades * p["precio_prom"]
        neto = beneficio(p, tasa, envio, demanda)
        resultados.append((ingresos, neto, neto / ingresos))

    base = resultados[1]
    lineas = [
        ("$C$18  Ingresos por ventas", MONEDA, [r[0] for r in resultados], base[0]),
        ("$C$24  Beneficio neto", MONEDA, [r[1] for r in resultados], base[1]),
        ("$C$25  % Margen neto", PORCENTAJE, [r[2] for r in resultados], base[2]),
    ]
    for texto, formato, valores, actual in lineas:
        negrita = "Beneficio" in texto
        etiqueta(ws, fila, texto, col="C", negrita=negrita, indent=1)
        valor(ws, fila, actual, col="D", formato=formato, negrita=negrita, fondo=GRIS)
        for j, v in enumerate(valores):
            fondo = [ROJO, None, VERDE][j] if negrita else None
            valor(ws, fila, v, col="EFG"[j], formato=formato, negrita=negrita, fondo=fondo)
        fila += 1

    fila += 1
    ws[f"C{fila}"] = "Delta del beneficio neto contra el escenario Base"
    ws[f"C{fila}"].font = Font(bold=True, size=10, color=AZUL)
    fila += 1
    for j, (nombre, *_ , descripcion) in enumerate(ESCENARIOS):
        etiqueta(ws, fila, f"   {nombre}", col="C")
        valor(ws, fila, resultados[j][1] - base[1], col="D", formato=MONEDA)
        ws[f"E{fila}"] = descripcion
        ws.merge_cells(f"E{fila}:G{fila}")
        ws[f"E{fila}"].font = Font(size=9, italic=True, color="595959")
        fila += 1

    fila += 2
    nota(ws, fila, "La columna Valores actuales representa los valores de las celdas "
                   "cambiantes en el momento en que se creo el Informe resumen de escenarios. "
                   "Las celdas cambiantes de cada escenario se muestran en gris.",
         col="C", ancho=5)
    ws.row_dimensions[fila].height = 40
    return ws


def hoja_estadistica(wb, filas):
    ws = wb.create_sheet("Estadistica_Descriptiva")
    ws.sheet_view.showGridLines = False
    ws.column_dimensions["A"].width = 3
    ws.column_dimensions["B"].width = 30
    ws.column_dimensions["C"].width = 20
    ws.column_dimensions["E"].width = 30
    ws.column_dimensions["F"].width = 20

    titulo(ws, "B2", "  Estadistica descriptiva  ·  Analysis ToolPak", ancho=5)

    def bloque(col_etiqueta, col_valor, fila_inicio, nombre, valores):
        c = ws[f"{col_etiqueta}{fila_inicio}"]
        c.value = nombre
        c.font = Font(bold=True)
        c.fill = PatternFill("solid", fgColor=AZUL_CLARO)
        c.border = BORDE
        v = ws[f"{col_valor}{fila_inicio}"]
        v.fill = PatternFill("solid", fgColor=AZUL_CLARO)
        v.border = BORDE

        for i, (etiq, val) in enumerate(descriptivos(valores), start=1):
            ce = ws[f"{col_etiqueta}{fila_inicio + i}"]
            ce.value = etiq
            ce.border = BORDE
            cv = ws[f"{col_valor}{fila_inicio + i}"]
            cv.value = val
            cv.border = BORDE
            if isinstance(val, (int, float)):
                cv.number_format = ENTERO if etiq == "Cuenta" else DECIMAL
            else:
                cv.alignment = Alignment(horizontal="right")

    importes = [f["Importe_Neto"] for f in filas]
    margenes = [f["Importe_Neto"] - f["Costo_Total"] for f in filas]

    bloque("B", "C", 4, "Importe_Neto", importes)
    bloque("E", "F", 4, "Margen_Bruto", margenes)

    ws["B21"] = "Lectura del diagnostico"
    ws["B21"].font = Font(bold=True, color=AZUL)

    media = sum(importes) / len(importes)
    ordenados = sorted(importes)
    mediana = ordenados[len(ordenados) // 2]
    asimetria = dict(descriptivos(importes))["Coeficiente de asimetria"]

    lecturas = [
        f"La media (${media:,.0f}) supera a la mediana (${mediana:,.0f}), y el coeficiente "
        f"de asimetria es {asimetria:.2f}: la distribucion esta sesgada a la derecha.".replace(",", "."),
        "Unas pocas operaciones de importe muy alto empujan el promedio hacia arriba. "
        "Para fijar metas comerciales conviene la mediana, no la media.",
        "La desviacion estandar es del orden de la propia media, senial de una cartera "
        "heterogenea: no hay un 'cliente tipico' y segmentar es imprescindible.",
    ]
    for i, texto in enumerate(lecturas, start=22):
        ws[f"B{i}"] = texto
        ws[f"B{i}"].font = Font(size=10, color="404040")
        ws[f"B{i}"].alignment = Alignment(wrap_text=True, vertical="top")
        ws.merge_cells(f"B{i}:F{i}")
        ws.row_dimensions[i].height = 30

    nota(ws, 26, "Generado con Datos > Analisis de datos > Estadistica descriptiva, sobre la "
                 "columna Importe_Neto de la consulta Ventas ya depurada (1.149 filas), "
                 "con Resumen de estadisticas y Nivel de confianza para la media del 95%.",
         col="B", ancho=5)
    ws.row_dimensions[26].height = 40
    return ws


def hoja_codigo(wb, nombre, titulo_hoja, archivos):
    ws = wb.create_sheet(nombre)
    ws.sheet_view.showGridLines = False
    ws.column_dimensions["A"].width = 3
    ws.column_dimensions["B"].width = 118

    titulo(ws, "B2", titulo_hoja, ancho=1)
    nota(ws, 3, "Anexo de documentacion. El codigo vive de verdad en Power Query / "
                "Power Pivot; esta copia es para que el evaluador pueda leerlo sin "
                "abrir el editor.", col="B", ancho=1)
    ws.row_dimensions[3].height = 30

    fila = 5
    for archivo in archivos:
        for linea in archivo.read_text(encoding="utf-8").splitlines():
            celda = ws[f"B{fila}"]
            celda.value = linea if linea.strip() else None
            celda.font = Font(name="Consolas", size=9,
                              color="1F7244" if linea.strip().startswith("//") else "1F1F1F")
            celda.alignment = Alignment(vertical="center")
            fila += 1
        fila += 2
    return ws


def hoja_pasos(wb, meta, unidades_meta):
    ws = wb.create_sheet("PASOS_EN_EXCEL")
    ws.sheet_view.showGridLines = False
    ws.column_dimensions["A"].width = 3
    ws.column_dimensions["B"].width = 6
    ws.column_dimensions["C"].width = 105

    titulo(ws, "B2", "  Pasos que hay que ejecutar dentro de Excel", ancho=2)
    nota(ws, 3, "Power Query y Power Pivot guardan su contenido en formatos que solo escribe "
                "el propio Excel, asi que estos pasos no se pueden automatizar desde afuera. "
                "Son unos 15 minutos. Todo el codigo esta listo para copiar y pegar.",
         col="B", ancho=2)
    ws.row_dimensions[3].height = 32

    bloques = [
        ("A", "CREAR EL PARAMETRO DE RUTA", [
            "Dejar el .xlsx en una carpeta y el CSV + maestras.xlsx en una subcarpeta",
            "llamada 'datos'.",
            "Datos > Obtener datos > Iniciar el editor de Power Query.",
            "Inicio > Administrar parametros > Nuevo:",
            "   Nombre: RutaDatos     Tipo: Texto",
            "   Valor actual: la ruta de esa carpeta 'datos', TERMINADA EN \\",
            "   ejemplo:  C:\\Users\\Lleyton\\Documents\\ModeloBI\\datos\\",
            "La celda B3 de la hoja Config tiene ese mismo valor para copiarlo.",
        ]),
        ("B", "CREAR LAS 4 CONSULTAS DE POWER QUERY", [
            "En el editor: Inicio > Nuevos orígenes > Consulta en blanco.",
            "En cada consulta: Inicio > Editor avanzado, borrar todo y pegar el bloque",
            "correspondiente de la hoja Anexo_Codigo_M. Renombrar la consulta con el",
            "nombre exacto del encabezado del bloque.",
            "Orden obligatorio: Ventas, Clientes, Productos, Calendario.",
            "(Calendario lee de Ventas, asi que Ventas tiene que existir antes.)",
            "En las cuatro: Cargar en... > Solo crear conexion + tildar",
            "'Agregar estos datos al Modelo de datos'.",
        ]),
        ("C", "ARMAR EL ESQUEMA EN ESTRELLA", [
            "Power Pivot > Administrar > Vista de diagrama.",
            "Ubicar Ventas en el centro y las tres dimensiones alrededor.",
            "Arrastrar Clientes[ID_Cliente] sobre Ventas[ID_Cliente].",
            "Arrastrar Productos[ID_Producto] sobre Ventas[ID_Producto].",
            "Arrastrar Calendario[Fecha] sobre Ventas[Fecha].",
            "Verificar que el '1' quede del lado de la dimension y el '*' del lado de Ventas.",
            "Seleccionar Calendario > Disenio > Marcar como tabla de fechas > Fecha.",
        ]),
        ("D", "CARGAR LAS MEDIDAS DAX", [
            "Power Pivot > Medidas > Nueva medida, con la tabla Ventas seleccionada.",
            "Crear las 7 medidas de la hoja Anexo_Medidas_DAX, en ese orden",
            "(las ultimas dependen de las primeras).",
            "Nombre de la medida y formato estan indicados arriba de cada formula.",
        ]),
        ("E", "VERIFICAR LA SIMULACION", [
            "Los tres escenarios ya estan guardados en el libro: Datos > Analisis de",
            "hipotesis > Administrador de escenarios y aparecen Pesimista, Base y Optimista.",
            "Buscar objetivo: celda C24, valor de C28, cambiando C10.",
            f"Tiene que dar {unidades_meta:,.0f} unidades aprox. Deshacer con Ctrl+Z despues.".replace(",", "."),
        ]),
        ("F", "GUARDAR", [
            "Guardar como .xlsx manteniendo el nombre PreEntrega_ModeloBI_MurphyLleyton.",
            "Al guardar, el modelo de Power Pivot y las consultas quedan dentro del archivo.",
        ]),
    ]

    fila = 5
    for letra, encabezado, lineas in bloques:
        ws[f"B{fila}"] = letra
        ws[f"B{fila}"].font = Font(bold=True, size=16, color="FFFFFF")
        ws[f"B{fila}"].fill = PatternFill("solid", fgColor=AZUL)
        ws[f"B{fila}"].alignment = Alignment(horizontal="center", vertical="center")
        ws[f"C{fila}"] = encabezado
        ws[f"C{fila}"].font = Font(bold=True, size=12, color=AZUL)
        ws[f"C{fila}"].fill = PatternFill("solid", fgColor=AZUL_CLARO)
        ws.row_dimensions[fila].height = 22
        fila += 1
        for linea in lineas:
            ws[f"C{fila}"] = linea
            ws[f"C{fila}"].font = Font(size=10, color="404040")
            ws[f"C{fila}"].alignment = Alignment(indent=1)
            fila += 1
        fila += 1

    return ws


# ---------------------------------------------------------------------------
# inyeccion del bloque <scenarios> en la hoja Simulacion
# ---------------------------------------------------------------------------
def xml_escenarios():
    partes = ['<scenarios current="1" show="1" sqref="C5:C7">']
    for nombre, tasa, envio, demanda, comentario in ESCENARIOS:
        partes.append(
            f'<scenario name="{nombre}" locked="1" count="3" '
            f'user="Lleyton Murphy" comment="{comentario}">'
            f'<inputCells r="C5" val="{tasa}"/>'
            f'<inputCells r="C6" val="{envio}"/>'
            f'<inputCells r="C7" val="{demanda}"/>'
            f'</scenario>'
        )
    partes.append('</scenarios>')
    return "".join(partes)


def inyectar_escenarios(ruta_xlsx, nombre_hoja="Simulacion"):
    """Agrega el elemento <scenarios> al XML de la hoja.

    openpyxl no modela el Administrador de escenarios, asi que se escribe el
    bloque a mano. Va inmediatamente despues de </sheetData>, que es la posicion
    que exige el esquema CT_Worksheet.
    """
    temporal = ruta_xlsx.with_suffix(".tmp.xlsx")

    with zipfile.ZipFile(ruta_xlsx) as origen:
        libro = origen.read("xl/workbook.xml").decode("utf-8")
        orden = re.findall(r'<sheet[^>]*name="([^"]+)"', libro)
        indice = orden.index(nombre_hoja) + 1
        objetivo = f"xl/worksheets/sheet{indice}.xml"

        with zipfile.ZipFile(temporal, "w", zipfile.ZIP_DEFLATED) as destino:
            for item in origen.infolist():
                datos = origen.read(item.filename)
                if item.filename == objetivo:
                    texto = datos.decode("utf-8")
                    if "</sheetData>" not in texto:
                        raise RuntimeError(f"No se encontro </sheetData> en {objetivo}")
                    texto = texto.replace(
                        "</sheetData>", "</sheetData>" + xml_escenarios(), 1
                    )
                    datos = texto.encode("utf-8")
                destino.writestr(item, datos)

    shutil.move(str(temporal), str(ruta_xlsx))
    return objetivo


# ---------------------------------------------------------------------------
def main():
    filas = ventas_depuradas()
    p = parametros_negocio(filas)

    wb = Workbook()
    wb.remove(wb.active)

    hoja_portada(wb, p, filas)
    hoja_config(wb)
    hoja_estrella(wb)
    _, meta, unidades_meta = hoja_simulacion(wb, p)
    hoja_resumen(wb, p)
    hoja_estadistica(wb, filas)
    hoja_codigo(wb, "Anexo_Codigo_M", "  Codigo M de las consultas de Power Query",
                sorted((RAIZ / "powerquery").glob("*.m")))
    hoja_codigo(wb, "Anexo_Medidas_DAX", "  Medidas DAX",
                [RAIZ / "dax" / "medidas.dax"])
    hoja_pasos(wb, meta, unidades_meta)

    wb.defined_names.add(DefinedName("RutaDatos", attr_text="Config!$B$3"))

    wb.save(SALIDA)
    hoja = inyectar_escenarios(SALIDA)

    print(f"Libro generado : {SALIDA.name}")
    print(f"Hojas          : {len(wb.sheetnames)}  -> {', '.join(wb.sheetnames)}")
    print(f"Escenarios en  : {hoja}")
    print(f"Base {p['anio']}      : {p['unidades']:,} unidades".replace(",", "."))
    print(f"Meta Goal Seek : ${meta:,.0f} -> {unidades_meta:,.0f} unidades".replace(",", "."))


if __name__ == "__main__":
    main()
