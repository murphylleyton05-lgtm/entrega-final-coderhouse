"""Controles automaticos sobre el libro generado.

Comprueba lo que se puede comprobar sin Excel: integridad del paquete OPC,
XML bien formado, presencia y forma del bloque <scenarios>, el rango con
nombre, la coherencia entre la hoja de resumen y el modelo financiero, y que
la cadena de formulas del modelo financiero devuelva lo esperado.
"""
import re
import sys
import xml.etree.ElementTree as ET
import zipfile
from pathlib import Path

from openpyxl import load_workbook

sys.path.insert(0, str(Path(__file__).resolve().parent))
from construir_libro import ESCENARIOS, beneficio, parametros_negocio  # noqa: E402
from estadistica import descriptivos, ventas_depuradas  # noqa: E402

RAIZ = Path(__file__).resolve().parent.parent
LIBRO = RAIZ / "PreEntrega_ModeloBI_MurphyLleyton.xlsx"

NS = "{http://schemas.openxmlformats.org/spreadsheetml/2006/main}"

fallos = []
oks = []


def check(condicion, mensaje, detalle=""):
    (oks if condicion else fallos).append(mensaje + (f"  [{detalle}]" if detalle else ""))
    return condicion


def paquete_integro():
    with zipfile.ZipFile(LIBRO) as z:
        danado = z.testzip()
        check(danado is None, "El paquete ZIP esta integro", str(danado))
        partes = z.namelist()
        for nombre in partes:
            if nombre.endswith(".xml") or nombre.endswith(".rels"):
                try:
                    ET.fromstring(z.read(nombre))
                except ET.ParseError as exc:
                    check(False, f"XML valido en {nombre}", str(exc))
        check(True, f"Los {len([p for p in partes if p.endswith(('.xml', '.rels'))])} XML del paquete son validos")
        return partes


def bloque_escenarios():
    with zipfile.ZipFile(LIBRO) as z:
        libro = z.read("xl/workbook.xml").decode("utf-8")
        orden = re.findall(r'<sheet[^>]*name="([^"]+)"', libro)
        indice = orden.index("Simulacion") + 1
        hoja = z.read(f"xl/worksheets/sheet{indice}.xml").decode("utf-8")

    if not check("<scenarios" in hoja, "La hoja Simulacion contiene el bloque <scenarios>"):
        return

    raiz = ET.fromstring(hoja)
    escenarios = raiz.find(f"{NS}scenarios")
    hijos = escenarios.findall(f"{NS}scenario")

    check(len(hijos) == 3, "Hay exactamente 3 escenarios guardados", f"{len(hijos)}")

    nombres = [e.get("name") for e in hijos]
    check(nombres == ["Pesimista", "Base", "Optimista"],
          "Los escenarios se llaman Pesimista / Base / Optimista", str(nombres))

    for esperado, elemento in zip(ESCENARIOS, hijos):
        nombre, tasa, envio, demanda, _ = esperado
        celdas = elemento.findall(f"{NS}inputCells")
        refs = [c.get("r") for c in celdas]
        vals = [float(c.get("val")) for c in celdas]
        check(refs == ["C5", "C6", "C7"],
              f"{nombre}: apunta a las celdas C5/C6/C7", str(refs))
        check(vals == [tasa, envio, demanda],
              f"{nombre}: guarda los valores {tasa} / {envio} / {demanda}", str(vals))

    # El orden del esquema exige scenarios inmediatamente despues de sheetData
    posterior = hoja.split("</sheetData>", 1)[1]
    check(posterior.lstrip().startswith("<scenarios"),
          "El bloque <scenarios> esta en la posicion que exige el esquema")


def rango_con_nombre():
    wb = load_workbook(LIBRO)
    check("RutaDatos" in wb.defined_names, "Existe el rango con nombre RutaDatos")
    if "RutaDatos" in wb.defined_names:
        destino = wb.defined_names["RutaDatos"].attr_text
        check(destino == "Config!$B$3", "RutaDatos apunta a Config!$B$3", destino)
        ruta = wb["Config"]["B3"].value
        check(isinstance(ruta, str) and ruta.endswith("\\"),
              "La ruta de datos termina en barra invertida", repr(ruta))
        check(isinstance(ruta, str) and re.match(r"^[A-Z]:\\", ruta),
              "La ruta de datos tiene formato Windows", repr(ruta))

    esperadas = {"Inicio", "Config", "Modelo_Estrella", "Simulacion", "Resumen_Escenarios",
                 "Estadistica_Descriptiva", "Anexo_Codigo_M", "Anexo_Medidas_DAX",
                 "PASOS_EN_EXCEL"}
    check(esperadas.issubset(set(wb.sheetnames)),
          "Estan las 9 hojas previstas", str(sorted(set(wb.sheetnames) ^ esperadas)))
    return wb


def contenido_anexos(wb):
    texto_m = "\n".join(
        str(c.value) for fila in wb["Anexo_Codigo_M"].iter_rows() for c in fila if c.value
    )
    check("VentasFiltradas" in texto_m,
          "El codigo M incluye el paso renombrado VentasFiltradas")
    check(texto_m.count("//") >= 10,
          "El codigo M tiene comentarios con //", f"{texto_m.count('//')} comentarios")
    for consulta in ["RutaDatos", "Ventas", "Clientes", "Productos", "Calendario"]:
        check(f"Consulta: {consulta}" in texto_m or f"Consulta: {consulta} " in texto_m,
              f"Esta documentada la consulta {consulta}")
    check("Excel.CurrentWorkbook" in texto_m,
          "La ruta de origen se resuelve sin hardcodear")

    texto_dax = "\n".join(
        str(c.value) for fila in wb["Anexo_Medidas_DAX"].iter_rows() for c in fila if c.value
    )
    medidas_var = len(re.findall(r"^VAR ", texto_dax, re.MULTILINE))
    check(medidas_var >= 5, "Hay medidas DAX con VAR", f"{medidas_var} sentencias VAR")
    check(texto_dax.count("RETURN") >= 5, "Cada VAR tiene su RETURN",
          f"{texto_dax.count('RETURN')} RETURN")
    for medida in ["Total Ventas", "Margen Neto", "% Margen", "Ticket Promedio"]:
        check(medida in texto_dax, f"Esta la medida {medida}")


def resumen_coherente(wb, p):
    """El resumen de escenarios debe coincidir con lo que da el modelo."""
    ws = wb["Resumen_Escenarios"]
    fila_beneficio = None
    for fila in ws.iter_rows(min_col=3, max_col=3):
        if fila[0].value and "Beneficio neto" in str(fila[0].value):
            fila_beneficio = fila[0].row
            break

    if not check(fila_beneficio is not None, "El resumen tiene la fila de beneficio neto"):
        return

    for j, (nombre, tasa, envio, demanda, _) in enumerate(ESCENARIOS):
        celda = ws[f"{'EFG'[j]}{fila_beneficio}"].value
        esperado = beneficio(p, tasa, envio, demanda)
        check(abs(celda - esperado) < 0.01,
              f"Resumen: el beneficio de {nombre} coincide con el modelo",
              f"{celda:,.2f} vs {esperado:,.2f}")

    pes = ws[f"E{fila_beneficio}"].value
    opt = ws[f"G{fila_beneficio}"].value
    base = ws[f"F{fila_beneficio}"].value
    check(pes < base < opt,
          "Los escenarios estan ordenados: Pesimista < Base < Optimista",
          f"{pes:,.0f} / {base:,.0f} / {opt:,.0f}")


def estadistica_coherente(wb, filas):
    ws = wb["Estadistica_Descriptiva"]
    leidos = {}
    for fila in range(5, 19):
        etiqueta = ws[f"B{fila}"].value
        if etiqueta:
            leidos[etiqueta] = ws[f"C{fila}"].value

    calculados = dict(descriptivos([f["Importe_Neto"] for f in filas]))
    for etiqueta in ["Media", "Mediana", "Desviacion estandar", "Rango", "Minimo", "Maximo"]:
        check(etiqueta in leidos, f"El reporte incluye {etiqueta}")
        if etiqueta in leidos:
            check(abs(leidos[etiqueta] - calculados[etiqueta]) < 0.01,
                  f"{etiqueta} coincide con el calculo directo",
                  f"{leidos[etiqueta]:,.2f}")

    check(leidos.get("Cuenta") == len(filas),
          "La cuenta del reporte coincide con las filas depuradas",
          f"{leidos.get('Cuenta')} vs {len(filas)}")


def _evaluar(ws, referencia, visitados=None):
    """Evaluador minimo de las formulas de la hoja Simulacion.

    Resuelve referencias en cadena y soporta las funciones usadas (SUM sobre un
    rango de una columna e IFERROR). No pretende ser un motor de calculo: sirve
    para detectar una referencia mal escrita, que es el error que de verdad
    puede colarse al escribir formulas desde un script.
    """
    visitados = visitados or set()
    if referencia in visitados:
        raise RuntimeError(f"Referencia circular en {referencia}")
    visitados = visitados | {referencia}

    valor_celda = ws[referencia].value
    if not isinstance(valor_celda, str) or not valor_celda.startswith("="):
        return float(valor_celda or 0)

    expresion = valor_celda[1:]

    def sumar_rango(match):
        col, desde, hasta = match.group(1), int(match.group(2)), int(match.group(3))
        return "(" + "+".join(
            str(_evaluar(ws, f"{col}{f}", visitados)) for f in range(desde, hasta + 1)
        ) + ")"

    expresion = re.sub(r"SUM\(([A-Z]+)(\d+):[A-Z]+(\d+)\)", sumar_rango, expresion)
    expresion = re.sub(r"IFERROR\((.+),\s*[\d.]+\)$", r"\1", expresion)
    expresion = re.sub(
        r"\b([A-Z]+\d+)\b",
        lambda m: repr(_evaluar(ws, m.group(1), visitados)),
        expresion,
    )
    return float(eval(expresion, {"__builtins__": {}}, {}))


def formulas_consistentes(wb, p):
    """Evalua la cadena de formulas y la compara contra el modelo en Python.

    LibreOffice no esta operativo en este entorno (no abre ni un .xlsx de dos
    celdas), asi que la comprobacion se hace con un evaluador propio.
    """
    ws = wb["Simulacion"]

    for celda in ["C17", "C18", "C19", "C20", "C21", "C22", "C23", "C24", "C25"]:
        check(isinstance(ws[celda].value, str) and ws[celda].value.startswith("="),
              f"{celda} contiene una formula, no un valor fijo", repr(ws[celda].value))

    check(_evaluar(ws, "C17") == p["unidades"],
          "Las unidades proyectadas parten de la base", f"{_evaluar(ws, 'C17')}")

    esperado = beneficio(p, 0.12, 0.06, 0.0)
    obtenido = _evaluar(ws, "C24")
    check(abs(obtenido - esperado) < 0.01,
          "La cadena de formulas da el mismo beneficio neto que el modelo",
          f"{obtenido:,.2f} vs {esperado:,.2f}")

    margen = _evaluar(ws, "C25")
    check(0 < margen < 1, "El % de margen neto cae en un rango razonable", f"{margen:.4f}")

    # Sensibilidad: subir la tasa de interes tiene que bajar el beneficio
    original = ws["C5"].value
    ws["C5"] = 0.30
    peor = _evaluar(ws, "C24")
    ws["C5"] = original
    check(peor < obtenido,
          "Subir la tasa de interes reduce el beneficio (el modelo reacciona)",
          f"{peor:,.0f} < {obtenido:,.0f}")

    # Buscar objetivo: el resultado documentado tiene que cumplir la meta
    meta = ws["C28"].value
    unidades_meta = ws["C29"].value
    ws["C10"] = unidades_meta
    alcanzado = _evaluar(ws, "C24")
    ws["C10"] = p["unidades"]
    check(abs(alcanzado - meta) / meta < 0.001,
          "El resultado documentado de Buscar Objetivo alcanza la meta",
          f"{alcanzado:,.0f} vs meta {meta:,.0f}")


def main():
    if not LIBRO.exists():
        print(f"No existe {LIBRO.name}: ejecutar antes construir_libro.py")
        return 1

    filas = ventas_depuradas()
    p = parametros_negocio(filas)

    paquete_integro()
    bloque_escenarios()
    wb = rango_con_nombre()
    contenido_anexos(wb)
    resumen_coherente(wb, p)
    estadistica_coherente(wb, filas)
    formulas_consistentes(wb, p)

    print(f"\n{'=' * 68}")
    for mensaje in oks:
        print(f"  OK    {mensaje}")
    for mensaje in fallos:
        print(f"  FALLA {mensaje}")
    print(f"{'=' * 68}")
    print(f"  {len(oks)} controles pasados, {len(fallos)} fallidos")
    return 1 if fallos else 0


if __name__ == "__main__":
    raise SystemExit(main())
