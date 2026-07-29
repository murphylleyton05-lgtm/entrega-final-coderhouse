"""Genera las fuentes de datos del checkpoint: el CSV transaccional (tabla de
hechos) y el libro de tablas maestras (dimensiones Clientes y Productos).

La dimension Calendario NO se genera aca a proposito: se construye por completo
en lenguaje M para que la consulta muestre codigo escrito a mano.
"""
import csv
import random
from datetime import date, timedelta
from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill
from openpyxl.worksheet.table import Table, TableStyleInfo

RAIZ = Path(__file__).resolve().parent.parent
DATOS = RAIZ / "datos"

SEMILLA = 20260729
FILAS = 1200

REGIONES = ["Centro", "Norte", "Sur", "Litoral", "Patagonia"]
SEGMENTOS = ["Corporativo", "PyME", "Consumo Final"]

CATEGORIAS = {
    "Tecnologia": ["Notebooks", "Perifericos", "Redes"],
    "Mobiliario": ["Escritorios", "Sillas", "Almacenamiento"],
    "Insumos":    ["Papeleria", "Cartuchos", "Limpieza"],
}

NOMBRES = [
    "Aguirre", "Benitez", "Cabrera", "Dominguez", "Escobar", "Ferreyra",
    "Gimenez", "Herrera", "Ibarra", "Juarez", "Ledesma", "Medina",
    "Navarro", "Ocampo", "Peralta", "Quiroga", "Rios", "Sosa",
    "Toledo", "Urbina", "Vera", "Zarate",
]
FORMAS = ["S.A.", "S.R.L.", "y Asociados", "Group", "Distribuciones"]


def generar_clientes(rnd, cantidad=60):
    clientes = []
    for i in range(1, cantidad + 1):
        clientes.append({
            "ID_Cliente": f"C{i:04d}",
            "Cliente": f"{rnd.choice(NOMBRES)} {rnd.choice(FORMAS)}",
            "Region": rnd.choice(REGIONES),
            "Segmento": rnd.choice(SEGMENTOS),
            "Antiguedad_Anios": rnd.randint(1, 15),
        })
    return clientes


def generar_productos(rnd, cantidad=40):
    productos = []
    for i in range(1, cantidad + 1):
        categoria = rnd.choice(list(CATEGORIAS))
        subcategoria = rnd.choice(CATEGORIAS[categoria])
        costo = round(rnd.uniform(1800, 95000), 2)
        productos.append({
            "ID_Producto": f"P{i:04d}",
            "Producto": f"{subcategoria[:-1]} {rnd.choice(['Pro', 'Basic', 'Plus', 'Max', 'Eco'])} {rnd.randint(100, 999)}",
            "Categoria": categoria,
            "Subcategoria": subcategoria,
            "Costo_Unitario": costo,
            "Precio_Lista": round(costo * rnd.uniform(1.35, 1.95), 2),
        })
    return productos


def generar_ventas(rnd, clientes, productos, filas=FILAS):
    """Tabla de hechos. Se ensucia a proposito para que Power Query tenga
    trabajo real: nulos, importes con formato texto y fechas en dos formatos."""
    inicio = date(2023, 1, 1)
    fin = date(2025, 12, 31)
    dias = (fin - inicio).days

    ventas = []
    for i in range(1, filas + 1):
        producto = rnd.choice(productos)
        cliente = rnd.choice(clientes)
        fecha = inicio + timedelta(days=rnd.randint(0, dias))
        unidades = rnd.randint(1, 40)
        precio = round(producto["Precio_Lista"] * rnd.uniform(0.92, 1.08), 2)
        descuento = rnd.choice([0, 0, 0, 0.05, 0.10, 0.15])

        fila = {
            "ID_Venta": f"V{i:05d}",
            "Fecha": fecha.strftime("%d/%m/%Y"),
            "ID_Cliente": cliente["ID_Cliente"],
            "ID_Producto": producto["ID_Producto"],
            "Unidades": unidades,
            "Precio_Unitario": f"{precio:.2f}",
            "Descuento": descuento,
            "Costo_Unitario": f"{producto['Costo_Unitario']:.2f}",
            "Canal": rnd.choice(["Online", "Sucursal", "Mayorista", "online", "SUCURSAL"]),
            "Estado": "Confirmada",
        }

        # Suciedad controlada: ~4% de filas con problemas que el ETL debe resolver.
        ruido = rnd.random()
        if ruido < 0.015:
            fila["Unidades"] = ""              # nulo en metrica -> se descarta
        elif ruido < 0.025:
            fila["ID_Cliente"] = ""            # FK huerfana -> se descarta
        elif ruido < 0.032:
            fila["Estado"] = "Anulada"         # se filtra por regla de negocio
        elif ruido < 0.040:
            fila["Precio_Unitario"] = "n/d"    # no numerico -> se descarta

        ventas.append(fila)

    # Duplicados exactos para que Table.Distinct tenga sentido.
    for fila in rnd.sample(ventas, 18):
        ventas.append(dict(fila))

    rnd.shuffle(ventas)
    return ventas


def escribir_csv(ventas):
    DATOS.mkdir(parents=True, exist_ok=True)
    ruta = DATOS / "ventas_historicas.csv"
    columnas = list(ventas[0].keys())
    with ruta.open("w", newline="", encoding="utf-8") as fh:
        escritor = csv.DictWriter(fh, fieldnames=columnas, delimiter=";")
        escritor.writeheader()
        escritor.writerows(ventas)
    return ruta


def _hoja_tabla(wb, titulo, registros, nombre_tabla):
    ws = wb.create_sheet(titulo)
    columnas = list(registros[0].keys())
    ws.append(columnas)
    for registro in registros:
        ws.append([registro[c] for c in columnas])

    cabecera = Font(bold=True, color="FFFFFF")
    relleno = PatternFill("solid", fgColor="1F3864")
    for celda in ws[1]:
        celda.font = cabecera
        celda.fill = relleno

    ref = f"A1:{chr(64 + len(columnas))}{len(registros) + 1}"
    tabla = Table(displayName=nombre_tabla, ref=ref)
    tabla.tableStyleInfo = TableStyleInfo(
        name="TableStyleMedium2", showRowStripes=True
    )
    ws.add_table(tabla)

    for idx, columna in enumerate(columnas, start=1):
        ancho = max(len(columna), *(len(str(r[columna])) for r in registros)) + 3
        ws.column_dimensions[chr(64 + idx)].width = min(ancho, 40)
    return ws


def escribir_maestras(clientes, productos):
    wb = Workbook()
    wb.remove(wb.active)
    _hoja_tabla(wb, "Clientes", clientes, "tbl_Clientes")
    _hoja_tabla(wb, "Productos", productos, "tbl_Productos")
    ruta = DATOS / "maestras.xlsx"
    wb.save(ruta)
    return ruta


def main():
    rnd = random.Random(SEMILLA)
    clientes = generar_clientes(rnd)
    productos = generar_productos(rnd)
    ventas = generar_ventas(rnd, clientes, productos)

    ruta_csv = escribir_csv(ventas)
    ruta_xlsx = escribir_maestras(clientes, productos)

    print(f"CSV     : {ruta_csv.name}  ({len(ventas)} filas)")
    print(f"Maestras: {ruta_xlsx.name} ({len(clientes)} clientes, {len(productos)} productos)")
    return clientes, productos, ventas


if __name__ == "__main__":
    main()
